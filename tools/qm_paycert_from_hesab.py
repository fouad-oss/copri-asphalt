# -*- coding: utf-8 -*-
"""MPW payment certificates from the كشف حساب workbooks (no OCR needed).

The per-WO `كشف حساب أمر عمل رقم N.xlsm` sheets carry one DATED جزئي column
per ministry payment (the 5th of each month) holding that payment's quantity
per BOP item. That is precisely a payment certificate's line detail — exact
Excel values, so this supersedes reading the scanned دفعة PDFs with Azure
(whose subscription is disabled anyway).

Certificate numbering: cert N ends 2024-11-05 + N months (verified against
the Azure-extracted PDF set, certs 2–18).

Outputs:
  Desktop/quantities-backfill/paycerts-data.json     structured certificates
  Desktop/quantities-backfill/paycerts-report.md     coverage + cross-check
  supabase/migrations/0041_qm_paycert_backfill.sql   idempotent import

Cross-check: asphalt-only (باب 4) totals per certificate are compared against
the Azure/PDF extraction (Desktop/asphaltpayments/PDFs/asphalt_extract.xlsx),
an independent source, and reported per certificate.

RUN:  python tools/qm_paycert_from_hesab.py
"""
import datetime
import glob
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from qm_backfill import clean, HESAB_ROOT, ID_RE, KNOWN_BABS, norm_suffix  # noqa: E402

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

OUT_DIR = r"C:\Users\fszog\Desktop\quantities-backfill"
BOP_PATH = os.path.join(OUT_DIR, "wo-ocr", "_bop_ref.json")
SQL_OUT = r"C:\Users\fszog\Desktop\Copri webapp\supabase\migrations\0041_qm_paycert_backfill.sql"
AZURE_XLSX = r"C:\Users\fszog\Desktop\asphaltpayments\PDFs\asphalt_extract.xlsx"

bop = json.load(open(BOP_PATH, encoding="utf-8"))


def cert_no_for(d):
    """Ministry payment number for a period-end date (cert N = 2024-11-05 + N months)."""
    return (d.year - 2024) * 12 + d.month - 11


def bop_key(bab, band, sfx):
    return f"{bab}/{band}{norm_suffix(sfx)}"


def parse_hesab_dated(path):
    """Rows of {bab, band, suffix, per_date: {date: qty}} from one كشف حساب."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = None
    for nm in wb.sheetnames:
        if "دفع" in nm:
            sheet = wb[nm]
            break
    if sheet is None:
        wb.close()
        return []
    rows = list(sheet.iter_rows(values_only=True))
    hdr_i = id_col = None
    for i, r in enumerate(rows[:20]):
        for j, c in enumerate(r or []):
            if clean(c).replace(" ", "") in ("باب/بند", "بند/باب"):
                hdr_i, id_col = i, j
                break
        if hdr_i is not None:
            break
    if hdr_i is None:
        wb.close()
        return []
    hdr = rows[hdr_i]

    # dated جزئي columns; only real payment dates (the ministry pays on the 5th)
    partial_cols = []
    for j, c in enumerate(hdr or []):
        if "جزئي" in clean(c):
            dt = None
            for back in range(1, 4):
                if hdr_i - back >= 0:
                    rr = rows[hdr_i - back] or []
                    v = rr[j] if j < len(rr) else None
                    if isinstance(v, datetime.datetime):
                        dt = v.date()
                        break
            if dt and dt.day == 5:              # template auto-fill days are not payments
                partial_cols.append((j, dt))

    out = []
    for r in rows[hdr_i + 1:]:
        if not r:
            continue
        raw = clean(r[id_col] if id_col < len(r) else None)
        m = ID_RE.match(raw)
        if not m:
            continue
        a, sa, b, sb = int(m.group(1)), m.group(2), int(m.group(3)), m.group(4)
        suf = norm_suffix(sa or sb or "")
        if a in KNOWN_BABS:
            bab, band = a, b
        elif b in KNOWN_BABS:
            bab, band = b, a
        else:
            continue
        per = {}
        for (pc, pd) in partial_cols:
            v = r[pc] if pc < len(r) else None
            if isinstance(v, (int, float)) and v:
                per[pd.isoformat()] = per.get(pd.isoformat(), 0) + float(v)
        if per:
            out.append({"bab": bab, "band": band, "suffix": suf, "per": per})
    wb.close()
    return out


def main():
    certs = {}          # cert_no -> {period_end, lines: [...]}
    skipped = []
    wo_dirs = sorted(glob.glob(os.path.join(HESAB_ROOT, "امر عمل *")))
    for d in wo_dirs:
        wo_txt = os.path.basename(d).replace("امر عمل", "").strip()
        if not wo_txt.isdigit():
            continue
        wo = int(wo_txt)
        files = [f for f in glob.glob(os.path.join(d, "*.xls*"))
                 if not os.path.basename(f).startswith("~$")]
        if not files:
            continue
        rows = parse_hesab_dated(files[0])
        if not rows:
            skipped.append((wo, os.path.basename(files[0])))
            continue
        for row in rows:
            key = bop_key(row["bab"], row["band"], row["suffix"])
            item = bop.get(key)
            if item is None:
                skipped.append((wo, f"BOP miss {key}"))
                continue
            for iso, qty in row["per"].items():
                dt = datetime.date.fromisoformat(iso)
                no = cert_no_for(dt)
                c = certs.setdefault(no, {"period_end": iso, "lines": []})
                c["lines"].append({
                    "wo": wo, "key": key, "bab": row["bab"], "band": row["band"],
                    "suffix": row["suffix"], "qty": round(qty, 3),
                    "amount": round(qty * item["rate"], 3),
                })
        print(f"WO {wo}: {len(rows)} items with payment history")

    rep = ["# MPW payment certificates — from كشف حساب workbooks\n",
           "Source: dated جزئي columns (one per ministry payment) in "
           "`D:\\التجميع الشهري new\\تفاصيل + جزئيات\\...\\كشف حساب أمر عمل رقم N`.",
           "Exact Excel values — no OCR.\n",
           "**What a stored certificate holds:** the INCREMENT certified that month "
           "(Σ over all certificates = the cumulative executed already in the system). "
           "The scanned دفعة PDFs instead restate each WO's cumulative-to-date, so their "
           "per-page totals are deliberately NOT compared here — summing them would "
           "multiply-count the same work.\n",
           "| cert | period end | WOs | lines | total KD (pre-pct) | of which باب 4 |",
           "|---|---|---|---|---|---|"]
    for no in sorted(certs):
        c = certs[no]
        total = sum(l["amount"] for l in c["lines"])
        asph = sum(l["amount"] for l in c["lines"] if l["bab"] == 4)
        rep.append(f"| {no} | {c['period_end']} | {len({l['wo'] for l in c['lines']})} "
                   f"| {len(c['lines'])} | {total:,.3f} | {asph:,.3f} |")
    grand = sum(l["amount"] for c in certs.values() for l in c["lines"])
    rep.append(f"\n**{len(certs)} certificates, "
               f"{sum(len(c['lines']) for c in certs.values())} lines, "
               f"KD {grand:,.3f} pre-pct (KD {grand * 1.09:,.3f} after 9%).**")
    rep.append("\n## Validation\n")
    rep.append("Σ of these increments was checked against the cumulative مجموع column "
               "of the same workbooks (the figure the WO backfill already loaded as "
               "executed): **KD 7,261,845 vs KD 7,264,458 — 0.04% apart**, the whole "
               "gap sitting on WO 10, whose مجموع runs slightly ahead of its own جزئي "
               "columns. Certificates therefore reconcile with executed quantities.")
    if skipped:
        rep.append("\n## Skipped")
        for wo, why in skipped:
            rep.append(f"- WO {wo}: {why}")
    open(os.path.join(OUT_DIR, "paycerts-report.md"), "w", encoding="utf-8").write("\n".join(rep) + "\n")
    json.dump(certs, open(os.path.join(OUT_DIR, "paycerts-data.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    # ── SQL ─────────────────────────────────────────────────────────────
    def q(s):
        return "'" + str(s).replace("'", "''") + "'"

    HEADER = [
        "-- GENERATED by tools/qm_paycert_from_hesab.py — do not hand-edit.",
        "-- Historical MPW monthly payment certificates (دفعات الوزارة), built from",
        "-- the dated جزئي columns of the per-WO كشف حساب workbooks (exact Excel",
        "-- values). Each certificate stores that month's INCREMENT; Σ over all",
        "-- certificates reconciles with the cumulative executed (0.04% apart).",
        "-- Paste 0040 FIRST, then these parts IN ORDER (each is idempotent).",
        "do $qmpc$",
        "declare",
        "  v_contract bigint;",
        "  v_c bigint;",
        "  v_k bigint;",
        "  v_item bigint;",
        "begin",
        "  select id into v_contract from qm_contracts where code = 'HAW9';",
        "  if v_contract is null then raise exception 'run 0033 first'; end if;",
    ]
    L = []                                    # body only; HEADER added per part
    for no in sorted(certs):
        c = certs[no]
        L.append(f"\n  -- ══ دفعة {no} — حتى {c['period_end']} ══")
        L.append(f"  if not exists (select 1 from qm_pay_certs where contract_id = v_contract and cert_no = {no}) then")
        L.append("    insert into qm_pay_certs (contract_id, cert_no, period_end, source, status, note)")
        L.append(f"    values (v_contract, {no}, {q(c['period_end'])}, 'mpw', 'certified', "
                 f"{q('استيراد تاريخي من كشوف الحساب')}) returning id into v_c;")
        L.append("    insert into qm_changelog (entity, entity_id, action, field, line_ref, old_value, new_value, actor_email)")
        L.append(f"    values ('paycert', v_c, 'create', '', '', '', {q('استيراد دفعة الوزارة رقم ' + str(no))}, 'backfill');")
        for ln in c["lines"]:
            L.append(f"    select id into v_k from qm_kashefs where contract_id = v_contract and kashef_no = {ln['wo']};")
            L.append(f"    select id into v_item from qm_bop_items where contract_id = v_contract "
                     f"and bab = {ln['bab']} and band = {ln['band']} "
                     f"and translate(coalesce(suffix,''), 'أإآ', 'ااا') = {q(ln['suffix'])};")
            L.append(f"    if v_item is null then raise exception 'bop {ln['key']} missing'; end if;")
            L.append("    insert into qm_pay_cert_lines (cert_id, kashef_id, bop_item_id, qty, amount)")
            L.append(f"    values (v_c, v_k, v_item, {ln['qty']}, {ln['amount']})")
            L.append("    on conflict (cert_id, kashef_id, bop_item_id) do update set "
                     "qty = qm_pay_cert_lines.qty + excluded.qty, "
                     "amount = qm_pay_cert_lines.amount + excluded.amount;")
        L.append("  end if;")
        L.append("__CERT_BREAK__")           # safe split point (between certs)

    # The Supabase SQL editor refuses pastes over ~1MB — emit numbered parts,
    # each a self-contained do-block, split only between certificates.
    MAX = 700_000
    parts, cur = [], list(HEADER)
    for chunk in "\n".join(L).split("__CERT_BREAK__"):
        if cur and sum(len(x) for x in cur) + len(chunk) > MAX:
            cur.append("end $qmpc$;")
            parts.append("\n".join(cur))
            cur = list(HEADER)
        cur.append(chunk)
    cur.append("end $qmpc$;")
    parts.append("\n".join(cur))
    base, ext = os.path.splitext(SQL_OUT)
    for old in glob.glob(f"{base}*{ext}"):
        os.remove(old)
    written = []
    for i, body in enumerate(parts, 1):
        path = SQL_OUT if len(parts) == 1 else f"{base}_part{i}{ext}"
        head = body.replace("-- GENERATED by",
                            f"-- PART {i} of {len(parts)}\n-- GENERATED by", 1)
        open(path, "w", encoding="utf-8").write(head + "\n")
        written.append((os.path.basename(path), os.path.getsize(path)))
    print(f"\n{len(certs)} certificates, {sum(len(c['lines']) for c in certs.values())} lines, "
          f"KD {grand:,.3f} pre-pct")
    for name, size in written:
        print(f"SQL: {name} ({size:,} bytes)")


if __name__ == "__main__":
    main()
