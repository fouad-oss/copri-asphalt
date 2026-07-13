# -*- coding: utf-8 -*-
"""NUBA vs SpectroNova journal reconciliation -> exceptions-only Excel report.

Join key: SpectroNova JV `ReferenceNumber` carries the NUBA manual voucher no
(6-digit `nnnnnn` or `Cnnnnnnnn`).  Two traps drive the design:

1. NUBA voucher numbering RESTARTS PER VOUCHER TYPE, so one number (e.g.
   260001) can name up to ~8 different NUBA vouchers.  A single SN reference
   can therefore cover several distinct ERP documents.  We match at the
   (ReferenceNumber, TrxNumber) level -- each TrxNumber is one ERP document --
   and disambiguate candidates by amount (exact within tolerance).
2. SN also stores petty-cash voucher numbers (4-digit) in ReferenceNumber;
   they fall inside the required ^\\d{4,6}$ net but are almost never NUBA
   journal vouchers.  They surface on the 'In SpectroNova not NUBA' sheet
   with their LinkSourceDocType so the reader can discount them.

Everything sums from source cells; the Summary sheet proves the reconciliation
identity on both sides (any residual > tolerance aborts the run).
Re-runnable monthly: defaults pick the newest NUBA .xls / SN JV xlsx.
Console output is ASCII-only (cp1252 console; Arabic lives in the xlsx).
"""
import argparse
import datetime
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import nuba_reader  # noqa: E402

REPO = Path(__file__).resolve().parents[2]
NUBA_REF = re.compile(r"^(\d{4,6}|C\d{8})$")

SN_HEADER = ("DocumentDate", "TrxNumber", "ReferenceNumber", "GLCode", "GLName",
             "Description", "Debit", "Credit", "LinkSourceDocID",
             "LinkSourceDocType", "LinkSourceDocumentNumber")


def clean_ref(v):
    """SN ReferenceNumber: literal 'NULL'/blank means no reference; numbers -> text."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).strip()
    return None if s in ("", "NULL") else s


def load_sn(path):
    """Return flat SN row list. Amounts to float, dates to date, ref cleaned."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["JV Data"]
    it = ws.iter_rows(values_only=True)
    hdr = tuple(str(h).strip() if h else "" for h in next(it))
    if hdr[:len(SN_HEADER)] != SN_HEADER:
        raise ValueError("SN header drifted: %r" % (hdr,))
    rows = []
    for r in it:
        if r[0] is None and r[1] is None and r[6] is None and r[7] is None:
            continue  # trailing blank rows in read_only mode
        d = r[0].date() if isinstance(r[0], datetime.datetime) else r[0]
        rows.append({
            "date": d,
            "trx": "" if r[1] is None else str(r[1]).strip(),
            "ref": clean_ref(r[2]),
            "gl": "" if r[3] is None else str(r[3]).strip(),
            "link": "" if r[9] in (None, "NULL") else str(r[9]).strip(),
            "debit": float(r[6] or 0.0),
            "credit": float(r[7] or 0.0),
        })
    wb.close()
    return rows


def reconcile(vouchers, sn_rows, tol, date_tol):
    """Match NUBA vouchers to SN (ref, trx) document groups. Returns result dict."""
    # SN side: one group per ERP document that carries a NUBA-shaped reference
    groups = {}
    for r in sn_rows:
        if r["ref"] and NUBA_REF.match(r["ref"]):
            g = groups.setdefault((r["ref"], r["trx"]),
                                  {"debit": 0.0, "dates": set(), "links": set(), "gl": set()})
            g["debit"] += r["debit"]
            if r["date"]:
                g["dates"].add(r["date"])
            if r["link"]:
                g["links"].add(r["link"])
            g["gl"].add(r["gl"].split("-")[0])
    by_ref = defaultdict(list)
    for (ref, trx), g in sorted(groups.items()):
        by_ref[ref].append((trx, g))

    by_no = defaultdict(list)
    for v in vouchers:
        by_no[v["voucher_no"]].append(v)

    matched = []        # (voucher, ref, trx, group)
    amount_mm = []      # (voucher, ref, trx, group)  1:1 leftover, amounts differ
    ambiguous = []      # (ref, leftover_vouchers, leftover_sn_(trx,group))
    extra_sn = []       # (ref, trx, group, ref_known_in_nuba)
    matched_nos = set()

    for ref in sorted(by_ref):
        sgs = by_ref[ref]
        cands = by_no.get(ref)
        if not cands:
            extra_sn.extend((ref, trx, g, False) for trx, g in sgs)
            continue
        matched_nos.add(ref)
        # amount disambiguation: NUBA numbering restarts per type, so pair each
        # candidate voucher with an SN document of identical value
        used = [False] * len(sgs)
        left_c = []
        for v in cands:
            vt = nuba_reader.voucher_total(v)
            pick = next((i for i, (_t, g) in enumerate(sgs)
                         if not used[i] and abs(g["debit"] - vt) <= tol), None)
            if pick is None:
                left_c.append(v)
            else:
                used[pick] = True
                matched.append((v, ref, sgs[pick][0], sgs[pick][1]))
        left_s = [i for i in range(len(sgs)) if not used[i]]
        if len(left_c) == 1 and len(left_s) == 1:
            trx, g = sgs[left_s[0]]
            amount_mm.append((left_c[0], ref, trx, g))
        elif left_c and left_s:
            ambiguous.append((ref, left_c, [sgs[i] for i in left_s]))
        else:
            extra_sn.extend((ref, sgs[i][0], sgs[i][1], True) for i in left_s)
            # left_c with no SN leftovers -> genuinely missing in SN

    consumed = {id(v) for v, *_ in matched} | {id(v) for v, *_ in amount_mm}
    consumed |= {id(v) for _ref, vs, _s in ambiguous for v in vs}
    missing = [v for v in vouchers if id(v) not in consumed]

    date_mm = []
    for v, ref, trx, g in matched:
        if not g["dates"]:
            continue
        diff = min(abs((d - v["date"]).days) for d in g["dates"])
        if diff > date_tol:
            date_mm.append((v, ref, trx, g, diff))

    return {"groups": groups, "matched": matched, "amount_mm": amount_mm,
            "ambiguous": ambiguous, "extra_sn": extra_sn, "missing": missing,
            "date_mm": date_mm}


# ---------------------------------------------------------------- xlsx layer

AMT = "#,##0.000"
DATEFMT = "dd/mm/yyyy"
BOLD = Font(bold=True)


def add_sheet(wb, title, headers, rows, widths, num_cols=(), date_cols=()):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font = BOLD
    for row in rows:
        ws.append(row)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        for i in num_cols:
            ws.cell(row=r, column=i).number_format = AMT
        for i in date_cols:
            ws.cell(row=r, column=i).number_format = DATEFMT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), max(ws.max_row, 1))
    return ws


def fmt_dates(dates):
    ds = sorted(dates)
    if not ds:
        return ""
    if len(ds) == 1:
        return ds[0].strftime("%d/%m/%Y")
    return "%s .. %s (%d)" % (ds[0].strftime("%d/%m/%Y"), ds[-1].strftime("%d/%m/%Y"), len(ds))


def build_report(out_path, vouchers, sn_rows, res, tol, date_tol):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    nuba_total = sum(nuba_reader.voucher_total(v) for v in vouchers)
    sn_total = sum(r["debit"] for r in sn_rows)
    sn_pat_rows = [r for r in sn_rows if r["ref"] and NUBA_REF.match(r["ref"])]
    sn_other_rows = [r for r in sn_rows if not (r["ref"] and NUBA_REF.match(r["ref"]))]
    sn_pat_total = sum(r["debit"] for r in sn_pat_rows)
    sn_other_total = sum(r["debit"] for r in sn_other_rows)

    m_sn = sum(g["debit"] for *_x, g in res["matched"])
    m_nuba = sum(nuba_reader.voucher_total(v) for v, *_x in res["matched"])
    mm_sn = sum(g["debit"] for *_x, g in res["amount_mm"])
    mm_nuba = sum(nuba_reader.voucher_total(v) for v, *_x in res["amount_mm"])
    amb_sn = sum(g["debit"] for _r, _vs, sgs in res["ambiguous"] for _t, g in sgs)
    amb_nuba = sum(nuba_reader.voucher_total(v) for _r, vs, _s in res["ambiguous"] for v in vs)
    ex_sn = sum(g["debit"] for _r, _t, g, _k in res["extra_sn"])
    miss_nuba = sum(nuba_reader.voucher_total(v) for v in res["missing"])

    # the identity MUST tie -- never publish a report that doesn't add up
    resid_sn = sn_total - (m_sn + mm_sn + amb_sn + ex_sn + sn_other_total)
    resid_nuba = nuba_total - (m_nuba + mm_nuba + amb_nuba + miss_nuba)
    if abs(resid_sn) > 0.001 or abs(resid_nuba) > 0.001:
        raise AssertionError("identity residual SN=%.6f NUBA=%.6f" % (resid_sn, resid_nuba))

    # ---- Summary
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 52
    for col, w in (("B", 16), ("C", 16), ("D", 16), ("E", 16)):
        ws.column_dimensions[col].width = w

    def put(label, *vals, bold=False, amt_cols=()):
        ws.append((label,) + vals)
        r = ws.max_row
        if bold:
            for c in ws[r]:
                c.font = BOLD
        for i in amt_cols:
            ws.cell(row=r, column=i).number_format = AMT
        return r

    put("JV RECONCILIATION - NUBA vs SpectroNova", bold=True)
    put("Period: %s to %s   |   run %s   |   tolerance %.3f KWD, date tolerance %d day(s)"
        % (min(v["date"] for v in vouchers).strftime("%d/%m/%Y"),
           max(v["date"] for v in vouchers).strftime("%d/%m/%Y"),
           datetime.date.today().strftime("%d/%m/%Y"), tol, date_tol))
    put("")
    put("GRAND TOTALS (debit = credit on both sides)", bold=True)
    put("NUBA daily journal total", nuba_total, amt_cols=(2,))
    put("SpectroNova JV total", sn_total, amt_cols=(2,))
    put("Difference (SpectroNova - NUBA)", sn_total - nuba_total, amt_cols=(2,))
    put("")
    put("MATCH RESULT", "count", "NUBA value", "SN value", bold=True)
    put("Matched vouchers (amount agrees within tolerance)", len(res["matched"]), m_nuba, m_sn, amt_cols=(3, 4))
    put("  of which dates differ (see 'Date mismatch')", len(res["date_mm"]))
    put("Amount mismatch (same voucher no, value differs)", len(res["amount_mm"]), mm_nuba, mm_sn, amt_cols=(3, 4))
    put("Ambiguous duplicates (see 'Ambiguous')",
        sum(len(vs) for _r, vs, _s in res["ambiguous"]), amb_nuba, amb_sn, amt_cols=(3, 4))
    put("NUBA vouchers missing in SpectroNova", len(res["missing"]), miss_nuba, amt_cols=(3,))
    put("SN documents with NUBA-style ref not matched", len(res["extra_sn"]), None, ex_sn, amt_cols=(4,))
    put("")
    put("RECONCILIATION IDENTITY (SpectroNova side, debit)", bold=True)
    put("Matched to NUBA vouchers", m_sn, amt_cols=(2,))
    put("+ Amount-mismatched refs", mm_sn, amt_cols=(2,))
    put("+ Ambiguous refs", amb_sn, amt_cols=(2,))
    put("+ Unmatched NUBA-style refs (mostly petty cash)", ex_sn, amt_cols=(2,))
    put("+ Rows with no NUBA-style reference", sn_other_total, amt_cols=(2,))
    put("= SpectroNova JV total", m_sn + mm_sn + amb_sn + ex_sn + sn_other_total, bold=True, amt_cols=(2,))
    put("")
    put("RECONCILIATION IDENTITY (NUBA side, debit)", bold=True)
    put("Matched", m_nuba, amt_cols=(2,))
    put("+ Amount-mismatched", mm_nuba, amt_cols=(2,))
    put("+ Ambiguous", amb_nuba, amt_cols=(2,))
    put("+ Missing in SpectroNova", miss_nuba, amt_cols=(2,))
    put("= NUBA journal total", m_nuba + mm_nuba + amb_nuba + miss_nuba, bold=True, amt_cols=(2,))
    put("")
    put("PER-MONTH TOTALS (debit)", bold=True)
    put("Month", "NUBA", "SN w/ NUBA ref", "SN other", "SN total", bold=True)
    months = sorted({v["date"].strftime("%Y-%m") for v in vouchers}
                    | {r["date"].strftime("%Y-%m") for r in sn_rows if r["date"]})
    nuba_m = defaultdict(float)
    for v in vouchers:
        nuba_m[v["date"].strftime("%Y-%m")] += nuba_reader.voucher_total(v)
    snp_m, sno_m = defaultdict(float), defaultdict(float)
    for r in sn_pat_rows:
        snp_m[r["date"].strftime("%Y-%m")] += r["debit"]
    for r in sn_other_rows:
        sno_m[r["date"].strftime("%Y-%m")] += r["debit"]
    for m in months:
        put(m, nuba_m[m], snp_m[m], sno_m[m], snp_m[m] + sno_m[m], amt_cols=(2, 3, 4, 5))
    put("Total", sum(nuba_m.values()), sum(snp_m.values()), sum(sno_m.values()),
        sum(snp_m.values()) + sum(sno_m.values()), bold=True, amt_cols=(2, 3, 4, 5))
    put("")
    put("HOW TO READ THIS REPORT", bold=True)
    for note in (
        "Only exceptions are listed on the other sheets; a voucher that matched on amount and date appears nowhere.",
        "Matching key: NUBA voucher number found in SpectroNova ReferenceNumber; one ERP document = one TrxNumber.",
        "NUBA voucher numbering restarts for every voucher type, so the same number can name several vouchers;",
        "those are told apart by amount. Where amounts cannot tell them apart, see the 'Ambiguous' sheet.",
        "'In SpectroNova not NUBA' includes 4-digit references: these are almost all petty-cash voucher numbers",
        "(LinkSourceDocType = PC_VOUCHER), not daily-journal vouchers - check that column before chasing them.",
        "'Missing in SpectroNova' means NO JV Data row carries that voucher number in ReferenceNumber.",
        "Most of that sheet is Purchase Invoice LPO/WO and Journal (Sanad Qayd) vouchers: SpectroNova books",
        "purchases through its own SupInv / Inventory / GRN document chain without quoting the NUBA number,",
        "so 'missing by reference' usually means 'recorded under a different document trail', not lost value.",
        "The gap between the two grand totals is SpectroNova activity with no NUBA-style reference",
        "(depreciation runs, HR process runs, inventory receipts, etc.) minus NUBA vouchers not referenced in SN;",
        "the identity blocks above quantify both directions to the fils.",
    ):
        put(note)
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.freeze_panes = "A3"

    # ---- Missing in SpectroNova
    add_sheet(
        wb, "Missing in SpectroNova",
        ("Date", "Voucher type", "Voucher no", "Amount (KWD)", "First line description"),
        [(v["date"], v["vtype"], v["voucher_no"], nuba_reader.voucher_total(v),
          v["lines"][0]["descr"] if v["lines"] else "")
         for v in sorted(res["missing"], key=lambda v: (v["date"], v["vtype"], v["voucher_no"]))],
        widths=(12, 28, 14, 16, 60), num_cols=(4,), date_cols=(1,))

    # ---- Amount mismatch
    add_sheet(
        wb, "Amount mismatch",
        ("Voucher no", "Voucher type", "NUBA date", "NUBA amount", "SN amount",
         "Difference (SN-NUBA)", "SN TrxNumber", "SN date(s)"),
        [(ref, v["vtype"], v["date"], nuba_reader.voucher_total(v), g["debit"],
          g["debit"] - nuba_reader.voucher_total(v), trx, fmt_dates(g["dates"]))
         for v, ref, trx, g in sorted(res["amount_mm"], key=lambda t: t[1])],
        widths=(14, 28, 12, 16, 16, 16, 20, 26), num_cols=(4, 5, 6), date_cols=(3,))

    # ---- Date mismatch
    add_sheet(
        wb, "Date mismatch",
        ("Voucher no", "Voucher type", "Amount (KWD)", "NUBA date", "SN date(s)",
         "Day diff", "SN TrxNumber"),
        [(ref, v["vtype"], g["debit"], v["date"], fmt_dates(g["dates"]), diff, trx)
         for v, ref, trx, g, diff in sorted(res["date_mm"], key=lambda t: (-t[4], t[1]))],
        widths=(14, 28, 16, 12, 26, 10, 20), num_cols=(3,), date_cols=(4,))

    # ---- In SpectroNova not NUBA
    add_sheet(
        wb, "In SpectroNova not NUBA",
        ("ReferenceNumber", "Ref exists in NUBA", "SN amount (KWD)", "SN date(s)",
         "TrxNumber", "LinkSourceDocType", "GL prefixes"),
        [(ref, "yes - other type/amount" if known else "no", g["debit"], fmt_dates(g["dates"]),
          trx, ", ".join(sorted(g["links"])) or "(none)", ", ".join(sorted(g["gl"])))
         for ref, trx, g, known in sorted(res["extra_sn"], key=lambda t: (t[3], t[0], t[1]))],
        widths=(16, 20, 16, 26, 20, 24, 16), num_cols=(3,))

    # ---- Ambiguous
    add_sheet(
        wb, "Ambiguous",
        ("Voucher no", "NUBA candidates (unmatched)", "SN documents (unmatched)",
         "NUBA value", "SN value"),
        [(ref,
          "; ".join("%s %s = %s" % (v["vtype"], v["date"].strftime("%d/%m"),
                                    format(nuba_reader.voucher_total(v), ",.3f")) for v in vs),
          "; ".join("%s = %s" % (trx, format(g["debit"], ",.3f")) for trx, g in sgs),
          sum(nuba_reader.voucher_total(v) for v in vs),
          sum(g["debit"] for _t, g in sgs))
         for ref, vs, sgs in sorted(res["ambiguous"])],
        widths=(14, 70, 55, 16, 16), num_cols=(4, 5))
    for row in wb["Ambiguous"].iter_rows(min_row=2, min_col=2, max_col=3):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {
        "nuba_total": nuba_total, "sn_total": sn_total,
        "sn_pat_total": sn_pat_total, "sn_other_total": sn_other_total,
        "m_nuba": m_nuba, "m_sn": m_sn, "mm_sn": mm_sn, "amb_sn": amb_sn, "ex_sn": ex_sn,
        "miss_nuba": miss_nuba,
    }


def main():
    ap = argparse.ArgumentParser(description="Reconcile NUBA daily journal vs SpectroNova JV")
    ap.add_argument("--nuba", type=Path,
                    default=None, help="NUBA .xls (default: newest in Accounting raw data/legacy-jv)")
    ap.add_argument("--sn", type=Path,
                    default=None, help="SpectroNova JV .xlsx (default: newest JV Data* in spectronova)")
    ap.add_argument("--out", type=Path, default=None,
                    help="output xlsx (default: Accounting raw data/output/JV_reconciliation_<today>.xlsx)")
    ap.add_argument("--tol", type=float, default=0.005, help="amount tolerance, KWD")
    ap.add_argument("--date-tol", type=int, default=0, help="date tolerance, days")
    args = ap.parse_args()

    raw = REPO / "Accounting raw data"
    nuba_path = args.nuba or max((raw / "legacy-jv").glob("*.xls"), key=lambda p: p.stat().st_mtime)
    sn_path = args.sn or max((raw / "spectronova").glob("JV Data*.xlsx"), key=lambda p: p.stat().st_mtime)
    out_path = args.out or raw / "output" / ("JV_reconciliation_%s.xlsx" % datetime.date.today().isoformat())

    print("NUBA source : %s" % nuba_path)
    print("SN source   : %s" % sn_path)

    vouchers, printed = nuba_reader.read_nuba(str(nuba_path))
    problems = nuba_reader.validate(vouchers, printed)
    if problems:
        for p in problems:
            print("NUBA FAIL: " + p)
        sys.exit(1)
    st = nuba_reader.stats(vouchers)
    print("NUBA        : %d vouchers, %d lines, total %.3f (ties printed footer)"
          % (st["vouchers"], st["lines"], st["debit"]))

    sn_rows = load_sn(sn_path)
    sn_deb = sum(r["debit"] for r in sn_rows)
    sn_cre = sum(r["credit"] for r in sn_rows)
    if abs(sn_deb - sn_cre) > 0.001:
        print("SN FAIL: debit %.3f != credit %.3f" % (sn_deb, sn_cre))
        sys.exit(1)
    print("SpectroNova : %d rows, total %.3f (balanced)" % (len(sn_rows), sn_deb))

    res = reconcile(vouchers, sn_rows, args.tol, args.date_tol)
    tot = build_report(out_path, vouchers, sn_rows, res, args.tol, args.date_tol)

    print("")
    print("Matched vouchers        : %6d  value %15s KWD" % (len(res["matched"]), format(tot["m_nuba"], ",.3f")))
    print("  with date mismatch    : %6d" % len(res["date_mm"]))
    print("Amount mismatches       : %6d  SN value %12s" % (len(res["amount_mm"]), format(tot["mm_sn"], ",.3f")))
    print("Ambiguous NUBA vouchers : %6d  (%d references)"
          % (sum(len(vs) for _r, vs, _s in res["ambiguous"]), len(res["ambiguous"])))
    print("Missing in SpectroNova  : %6d  value %15s" % (len(res["missing"]), format(tot["miss_nuba"], ",.3f")))
    print("SN refs not matched     : %6d  value %15s" % (len(res["extra_sn"]), format(tot["ex_sn"], ",.3f")))
    print("SN rows w/o NUBA ref    :         value %15s" % format(tot["sn_other_total"], ",.3f"))
    print("Identity check          : OK (both sides tie to the fils)")
    print("Report                  : %s" % out_path)


if __name__ == "__main__":
    main()
