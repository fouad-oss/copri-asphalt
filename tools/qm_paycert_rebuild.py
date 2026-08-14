# -*- coding: utf-8 -*-
"""Rebuild the historical payment certificates from the best source per WO.

Decided with Fouad on 2026-08-14. The tracking report (متابعة أوامـر العمــل
…-الجديد.xlsx) is transcribed from what the ministry actually approved, so it
is the CONTROL TOTAL per work order. Two candidate sources carry line detail:

  ours    — the dated جزئي columns of the per-WO كشف حساب workbooks (0041)
  folders — the monthly «جميع الشركات» sheets under دفعات الوزارة, summed
            per month across subcontractors, with copied-forward months
            dropped (a month whose sheet is identical to the previous
            month's for the same WO+company is a stale copy, not new work)

Per work order:
  · ours within 0.5% of the tracker            → keep ours
  · else folders within 5% of the tracker      → use the folders
  · else                                        → keep ours, FLAG for the QA
                                                  (never import a total we
                                                   cannot tie to the ministry)

Outputs:
  Desktop/quantities-backfill/paycerts-blended.json
  Desktop/quantities-backfill/paycert-rebuild-report.md   ← the QA worklist
  supabase/migrations/0041_qm_paycert_backfill_part*.sql  (regenerated)
"""
import datetime
import glob
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from qm_backfill import (  # noqa: E402
    classify_sheet, parse_company_sheet, payment_folders, wo_key_of_folder, norm_suffix,
)

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

OUT_DIR = r"C:\Users\fszog\Desktop\quantities-backfill"
TRACK_DIR = r"D:\التجميع الشهري new\اوامر العمل"
MAP_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "qm-backfill-map.json")
SQL_OUT = r"C:\Users\fszog\Desktop\Copri webapp\supabase\migrations\0041_qm_paycert_backfill.sql"
BOP = json.load(open(os.path.join(OUT_DIR, "wo-ocr", "_bop_ref.json"), encoding="utf-8"))
PCT = 1.09
OURS_TOL = 0.005     # ours must tie to the tracker this tightly to be kept
FOLDER_TOL = 0.05    # folders may differ this much and still be trusted


def cert_no_for(d):
    return (d.year - 2024) * 12 + d.month - 11


def key_of(bab, band, suf):
    return f"{bab}/{band}{norm_suffix(suf)}"


def value_of(lines):
    """pre-pct KD of {key: qty}"""
    return sum((BOP.get(k, {}).get("rate") or 0) * q for k, q in lines.items())


def load_tracker():
    path = [f for f in glob.glob(os.path.join(TRACK_DIR, "*.xlsx"))
            if "الجديد" in os.path.basename(f)
            and not os.path.basename(f).startswith("~$")
            and "-2.xlsx" not in os.path.basename(f)][0]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    best, bkey = None, None
    for nm in wb.sheetnames:
        m = re.match(r"\s*5-(\d{1,2})-(\d{4})", nm.strip())
        if not m:
            continue
        k = (int(m.group(2)), int(m.group(1)))
        if bkey is None or k > bkey or (k == bkey and "(2)" in nm):
            best, bkey = nm, k
    rows = list(wb[best].iter_rows(values_only=True))
    out = {}
    for r in rows[16:]:
        if not r or len(r) <= 17:
            continue
        m = re.match(r"^\s*(\d+)\s*$", str(r[17] or ""))
        if m:
            out[int(m.group(1))] = {
                "executed": ((r[1] or 0) + (r[2] or 0)) / PCT,   # store pre-pct
                "status": str(r[8] or "").strip(),
                "name": str(r[13] or "").strip(),
            }
    wb.close()
    return best, out


def parse_folders():
    """→ lines[wo][month][key] = qty  (copied-forward months removed),
       plus the list of dropped (wo, company, month) repeats."""
    aliases = json.load(open(MAP_PATH, encoding="utf-8")).get("aliases", {})
    raw = {}   # wo -> company -> month -> {key: qty}
    for fdate, fpath in payment_folders():
        iso = fdate.isoformat()
        print(f"  {os.path.basename(fpath)}")
        for sub in sorted(os.listdir(fpath)):
            spath = os.path.join(fpath, sub)
            if not os.path.isdir(spath):
                continue
            wk = wo_key_of_folder(sub, aliases)
            if wk is None or wk[0] != "wo":
                continue
            books = [b for b in glob.glob(os.path.join(spath, "جميع الشركات*.xls*"))
                     if not os.path.basename(b).startswith("~$")]
            if not books:
                continue
            try:
                wb = openpyxl.load_workbook(books[0], data_only=True, read_only=True)
            except Exception:
                continue
            slots = {}
            for nm in wb.sheetnames:
                kind, comp, slot = classify_sheet(nm)
                if slot and comp:
                    slots[slot] = comp
            for nm in wb.sheetnames:
                kind, comp, slot = classify_sheet(nm)
                if kind != "plain":              # قديم+جديد are per-batch, not monthly
                    continue
                if comp is None and slot is not None:
                    comp = slots.get(slot)
                if comp is None:
                    continue
                try:
                    parsed = parse_company_sheet(wb[nm])
                except Exception:
                    continue
                if not parsed:
                    continue
                lines = {}
                for (bab, band, suf), q in parsed.items():
                    lines[key_of(bab, band, suf)] = lines.get(key_of(bab, band, suf), 0) + q
                tgt = raw.setdefault(wk[1], {}).setdefault(comp, {}).setdefault(iso, {})
                for k, q in lines.items():
                    tgt[k] = tgt.get(k, 0) + q
            wb.close()

    out, dropped = {}, []
    for wo, comps in raw.items():
        for comp, months in comps.items():
            prev = None
            for iso in sorted(months):
                cur = {k: round(v, 6) for k, v in months[iso].items() if v}
                if prev is not None and cur == prev:
                    dropped.append((wo, comp, iso, value_of(cur)))
                    continue                      # folder copied forward, not new work
                prev = cur
                tgt = out.setdefault(wo, {}).setdefault(iso, {})
                for k, q in cur.items():
                    tgt[k] = tgt.get(k, 0) + q
    return out, dropped


def main():
    print("parsing دفعات الوزارة …")
    folders, dropped = parse_folders()
    sheet, tracker = load_tracker()
    ours_raw = json.load(open(os.path.join(OUT_DIR, "paycerts-data.json"), encoding="utf-8"))

    # ours: wo -> month -> {key: qty}
    ours = {}
    for no, c in ours_raw.items():
        for l in c["lines"]:
            ours.setdefault(l["wo"], {}).setdefault(c["period_end"], {})
            d = ours[l["wo"]][c["period_end"]]
            d[l["key"]] = d.get(l["key"], 0) + l["qty"]

    decisions = {}
    for wo in sorted(set(tracker) | set(ours) | set(folders)):
        t = tracker.get(wo, {}).get("executed")
        o = sum(value_of(m) for m in ours.get(wo, {}).values())
        f = sum(value_of(m) for m in folders.get(wo, {}).values())
        if t is None:
            pick, why = ("ours" if o else "none"), "not in the tracking report"
        elif t == 0:
            pick, why = ("none", "tracker shows nothing certified")
        elif o and abs(o - t) <= max(1.0, OURS_TOL * t):
            pick, why = "ours", "كشف حساب ties to the tracker"
        elif f and abs(f - t) <= FOLDER_TOL * t:
            pick, why = "folders", "payment folders tie to the tracker"
        else:
            pick, why = "flag", "neither source ties to the tracker"
        decisions[wo] = {"tracker": t, "ours": o, "folders": f, "pick": pick, "why": why,
                         "status": tracker.get(wo, {}).get("status", ""),
                         "name": tracker.get(wo, {}).get("name", "")}

    # ── assemble the certificates ───────────────────────────────────────
    certs = {}
    for wo, d in decisions.items():
        src = folders.get(wo, {}) if d["pick"] == "folders" else ours.get(wo, {})
        if d["pick"] == "none":
            continue
        for iso, lines in src.items():
            no = cert_no_for(datetime.date.fromisoformat(iso))
            c = certs.setdefault(no, {"period_end": iso, "lines": []})
            for k, qty in lines.items():
                item = BOP.get(k)
                if item is None or not qty:
                    continue
                c["lines"].append({
                    "wo": wo, "key": k, "bab": item["bab"], "band": item["band"],
                    "suffix": item["suffix"], "qty": round(qty, 3),
                    "amount": round(qty * item["rate"], 3),
                    "source": d["pick"],
                })

    grand = sum(l["amount"] for c in certs.values() for l in c["lines"])
    tot_tracker = sum(v["executed"] for v in tracker.values())

    # ── report ──────────────────────────────────────────────────────────
    picks = {}
    for d in decisions.values():
        picks[d["pick"]] = picks.get(d["pick"], 0) + 1
    R = ["# Payment certificates — rebuild from the best source per work order\n",
         f"Tracking report sheet **{sheet}** is the control total (transcribed from what "
         "the ministry approved). Values below are pre-نسبة العقد unless stated.\n",
         "| decision | work orders | rule |", "|---|---|---|",
         f"| kept our كشف حساب | {picks.get('ours', 0)} | ties to the tracker within 0.5% |",
         f"| rebuilt from the payment folders | {picks.get('folders', 0)} | ties within 5% |",
         f"| **flagged for the QA** | {picks.get('flag', 0)} | neither source ties — left as-is |",
         f"| nothing certified | {picks.get('none', 0)} | tracker shows zero |",
         "",
         f"**Result: {len(certs)} certificates, {sum(len(c['lines']) for c in certs.values())} lines, "
         f"KD {grand:,.3f} pre-pct (KD {grand * PCT:,.3f} after 9%)** against the tracker's "
         f"KD {tot_tracker * PCT:,.3f} after 9% — "
         f"{(grand / tot_tracker - 1) * 100:+.2f}%.\n",
         f"{len(dropped)} copied-forward month sheets were dropped "
         f"(a month identical to the previous one for the same WO + subcontractor), "
         f"worth KD {sum(v for *_, v in dropped) * PCT:,.3f} after 9%.\n",
         "## Work orders needing the QA's attention\n",
         "Left on their existing data — the certificates below understate these work orders.\n",
         "| WO | name | tracker (after 9%) | ours | folders | short by | status |",
         "|---|---|---|---|---|---|---|"]
    flagged = [(wo, d) for wo, d in decisions.items() if d["pick"] == "flag"]
    for wo, d in sorted(flagged, key=lambda x: -(x[1]["tracker"] or 0)):
        short = (d["tracker"] or 0) - d["ours"]
        R.append(f"| {wo} | {d['name'][:28]} | {(d['tracker'] or 0) * PCT:,.3f} | "
                 f"{d['ours'] * PCT:,.3f} | {d['folders'] * PCT:,.3f} | "
                 f"{short * PCT:,.3f} | {d['status']} |")
    R.append(f"\n**{len(flagged)} work orders, KD "
             f"{sum((d['tracker'] or 0) - d['ours'] for _, d in flagged) * PCT:,.3f} of certified "
             f"value still unaccounted for after 9%.**\n")
    R += ["## Every work order\n",
          "| WO | decision | tracker | ours | folders | why |", "|---|---|---|---|---|---|"]
    for wo, d in sorted(decisions.items()):
        R.append(f"| {wo} | {d['pick']} | {(d['tracker'] or 0) * PCT:,.3f} | {d['ours'] * PCT:,.3f} "
                 f"| {d['folders'] * PCT:,.3f} | {d['why']} |")
    if dropped:
        R += ["", "## Dropped copied-forward month sheets\n",
              "| WO | subcontractor | month | value (after 9%) |", "|---|---|---|---|"]
        for wo, comp, iso, v in sorted(dropped, key=lambda x: -x[3]):
            R.append(f"| {wo} | {comp} | {iso} | {v * PCT:,.3f} |")
    open(os.path.join(OUT_DIR, "paycert-rebuild-report.md"), "w", encoding="utf-8").write("\n".join(R) + "\n")
    json.dump({"decisions": decisions, "certs": certs},
              open(os.path.join(OUT_DIR, "paycerts-blended.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    # ── SQL ─────────────────────────────────────────────────────────────
    def q(s):
        return "'" + str(s).replace("'", "''") + "'"

    HEADER = [
        "-- GENERATED by tools/qm_paycert_rebuild.py — do not hand-edit.",
        "-- Historical MPW payment certificates. Per work order the source is",
        "-- whichever of (a) the كشف حساب جزئي columns or (b) the monthly",
        "-- «جميع الشركات» sheets under دفعات الوزارة ties to the ministry",
        "-- tracking report; copied-forward month sheets are excluded.",
        "-- Work orders that tie to neither are left on their existing data and",
        "-- listed in Desktop\\quantities-backfill\\paycert-rebuild-report.md.",
        "-- Paste 0040 FIRST, then these parts IN ORDER. Re-runnable: the",
        "-- backfill-sourced certificates are rebuilt, hand-made ones untouched.",
        "do $qmpc$",
        "declare",
        "  v_contract bigint;",
        "  v_c bigint;",
        "  v_k bigint;",
        "  v_item bigint;",
        "begin",
        "  select id into v_contract from qm_contracts where code = 'HAW9';",
        "  if v_contract is null then raise exception 'run 0033 first'; end if;",
    ]
    L = []
    for no in sorted(certs):
        c = certs[no]
        L.append(f"\n  -- ══ دفعة {no} — حتى {c['period_end']} ══")
        L.append(f"  delete from qm_pay_cert_lines where cert_id in "
                 f"(select id from qm_pay_certs where contract_id = v_contract "
                 f"and cert_no = {no} and source = 'mpw');")
        L.append(f"  delete from qm_pay_certs where contract_id = v_contract "
                 f"and cert_no = {no} and source = 'mpw';")
        L.append("  insert into qm_pay_certs (contract_id, cert_no, period_end, source, status, note)")
        L.append(f"  values (v_contract, {no}, {q(c['period_end'])}, 'mpw', 'certified', "
                 f"{q('استيراد تاريخي — دفعات الوزارة')}) returning id into v_c;")
        L.append("  insert into qm_changelog (entity, entity_id, action, field, line_ref, old_value, new_value, actor_email)")
        L.append(f"  values ('paycert', v_c, 'create', '', '', '', "
                 f"{q('استيراد دفعة الوزارة رقم ' + str(no))}, 'backfill');")
        for ln in c["lines"]:
            L.append(f"  select id into v_k from qm_kashefs where contract_id = v_contract and kashef_no = {ln['wo']};")
            L.append(f"  select id into v_item from qm_bop_items where contract_id = v_contract "
                     f"and bab = {ln['bab']} and band = {ln['band']} "
                     f"and translate(coalesce(suffix,''), 'أإآ', 'ااا') = {q(ln['suffix'])};")
            L.append(f"  if v_item is null then raise exception 'bop {ln['key']} missing'; end if;")
            L.append("  insert into qm_pay_cert_lines (cert_id, kashef_id, bop_item_id, qty, amount)")
            L.append(f"  values (v_c, v_k, v_item, {ln['qty']}, {ln['amount']})")
            L.append("  on conflict (cert_id, kashef_id, bop_item_id) do update set "
                     "qty = qm_pay_cert_lines.qty + excluded.qty, "
                     "amount = qm_pay_cert_lines.amount + excluded.amount;")
        L.append("__CERT_BREAK__")

    MAX = 700_000
    parts, cur = [], list(HEADER)
    for chunk in "\n".join(L).split("__CERT_BREAK__"):
        if cur and sum(len(x) for x in cur) + len(chunk) > MAX:
            cur.append("end $qmpc$;")
            parts.append("\n".join(cur))
            cur = list(HEADER)
        cur.append(chunk)
    cur.append("end $qmpc$;")
    parts.append("\n".join(cur))
    base, ext = os.path.splitext(SQL_OUT)
    for old in glob.glob(f"{base}*{ext}"):
        os.remove(old)
    for i, body in enumerate(parts, 1):
        path = SQL_OUT if len(parts) == 1 else f"{base}_part{i}{ext}"
        open(path, "w", encoding="utf-8").write(
            body.replace("-- GENERATED by", f"-- PART {i} of {len(parts)}\n-- GENERATED by", 1) + "\n")
        print(f"SQL: {os.path.basename(path)} ({os.path.getsize(path):,} bytes)")

    print(f"\ndecisions: {picks}")
    print(f"certs {len(certs)}, lines {sum(len(c['lines']) for c in certs.values())}, "
          f"KD {grand * PCT:,.3f} after 9% vs tracker {tot_tracker * PCT:,.3f}")
    print(f"dropped copies: {len(dropped)} | flagged WOs: {len(flagged)}")


if __name__ == "__main__":
    main()
