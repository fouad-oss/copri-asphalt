# -*- coding: utf-8 -*-
"""Expressway EXECUTED split — attribute طلبات التدقيق to subcontractors.

0051 booked all 1,090 Expressway طلبات التدقيق on «كوبري — تنفيذ ذاتي»
because the request cross-tabs name no subcontractor. 0054 split the
ALLOCATION tier from the subcontractor claim workbooks. This tool closes the
gap: it re-reads the same claim workbooks as a TIME SERIES and reassigns each
request (or, where a request straddles vendors, each request LINE) to the
subcontractor whose payment certificate absorbed it. Emits 0057.

── the evidence (Fouad, 2026-08-17) ───────────────────────────────────────
Each subcontractor folder holds numbered workbooks = payment certificates in
chronological order. `رئيسي` states «الاعمال المنتهية حتى dd/mm/yyyy» (the
period end, the 5th of the month, same calendar as the ministry certs) and
«كشف رقم N». Each numbered sheet is one work order with, per BOP item,
الكمية السابقة / الحالية / الاجمالية. Example: WO 54's طلبات التدقيق carry
33/2 = 5,078.75, and بحر الابداع (مدني) certificate 13, sheet 54, carries
exactly 5,078.75 cumulative — as does every other item on that sheet.

── method, per (work order, BOP item) ─────────────────────────────────────
Fold the folders onto vendors first (same map as 0054), building each
vendor's CUMULATIVE curve {period_end → qty}; deltas between consecutive
certificates are what that vendor billed in the period.

  A. no vendor claims the item                → all lines stay on كوبري.
  B. one vendor, Σclaim ≥ Σrequests (−0.1%)   → every request line → vendor
                                                (whole; claim above requests
                                                is capped — WO 1 ريكافكو).
  C. otherwise (vendor claims part, or several vendors):
     1. exact — a request line equals one certificate delta;
     2. combo — 2–5 request lines sum to a delta, or one line equals the
        sum of 2–3 consecutive deltas (billed across certificates);
     3. pro rata — whatever a vendor still needs after 1–2 is spread over
        the still-unassigned lines in proportion; the remainder of each
        line stays with كوبري. Reported separately so it can be judged.
     Nearest-date candidates are preferred, but DATES ARE ONLY A TIE-BREAK:
     the request dates and the certificate dates disagree by months in both
     directions (billing lags the request; some request dates are typos
     like 1901 or 2255), so quantity equality is the evidence, not timing.

A request whose lines land on more than one vendor is SPLIT: the existing
row is reassigned to the vendor with the largest share and its lines
rewritten; a new row (same serial, date, note) is inserted per other vendor.
Requests untouched by any subcontractor are not mentioned in the SQL.

── heals a 0051 gap ───────────────────────────────────────────────────────
Six requests were skipped by 0051's guard — five share (WO, serial, date)
with another request, and one UNDATED request (WO 15 طلب 514) shares its
serial with a dated one (undated rows were guarded on serial alone). They are inserted here with note «… (2)», which is also the
discriminator that keeps this migration idempotent.

Always writes the report + dataset. Gated on the same
tools/qm-expw-subs-map.json as 0054 (confirmed: true).
"""
import sys, io, os, re, json, glob, unicodedata, collections, datetime, itertools

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DESK = os.path.join(os.path.expanduser("~"), "Desktop")
SUBDIR = os.path.join(DESK, "ExpresswaysQMbackfill", "دفعات مقاولي الباطن")
OUT_DIR = os.path.join(DESK, "quantities-backfill")
MAP_PATH = os.path.join(HERE, "qm-expw-subs-map.json")
MIG = os.path.join(REPO, "supabase", "migrations", "0057_qm_expw_exec_split.sql")
COPRI = "كوبري — تنفيذ ذاتي"
NOTE = "استيراد تاريخي — الطرق السريعة"
AR = "اأإآبتثجحخدذرزسشصضطظعغفقكلمنهةوىي"
TOL = 0.011          # two-decimal equality
REL = 0.001          # 0.1 % on totals


def clean(v):
    if v is None:
        return ""
    s = "".join(ch for ch in str(v) if unicodedata.category(ch) != "Cf")
    return re.sub(r"\s+", " ", s.replace("\xa0", " ").replace("ـ", "")).strip()


def esc(s):
    return str(s).replace("'", "''")


def num(v):
    v = round(float(v), 4)
    if v == int(v):
        return str(int(v))
    return repr(v)


def kd(v):
    return "{:,.0f}".format(v)


def parse_code(s):
    s = clean(s)
    nums = re.findall(r"\d+", s)
    lets = re.findall("[" + AR + "]+", s)
    if len(nums) < 2:
        return None
    return "%d|%d|%s" % (int(nums[1]), int(nums[0]),
                         re.sub("[أإآ]", "ا", lets[0])[:1] if lets else "")


def cert_meta(wb):
    """period end («الاعمال المنتهية حتى dd/mm/yyyy») and كشف رقم from the
    non-numbered sheets."""
    out = {}
    for name in wb.sheetnames:
        if re.match(r"^\d+$", clean(name)):
            continue
        for row in wb[name].iter_rows(min_row=1, max_row=12, max_col=6, values_only=True):
            for v in row:
                s = clean(v)
                m = re.search(r"حتى\s*(\d{1,2})/(\d{1,2})/(\d{4})", s)
                if m and "date" not in out:
                    out["date"] = "%s-%02d-%02d" % (m.group(3), int(m.group(2)), int(m.group(1)))
                m2 = re.search(r"كشف رقم\s*:?\s*(\d+)", s)
                if m2 and "kashf" not in out:
                    out["kashf"] = int(m2.group(1))
    return out


def read_certs():
    files = [p for p in glob.glob(os.path.join(SUBDIR, "**", "*.xls*"), recursive=True)
             if not os.path.basename(p).startswith("~$")]
    groups = collections.defaultdict(list)
    for p in files:
        parts = os.path.relpath(p, SUBDIR).split(os.sep)
        sub = parts[0] if len(parts) > 1 else os.path.splitext(parts[0])[0]
        if len(parts) > 2:
            sub = parts[0] + " / " + parts[1]
        base = os.path.splitext(os.path.basename(p))[0]
        m = re.search(r"(\d+)\s*$", base)
        groups[sub].append((int(m.group(1)) if m else 0, p, base))
    certs = []
    for sub in sorted(groups):
        for n, p, base in sorted(groups[sub]):
            wb = openpyxl.load_workbook(p, data_only=True)
            meta = cert_meta(wb)
            wos = {}
            for name in wb.sheetnames:
                sn = clean(name)
                if not re.match(r"^\d+$", sn):
                    continue
                rows = [list(r) for r in wb[name].iter_rows(max_col=8, values_only=True)]
                hdr = next((i for i, row in enumerate(rows[:8])
                            if any(clean(v).startswith("باب") for v in row)), None)
                if hdr is None:
                    continue
                got = {}
                for row in rows[hdr + 1:]:
                    k = parse_code(row[0] if row else None)
                    if not k:
                        continue
                    try:
                        cum = float(row[5])
                    except (TypeError, ValueError, IndexError):
                        continue
                    if cum > 0:
                        got[k] = cum
                if got:
                    wos[int(sn)] = got
            wb.close()
            certs.append({"sub": sub, "n": n, "base": base, "date": meta.get("date"),
                          "kashf": meta.get("kashf"), "wos": wos})
    return groups, certs


def build_series(certs, vendor_key):
    """(wo, k) -> vk -> sorted [(date, cumulative)] — folders of one vendor
    are summed by carrying each folder's latest cumulative forward."""
    folder = collections.defaultdict(lambda: collections.defaultdict(dict))
    for c in certs:
        if not c["date"]:
            continue
        for wo, items in c["wos"].items():
            for k, cum in items.items():
                folder[(wo, k)][c["sub"]][c["date"]] = cum
    series = {}
    for key, subs in folder.items():
        dates = sorted({d for s in subs.values() for d in s})
        acc = collections.defaultdict(lambda: collections.defaultdict(float))
        for sub, s in subs.items():
            vk = vendor_key(sub)
            for d in dates:
                prior = [s[x] for x in sorted(s) if x <= d]
                acc[vk][d] += prior[-1] if prior else 0.0
        series[key] = {vk: sorted(m.items()) for vk, m in acc.items()}
    return series


def days(a, b):
    try:
        return abs((datetime.date.fromisoformat(a) - datetime.date.fromisoformat(b)).days)
    except (ValueError, TypeError):
        return 10 ** 6


def attribute(lines, vend):
    """lines: [{'date','qty','give':{vk:qty},'how':str}] (mutated)
    vend: vk -> [(date, cum)].  Returns per-pair stats dict."""
    st = {"method": None, "prorata_qty": 0.0, "capped": 0.0}
    tot_r = sum(l["qty"] for l in lines)
    final = {vk: s[-1][1] for vk, s in vend.items()}
    tot_c = sum(final.values())
    if len(vend) == 1 and tot_c >= tot_r * (1 - REL):
        vk = next(iter(vend))
        for l in lines:
            l["give"][vk] = l["qty"]
            l["how"] = "whole"
        st["method"] = "whole"
        st["capped"] = max(0.0, tot_c - tot_r)
        return st
    # deltas per vendor
    deltas = {}
    for vk, s in vend.items():
        prev, dl = 0.0, []
        for d, cum in s:
            if cum - prev > TOL:
                dl.append([d, cum - prev, False])   # [date, qty, used]
            prev = cum
        deltas[vk] = dl
    assigned = collections.defaultdict(float)

    def free():
        return [l for l in lines if not l["give"]]

    # 1. exact single line == single delta
    for vk in sorted(vend, key=lambda v: -final[v]):
        for dl in deltas[vk]:
            c = [l for l in free() if abs(l["qty"] - dl[1]) <= TOL]
            if c:
                l = min(c, key=lambda x: days(x["date"], dl[0]))
                l["give"][vk] = l["qty"]; l["how"] = "exact"
                dl[2] = True; assigned[vk] += l["qty"]
    # 2a. 2–3 lines == one delta
    for vk in sorted(vend, key=lambda v: -final[v]):
        for dl in deltas[vk]:
            if dl[2]:
                continue
            cand = sorted(free(), key=lambda x: days(x["date"], dl[0]))[:14]
            hit = None
            for r in (2, 3, 4, 5):
                for combo in itertools.combinations(cand, r):
                    if abs(sum(x["qty"] for x in combo) - dl[1]) <= TOL:
                        hit = combo; break
                if hit:
                    break
            if hit:
                for l in hit:
                    l["give"][vk] = l["qty"]; l["how"] = "combo"
                    assigned[vk] += l["qty"]
                dl[2] = True
    # 2b. one line == 2–3 consecutive unused deltas
    for vk in sorted(vend, key=lambda v: -final[v]):
        dl = deltas[vk]
        for i in range(len(dl)):
            for r in (2, 3):
                run = dl[i:i + r]
                if len(run) < r or any(x[2] for x in run):
                    continue
                s = sum(x[1] for x in run)
                c = [l for l in free() if abs(l["qty"] - s) <= TOL]
                if c:
                    l = min(c, key=lambda x: days(x["date"], run[-1][0]))
                    l["give"][vk] = l["qty"]; l["how"] = "combo"
                    assigned[vk] += l["qty"]
                    for x in run:
                        x[2] = True
                    break
    # 3. pro rata for what each vendor still needs
    need = {vk: final[vk] - assigned[vk] for vk in vend if final[vk] - assigned[vk] > TOL}
    rest = free()
    tot_u = sum(l["qty"] for l in rest)
    if need and rest and tot_u > 0:
        tot_n = sum(need.values())
        if tot_n >= tot_u * (1 - REL):
            # subs between them absorb every remaining line
            st["capped"] = max(0.0, tot_n - tot_u)
            fr = {vk: n / tot_n for vk, n in need.items()}
        else:
            fr = {vk: n / tot_u for vk, n in need.items()}
        for l in rest:
            for vk, f in fr.items():
                q = round(l["qty"] * f, 4)
                if q > 0:
                    l["give"][vk] = q
            l["how"] = "prorata"
            st["prorata_qty"] += sum(l["give"].values())
    elif need and not rest:
        st["capped"] = sum(need.values())
    st["method"] = "lines"
    return st


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--wo", type=int, nargs="*", default=None,
                    help="emit SQL for these work orders only (incremental, after an "
                         "incremental qm_expw_tadqiq.py --wo run); report + dataset stay complete")
    ap.add_argument("--mig", default=None, help="migration file name instead of 0057")
    args = ap.parse_args()
    mig_path = os.path.join(os.path.dirname(MIG), args.mig) if args.mig else MIG
    if args.wo and not args.mig:
        raise SystemExit("--wo needs --mig (never overwrite the applied 0057)")

    bop = {"%d|%d|%s" % (it["bab"], it["band"], it["suffix"] or ""): it
           for it in json.load(open(os.path.join(OUT_DIR, "expw-bop-data.json"), encoding="utf-8"))}
    td = json.load(open(os.path.join(OUT_DIR, "expw-tadqiq-data.json"), encoding="utf-8"))
    alloc = json.load(open(os.path.join(OUT_DIR, "expw-subs-data.json"), encoding="utf-8"))
    cfg = json.load(open(MAP_PATH, encoding="utf-8"))

    def vendor_key(sub):
        e = cfg["subs"].get(sub, {})
        if e.get("vendor_id"):
            return ("id", e["vendor_id"])
        if e.get("vendor_name"):
            return ("name", e["vendor_name"])
        return ("folder", sub)

    groups, certs = read_certs()
    series = build_series(certs, vendor_key)

    # ── requests → line objects, keyed by (wo, k) ─────────────────────
    # occurrence index disambiguates the 5 (wo, serial, date) duplicates
    seen = collections.Counter()
    reqs = []
    pair_lines = collections.defaultdict(list)
    for r in td:
        # mimic 0051's guard exactly: dated requests were guarded on
        # (kashef, vendor, serial, date); UNDATED ones on (kashef, vendor,
        # serial) only — so an undated request whose serial already appeared
        # in the same WO was skipped as well (WO 15 طلب 514).
        gkey = (r["wo"], r["serial"], r["date"]) if r["date"] else (r["wo"], r["serial"])
        occ = seen[gkey] + (0 if r["date"] else seen[(r["wo"], r["serial"], "any")])
        seen[gkey] += 1
        seen[(r["wo"], r["serial"], "any")] += 1
        rr = {"wo": r["wo"], "serial": r["serial"], "date": r["date"], "occ": occ, "lines": []}
        for l in r["lines"]:
            k = "%d|%d|%s" % (l["bab"], l["band"], l["suffix"] or "")
            lo = {"k": k, "date": r["date"] or "", "qty": float(l["qty"]),
                  "out": bool(l.get("out_of_kashef")), "give": {}, "how": "copri", "req": rr}
            rr["lines"].append(lo)
            pair_lines[(r["wo"], k)].append(lo)
        reqs.append(rr)

    # ── attribute ────────────────────────────────────────────────────
    pair_stats = {}
    for key, lines in pair_lines.items():
        vend = series.get(key)
        if not vend:
            continue
        pair_stats[key] = attribute(lines, vend)
    claim_no_req = [(key, {vk: s[-1][1] for vk, s in v.items()})
                    for key, v in series.items() if key not in pair_lines]

    # ── roll-ups ─────────────────────────────────────────────────────
    rate = lambda k: bop.get(k, {}).get("rate", 0.0)
    by_how = collections.Counter(); by_how_n = collections.Counter()
    exec_v = collections.defaultdict(float)     # vk -> value at BOP rates
    exec_q_pair = collections.defaultdict(float)  # (vk, wo, k) -> qty
    total_exec = 0.0
    for key, lines in pair_lines.items():
        for l in lines:
            v = l["qty"] * rate(l["k"])
            total_exec += v
            given = sum(l["give"].values())
            by_how[l["how"]] += v if l["how"] != "prorata" else given * rate(l["k"])
            by_how_n[l["how"]] += 1
            if l["how"] == "prorata":
                by_how["copri"] += (l["qty"] - given) * rate(l["k"])
            for vk, q in l["give"].items():
                exec_v[vk] += q * rate(l["k"])
                exec_q_pair[(vk, key[0], key[1])] += q
    copri_val = total_exec - sum(exec_v.values())

    def vlabel(vk):
        return ("vendors.id %s" % vk[1]) if vk[0] == "id" else vk[1]

    def alloc_label(vk):
        return ("('id', %s)" % vk[1]) if vk[0] == "id" else "('name', '%s')" % vk[1]

    # requests: touched / split
    touched, split = [], []
    for rr in reqs:
        vs = collections.defaultdict(float)
        for l in rr["lines"]:
            for vk, q in l["give"].items():
                vs[vk] += q
            rem = l["qty"] - sum(l["give"].values())
            if rem > 1e-6:
                vs[("copri",)] += rem
        if not any(vk != ("copri",) for vk in vs):
            if rr["occ"] > 0:
                touched.append(rr)   # dup healing, stays on COPRI
            continue
        rr["vendors"] = vs
        touched.append(rr)
        if len(vs) > 1:
            split.append(rr)

    # ── report ───────────────────────────────────────────────────────
    rep = ["# Expressway — executed (طلبات التدقيق) split by subcontractor\n"]
    rep.append("Source: `دفعات مقاولي الباطن` read as a time series — every numbered "
               "workbook is a payment certificate whose `رئيسي` states the period end "
               "(«الاعمال المنتهية حتى …», the 5th of the month) and كشف رقم; every "
               "numbered sheet is a work order with cumulative quantities per BOP item. "
               "%d certificates read across %d folders.\n" % (len(certs), len(groups)))
    rep.append("## Result\n")
    rep.append("| vendor | executed value (BOP rates, pre-pct) | allocated (0054) | exec / alloc | share of executed |")
    rep.append("|---|---|---|---|---|")
    # allocation keys in expw-subs-data.json are "('id', 10)|wo|bab|band|suf"
    alloc_by_vk = collections.defaultdict(float)
    for kk, q in alloc.items():
        s, wo, bab, band, suf = kk.split("|")
        alloc_by_vk[s] += q * rate("%s|%s|%s" % (bab, band, suf))
    for vk in sorted(exec_v, key=lambda v: -exec_v[v]):
        a = alloc_by_vk.get(alloc_label(vk), 0.0)
        rep.append("| %s | %s | %s | %s | %.1f%% |"
                   % (vlabel(vk), kd(exec_v[vk]), kd(a),
                      ("%.2f" % (exec_v[vk] / a)) if a else "—",
                      100 * exec_v[vk] / total_exec))
    rep.append("| **%s** (remainder) | **%s** | | | **%.1f%%** |"
               % (COPRI, kd(copri_val), 100 * copri_val / total_exec))
    rep.append("\n- executed total KD %s pre-pct (unchanged — only vendor_id moves); "
               "subcontracted KD %s (%.1f%%), self-performed KD %s."
               % (kd(total_exec), kd(sum(exec_v.values())),
                  100 * sum(exec_v.values()) / total_exec, kd(copri_val)))
    rep.append("- requests touched: **%d of %d** (%d split across vendors → %d extra rows); "
               "the rest stay on «%s» untouched."
               % (len(touched), len(reqs), len(split),
                  sum(len(r["vendors"]) - 1 for r in split), COPRI))

    rep.append("\n## How each request line was attributed\n")
    rep.append("| method | lines | value | share |")
    rep.append("|---|---|---|---|")
    legend = {"whole": "whole — one vendor claims the whole (WO, item), Σclaim = Σrequests",
              "exact": "exact — line equals one certificate delta",
              "combo": "combo — 2–5 lines sum to a delta / line spans 2–3 certificates",
              "prorata": "pro rata — vendor's unmatched remainder spread over unmatched lines",
              "copri": "كوبري — no subcontractor claim (or the pro-rata remainder)"}
    for h in ["whole", "exact", "combo", "prorata", "copri"]:
        rep.append("| %s | %d | %s | %.1f%% |"
                   % (legend[h], by_how_n[h], kd(by_how[h]), 100 * by_how[h] / total_exec))
    npair = collections.Counter(s["method"] for s in pair_stats.values())
    rep.append("\n- (WO, item) pairs: %d with a subcontractor claim (%d whole, %d line-matched), "
               "%d without." % (len(pair_stats), npair["whole"], npair["lines"],
                                len(pair_lines) - len(pair_stats)))

    pr = [(key, s) for key, s in pair_stats.items() if s["prorata_qty"] > 0]
    rep.append("\n## Pro-rata pairs (%d) — the judgement calls\n" % len(pr))
    rep.append("Vendor's certificate total could not be reproduced from whole request lines; "
               "the unmatched remainder was spread proportionally. Quantities are exact in "
               "total, only the split of these lines is estimated.\n")
    rep.append("| WO | bab/band | requests | claimed by | pro-rata qty | value |")
    rep.append("|---|---|---|---|---|---|")
    for key, s in sorted(pr, key=lambda x: -x[1]["prorata_qty"] * rate(x[0][1])):
        lines = pair_lines[key]
        rq = sum(l["qty"] for l in lines)
        who = ", ".join("%s %s" % (vlabel(vk), kd(sv[-1][1])) for vk, sv in series[key].items())
        b, band, suf = key[1].split("|")
        rep.append("| %d | %s/%s%s | %s | %s | %s | %s |"
                   % (key[0], b, band, suf, kd(rq), who, kd(s["prorata_qty"]),
                      kd(s["prorata_qty"] * rate(key[1]))))

    cap = [(key, s) for key, s in pair_stats.items() if s["capped"] > TOL]
    rep.append("\n## Claims above the requests (%d) — capped\n" % len(cap))
    rep.append("| WO | bab/band | requests | claimed | excess |")
    rep.append("|---|---|---|---|---|")
    for key, s in sorted(cap, key=lambda x: -x[1]["capped"] * rate(x[0][1])):
        rq = sum(l["qty"] for l in pair_lines[key])
        b, band, suf = key[1].split("|")
        rep.append("| %d | %s/%s%s | %s | %s | %s |"
                   % (key[0], b, band, suf, kd(rq), kd(rq + s["capped"]), kd(s["capped"])))

    rep.append("\n## Subcontractor claims with NO طلب تدقيق at all (%d pairs)\n" % len(claim_no_req))
    rep.append("Nothing to attribute — executed comes from the request sheets only. "
               "WO 3 has no تدقيق sheet in its workbook (BRIEF §5.3), which explains most of it.\n")
    rep.append("| WO | bab/band | vendor | claimed qty | value |")
    rep.append("|---|---|---|---|---|")
    for key, f in sorted(claim_no_req, key=lambda x: -sum(x[1].values()) * rate(x[0][1])):
        b, band, suf = key[1].split("|")
        for vk, q in f.items():
            rep.append("| %d | %s/%s%s | %s | %s | %s |"
                       % (key[0], b, band, suf, vlabel(vk), kd(q), kd(q * rate(key[1]))))

    rep.append("\n## Split requests (%d)\n" % len(split))
    rep.append("| WO | serial | date | vendors |")
    rep.append("|---|---|---|---|")
    for rr in sorted(split, key=lambda r: (r["wo"], r["serial"])):
        rep.append("| %d | %s | %s | %s |"
                   % (rr["wo"], rr["serial"], rr["date"] or "—",
                      "; ".join("%s %s" % (COPRI if vk == ("copri",) else vlabel(vk), kd(q))
                                for vk, q in rr["vendors"].items())))

    dups = [rr for rr in reqs if rr["occ"] > 0]
    rep.append("\n## Healed 0051 duplicates (%d)\n" % len(dups))
    rep.append("These collide with another request under 0051's guard (same serial+date, "
               "or undated with a dated same-serial sibling) and were skipped. Inserted here with note «%s (2)».\n" % NOTE)
    for rr in dups:
        rep.append("- WO %d — طلب %s (%s): %s" % (rr["wo"], rr["serial"], rr["date"],
                   ", ".join("%s × %s" % (l["k"], num(l["qty"])) for l in rr["lines"])))

    rep.append("\n## Certificates read\n")
    rep.append("| folder | file | كشف | period end | work orders |")
    rep.append("|---|---|---|---|---|")
    for c in certs:
        rep.append("| %s | %s | %s | %s | %s |"
                   % (c["sub"], c["base"], c["kashf"] or "—", c["date"] or "**none**",
                      ", ".join(str(w) for w in sorted(c["wos"]))))

    report = "\n".join(rep) + "\n"
    open(os.path.join(OUT_DIR, "expw-exec-report.md"), "w", encoding="utf-8").write(report)
    json.dump({"%s|%d|%s" % (vlabel(vk), wo, k): round(q, 4)
               for (vk, wo, k), q in exec_q_pair.items()},
              open(os.path.join(OUT_DIR, "expw-exec-data.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    print(report[:4000])

    if not cfg.get("confirmed"):
        print("\nNO SQL EMITTED — tools/qm-expw-subs-map.json not confirmed.")
        return

    # ── SQL ──────────────────────────────────────────────────────────
    if args.wo:
        touched = [rr for rr in touched if rr["wo"] in args.wo]
        split = [rr for rr in split if rr["wo"] in args.wo]
    vend_all = sorted({vk for rr in touched for vk in rr.get("vendors", {}) if vk != ("copri",)},
                      key=str)
    vvar = {}
    for i, vk in enumerate(vend_all):
        vvar[vk] = "v_s%d" % i
    vvar[("copri",)] = "v_copri"

    head = ["""-- %s — GENERATED by tools/qm_expw_exec_split.py, do not hand-edit.%s
-- Reassigns the Expressway طلبات التدقيق (0051, all on «%s») to the
-- subcontractor whose payment certificate absorbed each request, using
-- دفعات مقاولي الباطن read as a time series (see the tool docstring and
-- quantities-backfill/expw-exec-report.md). Quantities and dates are unchanged;
-- only vendor_id moves. %d of %d requests touched, %d split across vendors.
-- Also inserts the 6 requests 0051's guard skipped (note «… (2)»).
-- Idempotent. Paste after 0054 (needs its vendor rows).
do $qmexpwex$
declare
  v_contract bigint;
  v_copri bigint;
  v_k bigint;
  v_t bigint;
  v_item bigint;
%s
begin
  select id into v_contract from qm_contracts where code = 'EXPW';
  if v_contract is null then raise exception 'run 0047 first'; end if;
  select id into v_copri from vendors where name = '%s';
  if v_copri is null then raise exception 'COPRI pseudo-vendor missing'; end if;
""" % (os.path.basename(mig_path).replace(".sql", ""),
       ("\n-- INCREMENTAL: work orders %s only. Paste after their qm_expw_tadqiq --wo migration."
        % args.wo) if args.wo else "",
       esc(COPRI), len(touched), len(reqs), len(split),
       "\n".join("  %s bigint;" % vvar[vk] for vk in vend_all), esc(COPRI))]
    for vk in vend_all:
        if vk[0] == "id":
            head.append("  select id into %s from vendors where id = %d;\n"
                        "  if %s is null then raise exception 'vendor %d missing'; end if;"
                        % (vvar[vk], vk[1], vvar[vk], vk[1]))
        else:
            head.append("  select id into %s from vendors where name = '%s';\n"
                        "  if %s is null then raise exception 'vendor %s missing — run 0054'; end if;"
                        % (vvar[vk], esc(vk[1]), vvar[vk], esc(vk[1])))
    head.append("")

    def item_sql(k):
        b, band, suf = k.split("|")
        return ("    select id into v_item from qm_bop_items where contract_id = v_contract "
                "and bab = %s and band = %s and coalesce(suffix,'') = '%s';\n"
                "    if v_item is null then raise exception 'bop %s/%s missing'; end if;"
                % (b, band, esc(suf), band, b))

    body = []
    for rr in sorted(touched, key=lambda r: (r["wo"], r["serial"], r["date"] or "", r["occ"])):
        note = NOTE + (" (%d)" % (rr["occ"] + 1) if rr["occ"] > 0 else "")
        vs = rr.get("vendors", {("copri",): sum(l["qty"] for l in rr["lines"])})
        primary = max(vs, key=lambda v: (vs[v], v != ("copri",)))
        others = [v for v in vs if v != primary]
        date_expr = ("date '%s'" % rr["date"]) if rr["date"] else \
            "coalesce((select wo_date from qm_kashefs where id = v_k), (now() at time zone 'Asia/Kuwait')::date)"
        # undated rows carry the WO date (0051's coalesce) — constrain on it so
        # the lookup cannot pick up a dated sibling with the same serial
        date_where = ("and tadqiq_date = date '%s'" % rr["date"]) if rr["date"] else             "and tadqiq_date = " + date_expr

        def lines_for(v):
            out = []
            for l in rr["lines"]:
                q = l["give"].get(v) if v != ("copri",) else l["qty"] - sum(l["give"].values())
                if q is None or q <= 1e-6:
                    continue
                out.append(item_sql(l["k"]))
                out.append("    insert into qm_tadqiq_lines (tadqiq_id, bop_item_id, qty, out_of_kashef, over_allocation) "
                           "values (v_t, v_item, %s, %s, false);" % (num(q), "true" if l["out"] else "false"))
            return out

        body.append("\n  -- ── أمر عمل %d — طلب تدقيق %s (%s) → %s%s ──"
                    % (rr["wo"], rr["serial"], rr["date"] or "بلا تاريخ",
                       COPRI if primary == ("copri",) else vlabel(primary),
                       (" + " + ", ".join(COPRI if v == ("copri",) else vlabel(v) for v in others))
                       if others else ""))
        body.append("  select id into v_k from qm_kashefs where contract_id = v_contract and kashef_no = %d;" % rr["wo"])
        body.append("  if v_k is not null then")
        body.append("    select id into v_t from qm_tadqiq where kashef_id = v_k and serial_no = '%s' %s "
                    "and note = '%s' and not opening and vendor_id in (v_copri, %s) "
                    "order by (vendor_id = %s) desc limit 1;"
                    % (esc(rr["serial"]), date_where, esc(note), vvar[primary], vvar[primary]))
        if rr["occ"] > 0:
            # a duplicate 0051 skipped — insert it (with lines) if absent
            body.append("    if v_t is null then")
            body.append("      insert into qm_tadqiq (kashef_id, vendor_id, tadqiq_date, street_no, note, opening, serial_no)\n"
                        "      values (v_k, %s, %s, '', '%s', false, '%s') returning id into v_t;"
                        % (vvar[primary], date_expr, esc(note), esc(rr["serial"])))
            body.extend("  " + x for x in lines_for(primary))
            body.append("    else")
        else:
            # 0051 must have created it — fail loudly rather than re-create
            body.append("    if v_t is null then raise exception 'WO %d طلب %s missing — run 0051 first'; end if;"
                        % (rr["wo"], esc(rr["serial"])))
            body.append("    begin")
        body.append("      update qm_tadqiq set vendor_id = %s where id = v_t and vendor_id <> %s;"
                    % (vvar[primary], vvar[primary]))
        if others:
            body.append("      delete from qm_tadqiq_lines where tadqiq_id = v_t;")
            body.extend("  " + x for x in lines_for(primary))
        body.append("    end if;" if rr["occ"] > 0 else "    end;")
        for v in others:
            body.append("    delete from qm_tadqiq where kashef_id = v_k and serial_no = '%s' %s and note = '%s' "
                        "and not opening and vendor_id = %s;"
                        % (esc(rr["serial"]), date_where, esc(note), vvar[v]))
            body.append("    insert into qm_tadqiq (kashef_id, vendor_id, tadqiq_date, street_no, note, opening, serial_no)\n"
                        "    values (v_k, %s, %s, '', '%s', false, '%s') returning id into v_t;"
                        % (vvar[v], date_expr, esc(note), esc(rr["serial"])))
            body.extend(lines_for(v))
        body.append("  end if;")

    # ── split into ≤ ~600 KB parts, each a self-contained do block, cut only
    #    between requests (BRIEF §3: the SQL editor refuses pastes over ~1 MB)
    LIMIT = 600 * 1024
    chunks, cur, size = [], [], 0
    for blk in body:
        b = len(blk.encode()) + 1
        if cur and blk.startswith("\n  -- ── أمر عمل") and size + b > LIMIT:
            chunks.append(cur)
            cur, size = [], 0
        cur.append(blk)
        size += b
    if cur:
        chunks.append(cur)
    hdr = "\n".join(head)
    for old in glob.glob(mig_path.replace(".sql", "*.sql")):
        os.remove(old)
    for i, ch in enumerate(chunks, 1):
        path = mig_path.replace(".sql", "_part%d.sql" % i) if len(chunks) > 1 else mig_path
        out = (hdr.replace("do $qmexpwex$", "-- part %d of %d\ndo $qmexpwex$" % (i, len(chunks)))
               + "\n" + "\n".join(ch) + "\n\nend $qmexpwex$;\n")
        open(path, "w", encoding="utf-8").write(out)
        print("WROTE: %s (%.1f KB)" % (os.path.basename(path), len(out.encode()) / 1024))


if __name__ == "__main__":
    main()
