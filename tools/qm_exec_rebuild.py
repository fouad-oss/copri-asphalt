# -*- coding: utf-8 -*-
"""Rebuild the EXECUTED side (طلبات التدقيق opening balances) for the work
orders whose certificates now come from the دفعات الوزارة folders.

Why: the certificate rebuild (0041) re-sourced 17 work orders from the
payment folders so they tie to the ministry tracker, but their executed
quantities were left on the older, stale كشف حساب totals. That leaves
certified > executed — impossible — on 8 closed work orders, KD 1.5M in
total, and hides them from the "finished, awaiting certification" list.

This regenerates those work orders' opening entries from the SAME folder
sheets, and because those sheets are per subcontractor the result is
attributed correctly per vendor instead of being one lump per work order.
Allocations are then resynced to match (the backfill's Option B: allocated
= executed), so executed / allocated / certified all come from one source.

Untouched: work orders whose كشف حساب already ties to the tracker (their
executed is right), and the flagged ones in paycert-rebuild-report.md
(both sides are stale together — the QA's worklist).

Output: supabase/migrations/0045_qm_exec_rebuild.sql (+ a report)
"""
import glob
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from qm_backfill import (  # noqa: E402
    classify_sheet, parse_company_sheet, payment_folders, wo_key_of_folder, norm_suffix,
)

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

OUT_DIR = r"C:\Users\fszog\Desktop\quantities-backfill"
MAP_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "qm-backfill-map.json")
SQL_OUT = r"C:\Users\fszog\Desktop\Copri webapp\supabase\migrations\0045_qm_exec_rebuild.sql"
BOP = json.load(open(os.path.join(OUT_DIR, "wo-ocr", "_bop_ref.json"), encoding="utf-8"))
PCT = 1.09

# Company → vendors.id AFTER the 0037 canonicalisation (490→9, 498→10,
# 525→11, 500→12, 494→13, 468→14, 536→15, 538→16; the rest kept their id).
REMAP = {490: 9, 498: 10, 525: 11, 500: 12, 494: 13, 468: 14, 536: 15, 538: 16}
COPRI_VENDOR = 591          # كوبري — تنفيذ ذاتي (CCC folds in here too)


def vendor_map():
    m = json.load(open(MAP_PATH, encoding="utf-8"))
    out = {}
    for name, spec in m.get("subs", {}).items():
        if not spec.get("active", True):
            continue
        if spec.get("mergeInto") or spec.get("vendorId") is None:
            out[name] = COPRI_VENDOR       # كوبري + CCC
            continue
        vid = spec["vendorId"]
        out[name] = REMAP.get(vid, vid)
    out.setdefault("كوبري", COPRI_VENDOR)
    return out


def key_of(bab, band, suf):
    return f"{bab}/{band}{norm_suffix(suf)}"


def value_of(lines):
    return sum((BOP.get(k, {}).get("rate") or 0) * q for k, q in lines.items())


def parse_folders_by_company():
    """→ {wo: {company: {key: qty}}} cumulative, copied-forward months dropped."""
    aliases = json.load(open(MAP_PATH, encoding="utf-8")).get("aliases", {})
    raw = {}
    for fdate, fpath in payment_folders():
        iso = fdate.isoformat()
        print(f"  {os.path.basename(fpath)}")
        for sub in sorted(os.listdir(fpath)):
            spath = os.path.join(fpath, sub)
            if not os.path.isdir(spath):
                continue
            wk = wo_key_of_folder(sub, aliases)
            if wk is None or wk[0] != "wo":
                continue
            books = [b for b in glob.glob(os.path.join(spath, "جميع الشركات*.xls*"))
                     if not os.path.basename(b).startswith("~$")]
            if not books:
                continue
            try:
                wb = openpyxl.load_workbook(books[0], data_only=True, read_only=True)
            except Exception:
                continue
            slots = {}
            for nm in wb.sheetnames:
                kind, comp, slot = classify_sheet(nm)
                if slot and comp:
                    slots[slot] = comp
            for nm in wb.sheetnames:
                kind, comp, slot = classify_sheet(nm)
                if kind != "plain":
                    continue
                if comp is None and slot is not None:
                    comp = slots.get(slot)
                if comp is None:
                    continue
                try:
                    parsed = parse_company_sheet(wb[nm])
                except Exception:
                    continue
                if not parsed:
                    continue
                tgt = raw.setdefault(wk[1], {}).setdefault(comp, {}).setdefault(iso, {})
                for (bab, band, suf), q in parsed.items():
                    k = key_of(bab, band, suf)
                    tgt[k] = tgt.get(k, 0) + q
            wb.close()

    out, last_month = {}, {}
    for wo, comps in raw.items():
        for comp, months in comps.items():
            prev = None
            for iso in sorted(months):
                cur = {k: round(v, 6) for k, v in months[iso].items() if v}
                if prev is not None and cur == prev:
                    continue                       # copied forward, not new work
                prev = cur
                agg = out.setdefault(wo, {}).setdefault(comp, {})
                for k, q in cur.items():
                    agg[k] = agg.get(k, 0) + q
                if iso > last_month.get(wo, ""):
                    last_month[wo] = iso
    return out, last_month


def main():
    blended = json.load(open(os.path.join(OUT_DIR, "paycerts-blended.json"), encoding="utf-8"))
    targets = {int(w) for w, d in blended["decisions"].items() if d["pick"] == "folders"}
    print(f"work orders to rebuild: {len(targets)} → {sorted(targets)}")

    print("parsing دفعات الوزارة …")
    by_comp, last_month = parse_folders_by_company()
    vmap = vendor_map()

    plan, unknown = {}, {}
    for wo in sorted(targets):
        comps = by_comp.get(wo)
        if not comps:
            continue
        for comp, lines in comps.items():
            vid = vmap.get(comp)
            if vid is None:
                unknown.setdefault(comp, []).append(wo)
                continue
            tgt = plan.setdefault(wo, {}).setdefault(vid, {})
            for k, q in lines.items():          # فتيح + المثنى both land on 9
                if k in BOP and q:
                    tgt[k] = tgt.get(k, 0) + q

    def q(s):
        return "'" + str(s).replace("'", "''") + "'"

    L = ["-- ════════════════════════════════════════════════════════════════════",
         "-- 0045 — QUANTITIES: rebuild the executed side for the work orders whose",
         "--        certificates were re-sourced from دفعات الوزارة in 0041.",
         "--",
         "--   Those work orders' certificates now tie to the ministry tracker, but",
         "--   their executed quantities were still the older كشف حساب totals — so",
         "--   certified exceeded executed, which cannot happen, and the finished-",
         "--   awaiting-certification list could not see them.",
         "--",
         "--   Here the opening طلبات تدقيق for these work orders are rebuilt from",
         "--   the same folder sheets, PER SUBCONTRACTOR (the sheets are one per",
         "--   company), and allocations are resynced to match. Executed, allocated",
         "--   and certified then all come from one source.",
         "--",
         "--   Only opening (رصيد افتتاحي) entries are touched — anything the QA",
         "--   records in the app is left alone. Idempotent.",
         "-- ════════════════════════════════════════════════════════════════════",
         "do $qmex$",
         "declare",
         "  v_contract bigint;",
         "  v_k bigint;",
         "  v_t bigint;",
         "  v_item bigint;",
         "  v_line bigint;",
         "  v_wos int := 0;",
         "begin",
         "  select id into v_contract from qm_contracts where code = 'HAW9';",
         "  if v_contract is null then raise exception 'run 0033 first'; end if;"]

    n_lines = 0
    for wo in sorted(plan):
        date = last_month.get(wo) or "2026-08-05"
        L.append(f"\n  -- ══ أمر عمل {wo} — إعادة بناء المنفَّذ من دفعات الوزارة ══")
        L.append(f"  select id into v_k from qm_kashefs where contract_id = v_contract and kashef_no = {wo};")
        L.append("  if v_k is not null then")
        L.append("    v_wos := v_wos + 1;")
        # drop the previous opening balances for this WO (lines cascade)
        L.append("    delete from qm_tadqiq where kashef_id = v_k and opening;")
        for vid, lines in sorted(plan[wo].items()):
            if not lines:
                continue
            L.append(f"    select id into v_item from vendors where id = {vid} and qm_subcontractor;")
            L.append(f"    if v_item is null then raise exception 'vendor {vid} missing or not flagged'; end if;")
            L.append("    insert into qm_tadqiq (kashef_id, vendor_id, tadqiq_date, serial_no, street_no, note, opening)")
            L.append(f"    values (v_k, {vid}, {q(date)}, '', '', "
                     f"{q('رصيد افتتاحي — من دفعات الوزارة')}, true) returning id into v_t;")
            for k, qty in sorted(lines.items()):
                item = BOP[k]
                sfx = item["suffix"] or ""
                L.append(f"    select id into v_item from qm_bop_items where contract_id = v_contract "
                         f"and bab = {item['bab']} and band = {item['band']} "
                         f"and translate(coalesce(suffix,''), 'أإآ', 'ااا') = {q(sfx)};")
                L.append(f"    if v_item is null then raise exception 'bop {k} missing'; end if;")
                L.append("    insert into qm_tadqiq_lines (tadqiq_id, bop_item_id, qty, out_of_kashef, over_allocation)")
                L.append(f"    values (v_t, v_item, {round(qty, 3)}, "
                         "not exists (select 1 from qm_kashef_lines where kashef_id = v_k and bop_item_id = v_item), "
                         "false);")
                n_lines += 1
            L.append("    insert into qm_changelog (entity, entity_id, action, field, line_ref, old_value, new_value, actor_email)")
            L.append(f"    values ('tadqiq', v_k, 'tadqiq_create', '', '', '', "
                     f"{q('رصيد افتتاحي معاد بناؤه — ' + str(len(lines)) + ' بند')}, 'folders-rebuild');")
        L.append("  end if;")

    L += ["",
          "  -- ── resync allocations for the rebuilt work orders (allocated = executed) ──",
          "  delete from qm_allocations a using qm_kashef_lines kl, qm_kashefs k",
          "   where a.kashef_line_id = kl.id and kl.kashef_id = k.id",
          f"     and k.contract_id = v_contract and k.kashef_no in ({', '.join(str(w) for w in sorted(plan))});",
          "",
          "  insert into qm_allocations (kashef_line_id, vendor_id, qty)",
          "  select kl.id, t.vendor_id, sum(tl.qty)",
          "  from qm_tadqiq t",
          "  join qm_tadqiq_lines tl on tl.tadqiq_id = t.id",
          "  join qm_kashef_lines kl on kl.kashef_id = t.kashef_id and kl.bop_item_id = tl.bop_item_id",
          "  join qm_kashefs k on k.id = t.kashef_id",
          f"  where k.contract_id = v_contract and k.kashef_no in ({', '.join(str(w) for w in sorted(plan))})",
          "  group by kl.id, t.vendor_id",
          "  on conflict (kashef_line_id, vendor_id) do update set qty = excluded.qty;",
          "",
          "  raise notice 'work orders rebuilt: %', v_wos;",
          "end $qmex$;"]

    open(SQL_OUT, "w", encoding="utf-8").write("\n".join(L) + "\n")

    # ── report ──────────────────────────────────────────────────────────
    dec = blended["decisions"]
    R = ["# Executed-side rebuild (0045)\n",
         f"Rebuilt work orders: **{len(plan)}** — those whose certificates were re-sourced "
         "from دفعات الوزارة in 0041, so both sides now come from the same sheets.\n",
         "| WO | executed before | executed after | certified | per-subcontractor split |",
         "|---|---|---|---|---|"]
    for wo in sorted(plan):
        after = sum(value_of(l) for l in plan[wo].values()) * PCT
        before = dec.get(str(wo), {}).get("ours", 0) * PCT
        cert = dec.get(str(wo), {}).get("folders", 0) * PCT
        split = ", ".join(f"{vid}:{value_of(l) * PCT:,.0f}" for vid, l in sorted(plan[wo].items()))
        R.append(f"| {wo} | {before:,.3f} | {after:,.3f} | {cert:,.3f} | {split} |")
    tot_after = sum(sum(value_of(l) for l in v.values()) for v in plan.values()) * PCT
    R.append(f"\n**Executed across the rebuilt work orders: KD {tot_after:,.3f} after 9%, "
             f"now attributed across {len({v for p in plan.values() for v in p})} subcontractors "
             f"instead of one lump per work order.**")
    if unknown:
        R += ["", "## Company names with no vendor mapping (skipped)\n"]
        R += [f"- {c}: work orders {sorted(set(w))}" for c, w in unknown.items()]
    open(os.path.join(OUT_DIR, "exec-rebuild-report.md"), "w", encoding="utf-8").write("\n".join(R) + "\n")

    print(f"\nSQL: {os.path.basename(SQL_OUT)} ({os.path.getsize(SQL_OUT):,} bytes)")
    print(f"work orders {len(plan)}, tadqiq lines {n_lines}, "
          f"executed KD {tot_after:,.3f} after 9%")
    if unknown:
        print("unmapped companies:", {k: sorted(set(v)) for k, v in unknown.items()})


if __name__ == "__main__":
    main()
