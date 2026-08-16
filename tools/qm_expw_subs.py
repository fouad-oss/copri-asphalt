# -*- coding: utf-8 -*-
"""Expressway subcontractor split — دفعات مقاولي الباطن.

Reads the subcontractor claim workbooks and works out, per work order and
per BOP item, how much of the work each subcontractor did. Emits the
per-sub ALLOCATION split (migration 0054).

Like tools/import_pos.py and tools/qm_backfill.py, this ALWAYS writes the
report + dataset but REFUSES to emit SQL until "confirmed": true is set in
tools/qm-expw-subs-map.json — created on first run with a proposed
sub → vendor mapping for Fouad to correct. Two things need his answer:

  1. WHICH VENDOR each folder is. Some map to vendors that already exist
     from the Hawalli work (بحر الابداع, دالكو); most are new to this
     contract. `قشط الاسفلت` is a WORK TYPE ("asphalt milling"), not a
     company name, so it cannot be guessed at all.
  2. Whether the EXECUTED tier should be split too — see below.

── what this migration does and does not touch ─────────────────────────
It rewrites ALLOCATIONS only (who is assigned each work-order line), which
is unambiguous and non-destructive. It does NOT touch طلبات التدقيق: the
1,090 imported requests carry real numbers and dates but no subcontractor,
so executed stays on «كوبري — تنفيذ ذاتي».

The consequence is that qm_sub_totals will show each subcontractor with an
allocation and zero executed. Fixing that means REPLACING the 1,090 dated
requests with per-sub opening balances (the Hawalli 0045 model) — which
buys correct per-sub executed totals at the cost of the real request
history. That trade is Fouad's call, not this script's.

── source shape ────────────────────────────────────────────────────────
One workbook per subcontractor per payment, named `<sub>-<N>.xlsm`, with
ONE SHEET PER WORK ORDER (sheet name = the WO number) plus 'رئيسي' and
sometimes 'المواد'. Each WO sheet:
    باب / بند | الوصف | الوحدة | الكمية السابقة | الكمية الحالية |
    الكمية الاجمالية | سعر الوحدة | الاجمالي د.ك
`الكمية الاجمالية` is that subcontractor's CUMULATIVE quantity on that work
order as at that payment.

Gotchas confirmed here:
  * In-sheet headers are STALE COPIES — a دالكو file cites the *Hawalli*
    contract number and a سكوير file names بحر الابداع as the contractor.
    Identity comes from the file path only.
  * Take the latest file THAT MENTIONS EACH WORK ORDER, not simply the
    latest file: a sub's work order can appear in an early claim and be
    absent from the newest one. Using the newest file alone lost KD 2.7M.
  * Codes are band-first, same as everywhere else — verified here against
    the BOP descriptions: 7,152 matches and zero mismatches for band-first
    against 172/1,670 for bab-first.
  * The claim rate usually equals the BOP rate (6,057 of 6,252 lines) but
    not always — subcontractors are sometimes on negotiated rates. Values
    in this report are therefore computed at BOP rates, which measures the
    share of CONTRACT value executed, not what the sub is owed.
"""
import sys, io, os, re, json, glob, unicodedata, collections

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DESK = os.path.join(os.path.expanduser("~"), "Desktop")
SUBDIR = os.path.join(DESK, "ExpresswaysQMbackfill", "دفعات مقاولي الباطن")
OUT_DIR = os.path.join(DESK, "quantities-backfill")
MAP_PATH = os.path.join(HERE, "qm-expw-subs-map.json")
MIG = os.path.join(REPO, "supabase", "migrations", "0054_qm_expw_subs.sql")
COPRI = "كوبري — تنفيذ ذاتي"
AR = "اأإآبتثجحخدذرزسشصضطظعغفقكلمنهةوىي"

# vendors that already exist from the Hawalli contract (see 0037)
KNOWN_VENDORS = {"بحر الابداع": 10, "دالكو": 493}


def clean(v):
    if v is None:
        return ""
    s = "".join(ch for ch in str(v) if unicodedata.category(ch) != "Cf")
    return re.sub(r"\s+", " ", s.replace("\xa0", " ").replace("ـ", "")).strip()


def esc(s):
    return str(s).replace("'", "''")


def num(v):
    if float(v) == int(float(v)):
        return str(int(float(v)))
    return repr(round(float(v), 4))


def parse_code(s):
    s = clean(s)
    nums = re.findall(r"\d+", s)
    lets = re.findall("[" + AR + "]+", s)
    if len(nums) < 2:
        return None
    return (int(nums[1]), int(nums[0]),
            re.sub("[أإآ]", "ا", lets[0])[:1] if lets else "")


def read_claims(bop):
    files = [p for p in glob.glob(os.path.join(SUBDIR, "**", "*.xls*"), recursive=True)
             if not os.path.basename(p).startswith("~$")]
    groups = collections.defaultdict(list)
    for p in files:
        parts = os.path.relpath(p, SUBDIR).split(os.sep)
        sub = parts[0] if len(parts) > 1 else os.path.splitext(parts[0])[0]
        if len(parts) > 2:
            sub = parts[0] + " / " + parts[1]
        base = os.path.splitext(os.path.basename(p))[0]
        m = re.search(r"(\d+)\s*$", base)
        groups[sub].append((int(m.group(1)) if m else 0, p, base))

    claims, source, unres = {}, {}, collections.Counter()
    for sub in sorted(groups):
        for _, p, base in sorted(groups[sub]):
            wb = openpyxl.load_workbook(p, data_only=True)
            for name in wb.sheetnames:
                sn = clean(name)
                if not re.match(r"^\d+$", sn):
                    continue
                wo = int(sn)
                rows = [list(r) for r in wb[name].iter_rows(max_col=12, values_only=True)]
                hdr = next((i for i, row in enumerate(rows[:8])
                            if any(clean(v).startswith("باب") for v in row)), None)
                if hdr is None:
                    continue
                got = {}
                for row in rows[hdr + 1:]:
                    k = parse_code(row[0] if row else None)
                    if not k:
                        continue
                    try:
                        cum = float(row[5])
                    except (TypeError, ValueError, IndexError):
                        continue
                    if cum <= 0:
                        continue
                    if k not in bop:
                        unres[k] += 1
                        continue
                    got[k] = cum
                if got:
                    # the newest claim MENTIONING this work order replaces
                    # whatever an earlier one said about it
                    for key in [x for x in claims if x[0] == sub and x[1] == wo]:
                        claims.pop(key)
                    for k, q in got.items():
                        claims[(sub, wo, k)] = q
                    source[(sub, wo)] = base
            wb.close()
    return groups, claims, source, unres


def main():
    bop = {}
    for it in json.load(open(os.path.join(OUT_DIR, "expw-bop-data.json"),
                             encoding="utf-8")):
        bop[(it["bab"], it["band"], it["suffix"] or "")] = it
    wos = {w["wo"]: w for w in json.load(
        open(os.path.join(OUT_DIR, "expw-wo-data.json"), encoding="utf-8"))}
    woline = {}
    for w in wos.values():
        for l in w["lines"]:
            woline[(w["wo"], (l["bab"], l["band"], l["suffix"] or ""))] = l["qty"]

    groups, claims, source, unres = read_claims(bop)

    # ── the mapping file ──
    if os.path.exists(MAP_PATH):
        cfg = json.load(open(MAP_PATH, encoding="utf-8"))
    else:
        cfg = {"confirmed": False,
               "_help": "Set vendor_id for each folder (an existing vendors.id) "
                        "or vendor_name to have the row created. Set "
                        "confirmed=true to let the tool emit SQL.",
               "subs": {}}
    for sub in sorted(groups):
        entry = cfg["subs"].setdefault(sub, {})
        entry.setdefault("vendor_id", next(
            (v for k, v in KNOWN_VENDORS.items() if k in sub), None))
        entry.setdefault("vendor_name", None)
        entry.setdefault("note", "")
    json.dump(cfg, open(MAP_PATH, "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    # ── fold folders onto vendors BEFORE splitting ──────────────────
    # Several folders can be the same vendor (بحر الابداع is three: مدني,
    # نظافة and نظافة/العقد الجديد; قصر البيداء is two), and they overlap on
    # the same work orders. Allocations are keyed (kashef_line, vendor), so
    # folder-by-folder writes would OVERWRITE rather than sum. Aggregate on
    # the vendor key first.
    def vendor_key(sub):
        e = cfg["subs"].get(sub, {})
        if e.get("vendor_id"):
            return ("id", e["vendor_id"])
        if e.get("vendor_name"):
            return ("name", e["vendor_name"])
        return ("folder", sub)

    by_vendor = collections.defaultdict(float)    # (vkey, wo, k) -> qty
    folders_of = collections.defaultdict(set)
    for (sub, wo, k), q in claims.items():
        vk = vendor_key(sub)
        by_vendor[(vk, wo, k)] += q
        folders_of[vk].add(sub)

    # ── allocation split, capped at the work-order line quantity ──
    per_pair = collections.defaultdict(dict)      # (wo,key) -> vkey -> qty
    for (vk, wo, k), q in by_vendor.items():
        per_pair[(wo, k)][vk] = q
    alloc, capped, orphan = {}, [], []
    for (wo, k), vends in per_pair.items():
        line_qty = woline.get((wo, k))
        if line_qty is None:
            orphan.append((wo, k, sum(vends.values())))
            continue
        total = sum(vends.values())
        scale = 1.0
        if total > line_qty * 1.001:
            scale = line_qty / total
            capped.append((wo, k, total, line_qty))
        for vk, q in vends.items():
            alloc[(vk, wo, k)] = q * scale

    per_sub_val = collections.defaultdict(float)
    for (vk, wo, k), q in alloc.items():
        per_sub_val[vk] += q * bop[k]["rate"]
    total_alloc = sum(per_sub_val.values())
    wo_total = sum(w["value_calc"] for w in wos.values())

    # COPRI keeps the remainder of every line
    line_assigned = collections.defaultdict(float)
    for (sub, wo, k), q in alloc.items():
        line_assigned[(wo, k)] += q
    copri = {}
    for (wo, k), qty in woline.items():
        rem = qty - line_assigned.get((wo, k), 0.0)
        if rem > 1e-6:
            copri[(wo, k)] = rem
    copri_val = sum(q * bop[k]["rate"] for (wo, k), q in copri.items())

    # ── report ──
    rep = ["# Expressway subcontractors — claim extraction\n"]
    rep.append("Source: `دفعات مقاولي الباطن`, one workbook per subcontractor per "
               "payment, one sheet per work order.\n")
    rep.append("## Who claimed what\n")
    rep.append("Values are at **BOP rates** — the share of contract value each "
               "subcontractor executed, NOT what they are owed (the claim sheets "
               "sometimes carry negotiated rates: they equal the BOP rate on "
               "6,057 of 6,252 lines).\n")
    rep.append("| vendor | folders | work orders | allocated value pre-pct | share |")
    rep.append("|---|---|---|---|---|")
    for vk in sorted(per_sub_val, key=lambda s: -per_sub_val[s]):
        ws_ = sorted({w for (v2, w, _) in alloc if v2 == vk})
        label = ("vendors.id %s" % vk[1]) if vk[0] == "id" else vk[1]
        rep.append("| %s | %s | %s | %s | %.1f%% |"
                   % (label, " + ".join(sorted(folders_of[vk])),
                      ", ".join(str(x) for x in ws_),
                      "{:,.3f}".format(per_sub_val[vk]),
                      100 * per_sub_val[vk] / wo_total))
    rep.append("| **%s** (remainder) | — | — | **%s** | **%.1f%%** |"
               % (COPRI, "{:,.3f}".format(copri_val), 100 * copri_val / wo_total))
    rep.append("\n- subcontracted: **KD %s** · self-performed: **KD %s** · "
               "work-order total: **KD %s** (pre-pct)."
               % ("{:,.3f}".format(total_alloc), "{:,.3f}".format(copri_val),
                  "{:,.3f}".format(wo_total)))
    rep.append("- unresolved BOP codes in the claims: %d" % len(unres))

    rep.append("\n## Claims that had to be capped (%d)\n" % len(capped))
    if capped:
        rep.append("Σ of the subs' claims exceeded the work-order line, so they "
                   "were scaled down pro rata to fit.\n")
        rep.append("| WO | bab/band | claimed | work-order line | ratio |")
        rep.append("|---|---|---|---|---|")
        for wo, k, tot, line in sorted(capped, key=lambda x: -(x[2] / x[3])):
            rep.append("| %d | %d/%d%s | %s | %s | ×%.2f |"
                       % (wo, k[0], k[1], k[2] or "", "{:,.2f}".format(tot),
                          "{:,.2f}".format(line), tot / line))

    rep.append("\n## Claims for items not in the work order (%d) — skipped\n"
               % len(orphan))
    for wo, k, q in orphan:
        rep.append("- WO %d — bab %d band %d%s, qty %s"
                   % (wo, k[0], k[1], k[2] or "", "{:,.2f}".format(q)))

    rep.append("\n## Vendor mapping\n")
    rep.append("From `tools/qm-expw-subs-map.json` (confirmed: **%s**).\n"
               % cfg.get("confirmed"))
    rep.append("| folder | vendor | note |")
    rep.append("|---|---|---|")
    for sub in sorted(groups):
        e = cfg["subs"][sub]
        if e.get("vendor_id"):
            who = "`vendors.id %s`" % e["vendor_id"]
        elif e.get("vendor_name"):
            who = "%s *(by name)*" % e["vendor_name"]
        else:
            who = "**unmapped**"
        rep.append("| %s | %s | %s |" % (sub, who, e.get("note", "")))

    rep.append("\n## The executed tier — NEEDS FOUAD\n")
    rep.append("This split covers ALLOCATIONS only. طلبات التدقيق stay on "
               "«%s» because the request sheets name no subcontractor, so "
               "qm_sub_totals will show these vendors with an allocation and "
               "zero executed.\n" % COPRI)
    rep.append("Splitting executed too means replacing the 1,090 imported "
               "requests — which carry real numbers and dates — with per-sub "
               "opening balances, the model the Hawalli backfill used. That "
               "buys correct per-sub executed totals and loses the request "
               "history. Say which you prefer.\n")

    report = "\n".join(rep) + "\n"
    open(os.path.join(OUT_DIR, "expw-subs-report.md"), "w",
         encoding="utf-8").write(report)
    json.dump({"%s|%d|%d|%d|%s" % (s, w, k[0], k[1], k[2] or ""): round(q, 4)
               for (s, w, k), q in alloc.items()},
              open(os.path.join(OUT_DIR, "expw-subs-data.json"), "w",
                   encoding="utf-8"), ensure_ascii=False, indent=1)
    print(report[:3000])

    unmapped = [s for s in groups
                if not cfg["subs"][s]["vendor_id"] and not cfg["subs"][s]["vendor_name"]]
    if not cfg.get("confirmed") or unmapped:
        print("\nNO SQL EMITTED.")
        print("  confirmed: %s" % cfg.get("confirmed"))
        print("  unmapped folders: %d %s" % (len(unmapped), unmapped))
        print("  edit %s then re-run." % MAP_PATH)
        return

    # ── SQL (only once confirmed) ──
    body = ["""-- 0054_qm_expw_subs — GENERATED by tools/qm_expw_subs.py, do not hand-edit.
-- Per-subcontractor ALLOCATION split for the Expressway, from
-- دفعات مقاولي الباطن. Replaces the flat allocated:=executed rows that
-- 0051 seeded on «%s». Executed (طلبات التدقيق) is NOT touched.
-- Subcontracted KD %s of KD %s pre-pct; the remainder stays with COPRI.
-- Idempotent. Paste after 0051.
do $qmexpwsub$
declare
  v_contract bigint;
  v_k bigint;
  v_item bigint;
  v_line bigint;
  v_vendor bigint;
  v_copri bigint;
begin
  select id into v_contract from qm_contracts where code = 'EXPW';
  if v_contract is null then raise exception 'run 0047 first'; end if;
  select id into v_copri from vendors where name = '%s';
  if v_copri is null then raise exception 'COPRI pseudo-vendor missing'; end if;
""" % (esc(COPRI), "{:,.3f}".format(total_alloc), "{:,.3f}".format(wo_total),
       esc(COPRI))]

    for vk in sorted(per_sub_val, key=lambda s: -per_sub_val[s]):
        label = " + ".join(sorted(folders_of[vk]))
        if vk[0] == "id":
            # fail loudly rather than silently allocating to the wrong row
            body.append("""
  -- %s
  select id into v_vendor from vendors where id = %d;
  if v_vendor is null then raise exception 'vendor %d not found (%s)'; end if;"""
                        % (esc(label), vk[1], vk[1], esc(label)))
        else:
            body.append("""
  -- %s
  select id into v_vendor from vendors where name = '%s';
  if v_vendor is null then
    insert into vendors (name, kind, notes) values ('%s', 'subcontractor',
      'quantities module — Expressway subcontractor (backfill)')
    returning id into v_vendor;
  end if;""" % (esc(label), esc(vk[1]), esc(vk[1])))
        body.append("  update vendors set qm_subcontractor = true where id = "
                    "v_vendor and not qm_subcontractor;")
        for (v2, wo, k), q in sorted(alloc.items(), key=lambda x: (x[0][1], x[0][2])):
            if v2 != vk:
                continue
            body.append("""  select id into v_k from qm_kashefs where contract_id = v_contract and kashef_no = %d;
  select id into v_item from qm_bop_items where contract_id = v_contract and bab = %d and band = %d and coalesce(suffix,'') = '%s';
  select id into v_line from qm_kashef_lines where kashef_id = v_k and bop_item_id = v_item;
  if v_line is not null then
    insert into qm_allocations (kashef_line_id, vendor_id, qty) values (v_line, v_vendor, %s)
      on conflict (kashef_line_id, vendor_id) do update set qty = excluded.qty;
  end if;""" % (wo, k[0], k[1], esc(k[2] or ""), num(q)))

    body.append("\n  -- COPRI keeps the remainder of every line")
    for (wo, k), q in sorted(copri.items()):
        body.append("""  select id into v_k from qm_kashefs where contract_id = v_contract and kashef_no = %d;
  select id into v_item from qm_bop_items where contract_id = v_contract and bab = %d and band = %d and coalesce(suffix,'') = '%s';
  select id into v_line from qm_kashef_lines where kashef_id = v_k and bop_item_id = v_item;
  if v_line is not null then
    insert into qm_allocations (kashef_line_id, vendor_id, qty) values (v_line, v_copri, %s)
      on conflict (kashef_line_id, vendor_id) do update set qty = excluded.qty;
  end if;""" % (wo, k[0], k[1], esc(k[2] or ""), num(q)))
    body.append("\nend $qmexpwsub$;\n")

    out = "\n".join(body)
    open(MIG, "w", encoding="utf-8").write(out)
    print("\nWROTE: %s (%.1f KB)" % (os.path.basename(MIG), len(out.encode()) / 1024))


if __name__ == "__main__":
    main()
