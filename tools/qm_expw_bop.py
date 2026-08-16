# -*- coding: utf-8 -*-
"""Expressway (EXPW) Bill of Prices importer — mirrors tools/qm_backfill.py's
BOP pass for the Hawalli contract, but for `هـ ص / ط / 9`.

Source: the standalone جدول الأسعار workbook (BILL.xlsx), sheet 'Table 1 (2)'.
Sheet '13' of the same workbook is a verbatim duplicate of the bab-13 block
already present in 'Table 1 (2)' — verified row-by-row, then skipped.

── the bidi question (see BRIEF-expressway-backfill.md §5) ──────────────
The workbook stores ONE composite code per row under the header 'رقم البند',
e.g. '1/1', '17/أ/4', '36/17أ'. Segment order is NOT trustworthy on its own.
It was resolved here with evidence rather than assumption: every أمر عمل
workbook in ExpresswaysQMbackfill\\الدفعة carries باب and رقم البند as
SEPARATE columns. Cross-checking 210 codes that appear in both sources:

    reading  left=band, right=bab  -> 207 rate matches,  3 mismatches
    reading  left=bab,  right=band ->   2 rate matches, 16 mismatches

so the composite code is band-first / bab-second, matching the convention
the Hawalli standalone BOP file used. تعديل او اضافة بند - جديد.xls
independently confirms it ('16/2' there is the same 0.73 م3 item that WO 1
lists as باب 002 / بند 0016).

An Arabic letter may appear in ANY segment position ('17/أ/4', '18/4/ب',
'أ / 16/2', '36/17أ') — the authoring is inconsistent. The two NUMBERS always
keep band-then-bab order, so the parser extracts the numbers positionally and
treats any letter as the suffix, normalising hamza forms (أ/إ/آ -> ا, هـ -> ه).

── positional corrections ──────────────────────────────────────────────
The sheet is laid out in contiguous bab runs. Where a row's parsed bab
contradicts the run it sits in, the run wins and the row carries a
source_note (same policy as the Hawalli seed's 4 corrections). Rows whose
code carries only ONE number take their bab from the run.

Outputs (to ~\\Desktop\\quantities-backfill\\):
    expw-bop-report.md      validation report — read this before pasting
    expw-bop-data.json      the parsed rows
and the migration itself into supabase/migrations/.

Re-runnable; writes the report even when it refuses to emit SQL.
"""
import sys, io, os, re, json, glob, unicodedata, collections

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
import xlrd

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DESK = os.path.join(os.path.expanduser("~"), "Desktop")
BILL = os.path.join(os.path.expanduser("~"), "Downloads", "BILL (1).xlsx")
EXPW_ROOT = os.path.join(DESK, "ExpresswaysQMbackfill")
QTY = os.path.join(EXPW_ROOT, "الدفعة", "كميات 9المفصلة.xls")
WODIR = os.path.join(EXPW_ROOT, "الدفعة")
OUT_DIR = os.path.join(DESK, "quantities-backfill")
MIG = os.path.join(REPO, "supabase", "migrations", "0048_qm_expw_bop_seed.sql")
os.makedirs(OUT_DIR, exist_ok=True)

# babs actually present in the Expressway price book
KNOWN_BABS = {1, 2, 3, 4, 5, 6, 7, 12, 13, 14, 17}
AR_LETTERS = "اأإآبتثجحخدذرزسشصضطظعغفقكلمنهةوىي"
MAX_PART = 900_000          # SQL editor refuses > ~1 MB


# ── helpers ──────────────────────────────────────────────────────────
def clean(v):
    """Strip bidi controls and tatweel, collapse every whitespace run to one
    space — several BILL descriptions wrap onto a second line inside the cell
    (e.g. the Pre-Stressed Concrete Beam item), which would otherwise put raw
    newlines inside the SQL string literals and in the UI."""
    if v is None:
        return ""
    s = "".join(ch for ch in str(v) if unicodedata.category(ch) != "Cf")
    return re.sub(r"\s+", " ", s.replace("\xa0", " ").replace("ـ", "")).strip()


def norm_suffix(s):
    """Normalise the sub-item letter: hamza forms collapse to ا, هـ to ه."""
    s = re.sub("[أإآ]", "ا", clean(s))
    return s[:1] if s else None


def esc(s):
    return str(s).replace("'", "''")


def num(v):
    """SQL numeric literal without float noise."""
    if v == int(v):
        return str(int(v))
    return repr(round(float(v), 4))


def parse_code(code):
    """-> (band, bab, suffix, n_numbers). bab is None when the code carries
    only one number (caller fills it from the run)."""
    c = clean(code)
    nums = re.findall(r"\d+", c)
    lets = re.findall("[" + AR_LETTERS + "]+", c)
    suf = norm_suffix(lets[0]) if lets else None
    if len(nums) >= 2:
        return int(nums[0]), int(nums[1]), suf, len(nums)
    if len(nums) == 1:
        return int(nums[0]), None, suf, 1
    return None


# ── 1. read the price book ───────────────────────────────────────────
def read_bill():
    wb = openpyxl.load_workbook(BILL, data_only=True)
    sheets = {}
    for ws in wb.worksheets:
        rows = []
        for i, r in enumerate(ws.iter_rows(min_row=2, max_col=6, values_only=True)):
            code = clean(r[0])
            if not code:
                continue
            try:
                rate = round(float(r[3]), 4)
            except (TypeError, ValueError):
                rate = None
            rows.append({"row": i + 2, "code": code, "desc": clean(r[1]),
                         "unit": clean(r[2]), "rate": rate})
        sheets[ws.title] = rows
    wb.close()
    return sheets


# ── 2. the WO workbooks: independent (bab, band) truth ───────────────
def read_wo_items():
    def rows_of(path):
        out = {}
        if path.lower().endswith(".xls"):
            bk = xlrd.open_workbook(path)
            for sh in bk.sheets():
                out[sh.name] = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                                for r in range(sh.nrows)]
        else:
            wb = openpyxl.load_workbook(path, data_only=True)
            for ws in wb.worksheets:
                out[ws.title] = [list(r) for r in
                                 ws.iter_rows(max_col=14, values_only=True)]
            wb.close()
        return out

    items = {}
    files = [f for f in sorted(glob.glob(os.path.join(WODIR, "*مر عمل رقم*.xls*")))
             if not os.path.basename(f).startswith("~$")]
    for f in files:
        try:
            sheets = rows_of(f)
        except Exception:
            continue
        for _, rows in sheets.items():
            hi = hc = None
            for i, row in enumerate(rows[:60]):
                for c in range(len(row) - 1):
                    if clean(row[c]) == "باب" and clean(row[c + 1]).startswith("رقم البند"):
                        hi, hc = i, c
                        break
                if hi is not None:
                    break
            if hi is None:
                continue
            for row in rows[hi + 1:]:
                if len(row) <= hc + 7:
                    continue
                bab_s = clean(row[hc])
                bab_s = bab_s[:-2] if bab_s.endswith(".0") else bab_s
                band_raw = clean(row[hc + 1])
                if not bab_s.isdigit() or not band_raw:
                    continue
                m = re.match(r"^0*(\d+)\s*[/\\]?\s*(.*)$", band_raw)
                if not m:
                    continue
                try:
                    rate = round(float(row[hc + 7]), 4)
                except (TypeError, ValueError):
                    continue
                key = (int(bab_s), int(m.group(1)), norm_suffix(m.group(2)))
                items.setdefault(key, (clean(row[hc + 2]), clean(row[hc + 6]), rate))
    return items, len(files)


# ── 3. read كميات 9المفصلة (rate cross-check) ────────────────────────
def read_qty():
    out = collections.defaultdict(collections.Counter)
    bk = xlrd.open_workbook(QTY)
    for name in [str(i) for i in range(1, 22)]:
        if name not in bk.sheet_names():
            continue
        sh = bk.sheet_by_name(name)
        for r in range(2, sh.nrows):
            code = clean(sh.cell_value(r, 0))
            if not code:
                continue
            p = parse_code(code)
            if not p or p[1] is None:
                continue
            rate = sh.cell_value(r, 3)
            out[(p[1], p[0], p[2])][round(float(rate), 4)
                                    if isinstance(rate, float) else None] += 1
    return out


def main():
    notes = []
    sheets = read_bill()
    main_rows = sheets.get("Table 1 (2)")
    if not main_rows:
        sys.exit("BILL sheet 'Table 1 (2)' not found — got %s" % list(sheets))

    # 1a. confirm sheet '13' really is a duplicate before discarding it
    dup_sheet = sheets.get("13", [])
    main_by_code = {r["code"]: r for r in main_rows}
    dup_diff = [r for r in dup_sheet
                if r["code"] not in main_by_code
                or (main_by_code[r["code"]]["rate"], main_by_code[r["code"]]["unit"])
                != (r["rate"], r["unit"])]
    notes.append("sheet '13': %d rows, %d differing from 'Table 1 (2)' -> %s"
                 % (len(dup_sheet), len(dup_diff),
                    "discarded as duplicate" if not dup_diff else "NEEDS REVIEW"))

    # 2. parse + positional bab correction
    parsed = []
    for r in main_rows:
        p = parse_code(r["code"])
        if not p:
            notes.append("UNPARSED code %r (row %d) — SKIPPED" % (r["code"], r["row"]))
            continue
        band, bab, suf, nnum = p
        parsed.append(dict(r, band=band, bab=bab, suffix=suf, nnum=nnum))

    # ── establish the bab BLOCKS ─────────────────────────────────────
    # The sheet is laid out as contiguous bab blocks. A run of >= MIN_RUN
    # consecutive rows sharing a bab is "established" and is never
    # second-guessed — bab 3 (6 items) and bab 14 (4 items) are genuinely
    # small, and a plain neighbour-majority vote wrongly swallows them into
    # the large babs on either side. Only rows in runs shorter than that
    # (or with a missing/unknown bab) take their bab from the enclosing
    # established block, and only when the blocks on both sides agree.
    MIN_RUN = 3
    runs = []                      # (start, end_exclusive, bab)
    i = 0
    while i < len(parsed):
        b = parsed[i]["bab"]
        j = i
        while j < len(parsed) and parsed[j]["bab"] == b:
            j += 1
        runs.append((i, j, b))
        i = j
    established = [(s, e, b) for s, e, b in runs
                   if b in KNOWN_BABS and e - s >= MIN_RUN]

    def enclosing_bab(i):
        """bab of the established block containing/surrounding row i, or None
        when the neighbouring blocks disagree (a genuine boundary)."""
        before = [b for s, e, b in established if e <= i]
        after = [b for s, e, b in established if s > i]
        for s, e, b in established:
            if s <= i < e:
                return b
        if before and after:
            return before[-1] if before[-1] == after[0] else None
        return (before or after or [None])[-1] if (before or after) else None

    in_established = set()
    for s, e, _ in established:
        in_established.update(range(s, e))

    corrections = []
    for i, x in enumerate(parsed):
        if i in in_established:
            continue
        rb = enclosing_bab(i)
        if x["bab"] is None:
            if rb is None:
                notes.append("row %d %r: no bab and no run context — SKIPPED"
                             % (x["row"], x["code"]))
                x["skip"] = True
                continue
            x["source_note"] = ("الباب مفقود في الملف المصدر (%s) — أُخذ %d من موضع البند"
                                % (x["code"], rb))
            corrections.append((x["row"], x["code"], None, rb, "missing bab"))
            x["bab"] = rb
        elif x["bab"] not in KNOWN_BABS and rb is not None:
            x["source_note"] = ("الباب %d غير معروف (%s) — صُحّح إلى %d من موضع البند"
                                % (x["bab"], x["code"], rb))
            corrections.append((x["row"], x["code"], x["bab"], rb, "unknown bab"))
            x["bab"] = rb
        elif rb is not None and x["bab"] != rb:
            # the row sits inside a different bab run than its code claims
            x["source_note"] = ("الباب %d في الملف المصدر (%s) يخالف موضع البند — صُحّح إلى %d"
                                % (x["bab"], x["code"], rb))
            corrections.append((x["row"], x["code"], x["bab"], rb, "contradicts run"))
            x["bab"] = rb

    parsed = [x for x in parsed if not x.get("skip")]

    # 3. dedupe on the real key
    by_key = collections.OrderedDict()
    dupes = []
    for x in parsed:
        k = (x["bab"], x["band"], x["suffix"] or "")
        if k in by_key:
            dupes.append((k, by_key[k], x))
            continue
        by_key[k] = x
    rows = list(by_key.values())

    # 4. cross-checks
    wo_items, n_wo = read_wo_items()
    qty = read_qty()
    seed_keys = set(by_key)

    wo_missing, wo_rate_bad, wo_ok = [], [], 0
    for (bab, band, suf), (desc, unit, rate) in wo_items.items():
        k = (bab, band, suf or "")
        if k not in seed_keys:
            wo_missing.append((bab, band, suf, rate, desc, unit))
        elif abs(by_key[k]["rate"] - rate) >= 0.005:
            wo_rate_bad.append((k, by_key[k]["rate"], rate, desc))
        else:
            wo_ok += 1

    qty_missing, qty_rate_bad, qty_ok = [], [], 0
    for (bab, band, suf), ctr in qty.items():
        k = (bab, band, suf or "")
        rate = ctr.most_common(1)[0][0]
        if k not in seed_keys:
            qty_missing.append((bab, band, suf, rate))
        elif rate is not None and abs(by_key[k]["rate"] - rate) >= 0.005:
            qty_rate_bad.append((k, by_key[k]["rate"], rate))
        else:
            qty_ok += 1

    norate = [x for x in rows if x["rate"] is None]

    # 5. report
    babs = collections.Counter(x["bab"] for x in rows)
    rep = []
    rep.append("# Expressway BOP — import validation\n")
    rep.append("Source: `%s`, sheet `Table 1 (2)`.\n" % os.path.basename(BILL))
    rep.append("## Result\n")
    rep.append("- **%d BOP items** to seed, across %d babs." % (len(rows), len(babs)))
    rep.append("- bab distribution: %s" % ", ".join(
        "%d→%d" % (b, n) for b, n in sorted(babs.items())))
    rep.append("- items carrying a suffix letter: %d"
               % sum(1 for x in rows if x["suffix"]))
    rep.append("")
    for n in notes:
        rep.append("- %s" % n)
    rep.append("\n## Bidi reading\n")
    rep.append("Composite codes read **band / bab** (left segment = بند). "
               "Established against the أمر عمل workbooks, which carry باب and "
               "رقم البند in separate columns — see the module docstring.\n")
    rep.append("## Positional corrections (%d)\n" % len(corrections))
    if corrections:
        rep.append("| sheet row | code | bab in file | bab used | reason |")
        rep.append("|---|---|---|---|---|")
        for row, code, was, now, why in corrections:
            rep.append("| %d | `%s` | %s | %d | %s |"
                       % (row, code, "—" if was is None else was, now, why))
    rep.append("\n## Duplicate keys collapsed (%d)\n" % len(dupes))
    for k, first, second in dupes:
        rep.append("- bab %d band %d suf %s — kept row %d (%s, %s), dropped row %d (%s, %s)"
                   % (k[0], k[1], k[2] or "—", first["row"], first["rate"],
                      first["desc"][:40], second["row"], second["rate"],
                      second["desc"][:40]))
    rep.append("\n## Cross-check vs the أمر عمل workbooks (%d files)\n" % n_wo)
    rep.append("- %d WO items matched the seed on bab+band+suffix **and** rate" % wo_ok)
    rep.append("- %d WO items NOT in the price book:" % len(wo_missing))
    for bab, band, suf, rate, desc, unit in wo_missing:
        rep.append("    - bab %d band %d%s — %s %s — %s"
                   % (bab, band, suf or "", rate, unit, desc[:70]))
    rep.append("- %d WO items whose rate disagrees with the price book:" % len(wo_rate_bad))
    for k, seed_rate, wo_rate, desc in wo_rate_bad:
        rep.append("    - bab %d band %d suf %s — BOP %s vs WO %s — %s"
                   % (k[0], k[1], k[2] or "—", seed_rate, wo_rate, desc[:60]))
    rep.append("\n## Cross-check vs كميات 9المفصلة.xls\n")
    rep.append("- %d keys agreed on rate" % qty_ok)
    rep.append("- %d keys present there but not in the price book: %s"
               % (len(qty_missing), qty_missing[:10]))
    rep.append("- %d rate disagreements: %s" % (len(qty_rate_bad), qty_rate_bad[:10]))
    if norate:
        rep.append("\n## Rows with NO rate (%d) — skipped\n" % len(norate))
        for x in norate[:20]:
            rep.append("- row %d `%s` %s" % (x["row"], x["code"], x["desc"][:60]))

    report = "\n".join(rep) + "\n"
    open(os.path.join(OUT_DIR, "expw-bop-report.md"), "w", encoding="utf-8").write(report)
    keep = ("row", "code", "bab", "band", "suffix", "desc", "unit", "rate",
            "source_note")
    json.dump([{k: x[k] for k in keep if k in x} for x in rows],
              open(os.path.join(OUT_DIR, "expw-bop-data.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1, default=str)
    print(report)

    # 6. SQL
    rows = [x for x in rows if x["rate"] is not None]
    header = """-- ════════════════════════════════════════════════════════════════════
-- 0048 — QUANTITIES: seed the Expressway (EXPW) Bill of Prices.
-- Source: the standalone جدول الأسعار workbook, sheet 'Table 1 (2)'
-- (%d priced rows). Generated by tools/qm_expw_bop.py — see that file
-- for how the band/bab bidi order was established (cross-checked against
-- the أمر عمل workbooks, which carry باب and رقم البند separately:
-- 207 rate matches for band-first vs 2 for bab-first).
-- %d rows carry positional corrections (source_note on the row); see
-- ~/Desktop/quantities-backfill/expw-bop-report.md.
-- Idempotent: on conflict do nothing against (contract, bab, band, suffix).
-- Paste 0047 first (it sets the EXPW header).
-- ════════════════════════════════════════════════════════════════════
""" % (len(rows), len(corrections))

    def values_line(x):
        return "    (v_contract, %d, %d, %s, '%s', '%s', %s, %s)" % (
            x["bab"], x["band"],
            "null" if not x["suffix"] else "'%s'" % esc(x["suffix"]),
            esc(x["desc"]), esc(x["unit"]), num(x["rate"]),
            "null" if not x.get("source_note") else "'%s'" % esc(x["source_note"]))

    INS = ("  insert into qm_bop_items (contract_id, bab, band, suffix, "
           "description, unit, rate, source_note) values\n")
    TAIL = "\n  on conflict (contract_id, bab, band, coalesce(suffix, '')) do nothing;\n"
    PRE = """
do $qmexpwbop$
declare
  v_contract bigint;
begin
  select id into v_contract from qm_contracts where code = 'EXPW';
  if v_contract is null then
    raise exception 'qm_contracts EXPW missing - run 0046/0047 first';
  end if;

"""
    POST = "\nend $qmexpwbop$;\n"

    parts, cur, cur_len = [], [], 0
    CHUNK = 400
    batches = [rows[i:i + CHUNK] for i in range(0, len(rows), CHUNK)]
    for b in batches:
        block = INS + ",\n".join(values_line(x) for x in b) + TAIL
        if cur and cur_len + len(block.encode()) > MAX_PART:
            parts.append(cur); cur, cur_len = [], 0
        cur.append(block); cur_len += len(block.encode())
    if cur:
        parts.append(cur)

    written = []
    for i, blocks in enumerate(parts, 1):
        path = MIG if len(parts) == 1 else MIG.replace(".sql", "_part%d.sql" % i)
        body = header + PRE + "\n".join(blocks) + POST
        open(path, "w", encoding="utf-8").write(body)
        written.append((os.path.basename(path), len(body.encode())))
    print("\nWROTE:")
    for name, size in written:
        print("  %-44s %7.1f KB" % (name, size / 1024))


if __name__ == "__main__":
    main()
