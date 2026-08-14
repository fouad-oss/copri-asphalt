# -*- coding: utf-8 -*-
"""Audit D:\\التجميع الشهري new\\دفعات الوزارة — the folders the ministry
payment certificates are generated from — against the tracking report and
against what the module currently holds.

Per payment folder (a month) and per work-order sub-folder it reads the
«جميع الشركات» workbook, whose sheets are one per subcontractor and are
either PLAIN (that month's quantities) or «قديم+جديد» (cumulative to date).

Produces, per work order:
  · executed to date  = latest cumulative sheet set (قديم+جديد), falling
    back to the sum of the plain monthly sheets when no cumulative exists
  · a month-by-month series of the plain sheets
and compares that with
  · the tracking report's executed column   (متابعة أوامـر العمــل …-الجديد)
  · the module's executed (كشف حساب مجموع, what the WO backfill loaded)
  · the payment certificates built from the كشف حساب جزئي columns (0041)

Output: Desktop/quantities-backfill/payfolders-audit.md  (+ .json)
"""
import datetime
import glob
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from qm_backfill import (  # noqa: E402
    clean, norm_ar, norm_suffix, classify_sheet, parse_company_sheet,
    payment_folders, wo_key_of_folder, PAY_ROOT,
)

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

OUT_DIR = r"C:\Users\fszog\Desktop\quantities-backfill"
TRACK_DIR = r"D:\التجميع الشهري new\اوامر العمل"
MAP_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "qm-backfill-map.json")
BOP = json.load(open(os.path.join(OUT_DIR, "wo-ocr", "_bop_ref.json"), encoding="utf-8"))
PCT = 1.09


def val(lines):
    """KD value (after نسبة العقد) of a {(bab,band,suffix): qty} dict."""
    tot = 0.0
    for (bab, band, suf), qty in lines.items():
        r = BOP.get(f"{bab}/{band}{norm_suffix(suf)}", {}).get("rate")
        if r:
            tot += qty * r
    return tot * PCT


def load_tracker():
    path = [f for f in glob.glob(os.path.join(TRACK_DIR, "*.xlsx"))
            if "الجديد" in os.path.basename(f)
            and not os.path.basename(f).startswith("~$")
            and "-2.xlsx" not in os.path.basename(f)][0]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    best, bkey = None, None
    for nm in wb.sheetnames:
        m = re.match(r"\s*5-(\d{1,2})-(\d{4})", nm.strip())
        if not m:
            continue
        key = (int(m.group(2)), int(m.group(1)))
        if bkey is None or key > bkey or (key == bkey and "(2)" in nm):
            best, bkey = nm, key
    rows = list(wb[best].iter_rows(values_only=True))
    out = {}
    for r in rows[16:]:
        if not r or len(r) <= 17:
            continue
        m = re.match(r"^\s*(\d+)\s*$", str(r[17] or ""))
        if not m:
            continue
        out[int(m.group(1))] = {
            "executed": (r[1] or 0) + (r[2] or 0),
            "value": r[12] or r[11],
            "status": str(r[8] or "").strip(),
            "name": str(r[13] or "").strip(),
        }
    wb.close()
    return best, out


def main():
    aliases = json.load(open(MAP_PATH, encoding="utf-8")).get("aliases", {})
    folders = payment_folders()
    print(f"{len(folders)} payment folders under {PAY_ROOT}")

    # wo -> {'months': {iso: {company: qty_dict}}, 'cum': {company: (date, qty_dict)}}
    data = {}
    stats = {"folders": 0, "wo_folders": 0, "books": 0, "no_book": [], "unreadable": []}
    for fdate, fpath in folders:
        stats["folders"] += 1
        iso = fdate.isoformat()
        print(f"  {os.path.basename(fpath)}")
        for sub in sorted(os.listdir(fpath)):
            spath = os.path.join(fpath, sub)
            if not os.path.isdir(spath):
                continue
            wk = wo_key_of_folder(sub, aliases)
            if wk is None:
                continue
            stats["wo_folders"] += 1
            books = [b for b in glob.glob(os.path.join(spath, "جميع الشركات*.xls*"))
                     if not os.path.basename(b).startswith("~$")]
            if not books:
                stats["no_book"].append(f"{os.path.basename(fpath)} / {sub}")
                continue
            try:
                wb = openpyxl.load_workbook(books[0], data_only=True, read_only=True)
            except Exception as e:
                stats["unreadable"].append(f"{os.path.basename(fpath)} / {sub}: {e}")
                continue
            stats["books"] += 1
            slot_names = {}
            for nm in wb.sheetnames:
                kind, comp, slot = classify_sheet(nm)
                if slot and comp:
                    slot_names[slot] = comp
            rec = data.setdefault(wk, {"months": {}, "cum": {}, "folders": []})
            rec["folders"].append(iso)
            for nm in wb.sheetnames:
                kind, comp, slot = classify_sheet(nm)
                if kind in ("skip", "carry"):
                    continue
                if comp is None and slot is not None:
                    comp = slot_names.get(slot)
                if comp is None:
                    continue
                try:
                    lines = parse_company_sheet(wb[nm])
                except Exception:
                    continue
                if not lines:
                    continue
                if kind == "cumulative":
                    prev = rec["cum"].get(comp)
                    if prev is None or iso >= prev[0]:
                        rec["cum"][comp] = (iso, lines)
                else:
                    rec["months"].setdefault(iso, {}).setdefault(comp, {})
                    tgt = rec["months"][iso][comp]
                    for k, v in lines.items():
                        tgt[k] = tgt.get(k, 0) + v
            wb.close()

    # ── per-WO totals from the payment folders ──────────────────────────
    folder_exec = {}
    detail = {}
    for wk, rec in data.items():
        if wk[0] != "wo":
            continue
        wo = wk[1]
        comps = set(rec["cum"]) | {c for m in rec["months"].values() for c in m}
        total = 0.0
        per_comp = {}
        for comp in comps:
            # A قديم+جديد sheet is cumulative THROUGH its own folder's month;
            # months after it are only present as plain sheets, so add those.
            if comp in rec["cum"]:
                iso, lines = rec["cum"][comp]
                v = val(lines)
                after = {m: q for m, q in rec["months"].items() if m > iso and comp in q}
                extra = 0.0
                for m, q in after.items():
                    extra += val(q[comp])
                per_comp[comp] = {
                    "basis": f"قديم+جديد @ {iso}" + (f" + {len(after)} later monthly" if after else ""),
                    "value": v + extra}
            else:
                agg = {}
                for m in rec["months"].values():
                    for k, q in (m.get(comp) or {}).items():
                        agg[k] = agg.get(k, 0) + q
                v = val(agg)
                n = sum(1 for m in rec["months"].values() if comp in m)
                per_comp[comp] = {"basis": f"Σ {n} monthly sheets", "value": v}
            total += per_comp[comp]["value"]
        monthly = {iso: sum(val(q) for q in comps_.values())
                   for iso, comps_ in sorted(rec["months"].items())}
        folder_exec[wo] = total
        detail[wo] = {"perCompany": per_comp, "monthly": monthly,
                      "folders": sorted(set(rec["folders"]))}

    # ── the other three sources ─────────────────────────────────────────
    sheet_name, track = load_tracker()
    bf = json.load(open(os.path.join(OUT_DIR, "qm-backfill-data.json"), encoding="utf-8"))
    ours = {}
    for k, rec in bf.items():
        m = re.match(r"\('wo', (\d+)\)", k)
        if m:
            ours[int(m.group(1))] = sum(
                (BOP.get(kk, {}).get("rate", 0) or 0) * q
                for kk, q in (rec.get("execMinistry") or {}).items()) * PCT
    certs = json.load(open(os.path.join(OUT_DIR, "paycerts-data.json"), encoding="utf-8"))
    cert_by_wo, cert_by_month = {}, {}
    for no, c in certs.items():
        for l in c["lines"]:
            cert_by_wo[l["wo"]] = cert_by_wo.get(l["wo"], 0) + l["amount"] * PCT
            cert_by_month[c["period_end"]] = cert_by_month.get(c["period_end"], 0) + l["amount"] * PCT

    # ── report ──────────────────────────────────────────────────────────
    R = ["# دفعات الوزارة folders — audit vs the tracking report\n",
         f"Parsed **{stats['folders']} payment folders**, {stats['wo_folders']} work-order "
         f"sub-folders, {stats['books']} «جميع الشركات» workbooks.",
         f"Tracker sheet: **{sheet_name}**. All values after نسبة العقد (+9%).\n",
         "`folders` = executed to date from the payment folders (latest قديم+جديد per "
         "company, else Σ of that company's monthly sheets).\n",
         "| WO | tracker | folders | Δ (folders−tracker) | our executed | our certs | "
         "months seen | status |",
         "|---|---|---|---|---|---|---|---|"]
    rows = []
    for wo in sorted(set(track) | set(folder_exec)):
        t = track.get(wo, {}).get("executed")
        f = folder_exec.get(wo)
        o = ours.get(wo)
        c = cert_by_wo.get(wo)
        d = (f - t) if (f is not None and t is not None) else None
        months = len(detail.get(wo, {}).get("monthly", {}))
        st = track.get(wo, {}).get("status", "—")
        rows.append((wo, t, f, d, o, c, months, st))
        fmt = lambda v: f"{v:,.3f}" if isinstance(v, (int, float)) else "—"
        R.append(f"| {wo} | {fmt(t)} | {fmt(f)} | {fmt(d)} | {fmt(o)} | {fmt(c)} | "
                 f"{months or '—'} | {st} |")

    tot_t = sum(v["executed"] for v in track.values())
    tot_f = sum(folder_exec.values())
    tot_o = sum(ours.values())
    tot_c = sum(cert_by_wo.values())
    R += ["", "## Totals\n",
          "| source | executed (after 9%) |", "|---|---|",
          f"| tracking report | {tot_t:,.3f} |",
          f"| دفعات الوزارة folders | {tot_f:,.3f} |",
          f"| module — كشف حساب مجموع | {tot_o:,.3f} |",
          f"| module — payment certificates (0041) | {tot_c:,.3f} |"]

    close = [r for r in rows if r[3] is not None and abs(r[3]) <= 1000]
    far = sorted([r for r in rows if r[3] is not None and abs(r[3]) > 1000],
                 key=lambda r: -abs(r[3]))
    missing_folder = [r[0] for r in rows if r[2] is None]
    R += ["", f"## Agreement\n",
          f"- Work orders where the folders match the tracker within KD 1,000: "
          f"**{len(close)}** of {len([r for r in rows if r[3] is not None])}",
          f"- Work orders with no folder data at all: {missing_folder or 'none'}",
          "", f"### {len(far)} work orders where the folders and the tracker disagree by > KD 1,000\n",
          "| WO | tracker | folders | Δ | our executed | months seen | status |",
          "|---|---|---|---|---|---|---|"]
    for wo, t, f, d, o, c, months, st in far:
        fmt = lambda v: f"{v:,.3f}" if isinstance(v, (int, float)) else "—"
        R.append(f"| {wo} | {fmt(t)} | {fmt(f)} | {fmt(d)} | {fmt(o)} | {months or '—'} | {st} |")

    R += ["", "## Month-by-month (all work orders)\n",
          "| month folder | folders Σ monthly sheets | our certificate (same month) | Δ |",
          "|---|---|---|---|"]
    all_months = sorted({m for d in detail.values() for m in d["monthly"]}
                        | set(cert_by_month))
    for m in all_months:
        fv = sum(d["monthly"].get(m, 0) for d in detail.values())
        cv = cert_by_month.get(m)
        dv = (fv - cv) if cv is not None else None
        fmt = lambda v: f"{v:,.3f}" if isinstance(v, (int, float)) else "—"
        R.append(f"| {m} | {fmt(fv)} | {fmt(cv)} | {fmt(dv)} |")

    if stats["no_book"]:
        R += ["", "## Work-order folders with no «جميع الشركات» workbook\n"]
        R += [f"- {x}" for x in stats["no_book"]]
    if stats["unreadable"]:
        R += ["", "## Unreadable workbooks\n"] + [f"- {x}" for x in stats["unreadable"]]

    open(os.path.join(OUT_DIR, "payfolders-audit.md"), "w", encoding="utf-8").write("\n".join(R) + "\n")
    json.dump({"folderExec": folder_exec, "detail": detail,
               "tracker": {k: v["executed"] for k, v in track.items()}},
              open(os.path.join(OUT_DIR, "payfolders-audit.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    print(f"\ntracker {tot_t:,.3f} | folders {tot_f:,.3f} | ours {tot_o:,.3f} | certs {tot_c:,.3f}")
    print(f"within 1k: {len(close)} | diverging: {len(far)} | no folder data: {missing_folder}")
    print("written:", os.path.join(OUT_DIR, "payfolders-audit.md"))


if __name__ == "__main__":
    main()
