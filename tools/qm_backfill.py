# -*- coding: utf-8 -*-
"""Quantities module historical backfill (QUANTITIES_MODULE_BRIEF.md §6).

Reads the QA's corpus and emits ONE idempotent SQL paste:
  1. WO headers      — سجل أوامر العمل والدفعات - عقد 9.xlsx (ministry-PDF-derived register)
  2. Kashef lines    — D:\\التجميع الشهري new\\تفاصيل + جزئيات\\تفاصيل + جزئيات-2\\امر عمل N\\كشف حساب…
                       sheet الدفعه: كمية امر العمل per bab/band + جزئي 1..8 dated monthly
                       executed + مجموع cumulative (ministry-facing).
  3. Per-sub executed — دفعات الوزارة\\دفعة M-YYYY\\امر عمل N\\جميع الشركات-*.xlsm company sheets:
                       قديم+جديد sheet = that sub's cumulative per line (authoritative);
                       companies without قديم+جديد sheets (كوبري، المد الاخضر) are summed
                       across their monthly plain sheets.

Import model (approved by Fouad 2026-08-12):
  - kashefs created in `wo` status with real WO numbers + issue dates; unnumbered
    kashef folders (سلوى ق12 variants…) imported as live `kashef` status.
  - executed history = ONE opening طلب تدقيق per WO × sub (opening=true, cumulative
    qty, dated at that WO's last دفعة) — never fabricated monthly history.
  - COPRI self-performed works = the configured pseudo-sub vendor; its quantities are
    what the كوبري sheets carry (cross-checked against ministry مجموع − Σ subs).
  - allocations seeded = that sub's cumulative executed (baseline; QA adjusts forward).

Like tools/import_pos.py: the script ALWAYS writes the report + dataset, but refuses
to emit SQL until "confirmed": true is set in tools/qm-backfill-map.json (created on
first run with proposed sub→vendor mappings for Fouad to correct).

In-sheet titles inside جميع الشركات workbooks are STALE COPIES (wrong WO/site) —
identity comes from folder + file names only.
"""
import sys, io, os, re, json, glob, datetime, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
D_ROOT = r"D:\التجميع الشهري new"
HESAB_ROOT = os.path.join(D_ROOT, "تفاصيل + جزئيات", "تفاصيل + جزئيات-2")
PAY_ROOT = os.path.join(D_ROOT, "دفعات الوزارة")
REGISTER = os.path.join(os.path.expanduser("~"), "Downloads", "سجل أوامر العمل والدفعات - عقد 9.xlsx")
SEED_SQL = os.path.join(REPO, "supabase", "migrations", "0034_qm_bop_seed.sql")
CONFIG_PATH = os.path.join(HERE, "qm-backfill-map.json")
OUT_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "quantities-backfill")
os.makedirs(OUT_DIR, exist_ok=True)

KNOWN_BABS = {1, 2, 3, 4, 5, 6, 7, 12, 14, 17, 22}
AR_LETTERS = "اأإآبجدهوزحطي"
ID_RE = re.compile(r"^(\d+)\s*([" + AR_LETTERS + r"]?)\s*/\s*(\d+)\s*([" + AR_LETTERS + r"]?)$")

# ── helpers ──────────────────────────────────────────────────────────


def clean(v):
    if v is None:
        return ""
    s = str(v)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Cf")
    return s.replace("\xa0", " ").strip()


def norm_ar(s):
    """Normalize Arabic for fuzzy company matching."""
    s = clean(s)
    s = re.sub("[أإآ]", "ا", s)
    s = s.replace("ى", "ي").replace("ة", "ه")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def norm_suffix(s):
    """BOP suffix letters: normalize hamza variants (sheets write أ where the BOP has ا)."""
    return re.sub("[أإآ]", "ا", clean(s))


def esc(s):
    return str(s).replace("'", "''")


def num(v):
    """SQL numeric literal without float noise."""
    return repr(round(float(v), 3)).rstrip("0").rstrip(".") if v else "0"


# ── BOP keys from the seed migration (matches the DB exactly) ────────


def load_bop_keys():
    keys = {}
    txt = open(SEED_SQL, encoding="utf-8").read()
    for m in re.finditer(
        r"\(v_contract, (\d+), (\d+), (null|'[^']*'), '((?:[^']|'')*)', '((?:[^']|'')*)', ([\d.]+),", txt
    ):
        bab, band = int(m.group(1)), int(m.group(2))
        suf = "" if m.group(3) == "null" else norm_suffix(m.group(3).strip("'"))
        keys[(bab, band, suf)] = {"rate": float(m.group(6)), "unit": m.group(5).replace("''", "'")}
    if len(keys) < 1000:
        raise SystemExit(f"BOP seed parse failed: only {len(keys)} keys")
    return keys


# ── company canon ────────────────────────────────────────────────────

COMPANY_TOKENS = [
    ("الجارحي", ["الجارحي", "الجارحى"]),
    ("الكندية", ["الكنديه", "الكندية", "الكندي", "الكند"]),
    ("فتيح", ["فتيح"]),
    ("اليمامة", ["اليمامه", "اليمامة"]),
    ("المد الاخضر", ["المد الاخضر", "الخط الاخضر"]),
    ("بحر الابداع", ["بحر الابداع"]),
    ("عبيد", ["عبيد"]),
    ("المثنى", ["المثني", "المثنى"]),
    ("الوفرة", ["الوفره", "الوفرة"]),
    ("بوبيان", ["بوبيان"]),
    ("وايت بروجكت", ["وايت"]),
    ("دالكو", ["دالكو"]),
    ("الجود", ["الجود"]),
    ("دانة الرتاج", ["دانه الرتاج", "دانة الرتاج"]),
    ("CCC", ["ccc"]),
    ("كوبري", ["كوبري", "كوبرى"]),
]

SKIP_SHEET_PAT = re.compile(
    "|".join(["مجمع", "صافي", "متبقي", "كميات", "الدفعه الفعليه", "الدفعة الفعلية", "نسبه", "نسبة", "بدون طلبات", "بلا طلبات"])
)


def company_of(sheet_name):
    s = norm_ar(sheet_name).lower()
    for canon, toks in COMPANY_TOKENS:
        for t in toks:
            if norm_ar(t).lower() in s:
                return canon
    return None


def classify_sheet(name):
    """→ ('skip'|'cumulative'|'carry'|'plain', company|None, slot|None)"""
    s = norm_ar(name)
    if s in ("10", "الدفعه") or s.isdigit():
        return ("skip", None, None)
    if SKIP_SHEET_PAT.search(s):
        return ("skip", None, None)
    comp = company_of(s)
    slot = None
    m = re.search(r"مقاول\s*باطن\s*(\d+)", s)
    if m:
        slot = int(m.group(1))
    if "قديم" in s and "جديد" in s:
        return ("cumulative", comp, slot)
    if "مرحل" in s:
        return ("carry", comp, slot)
    return ("plain", comp, slot)


# ── generic line-sheet parser (جميع الشركات + قديم+جديد layout) ──────
# cols: 0 total, 1 rate, 2 unit, 3 qty, 4 desc, 8 باب/بند, 9 باب, 10 بند, 11 حرف


def parse_company_sheet(ws):
    lines = {}
    rows = ws.iter_rows(values_only=True)
    started = False
    for r in rows:
        if not r or len(r) < 11:
            continue
        if not started:
            if clean(r[8] if len(r) > 8 else None).replace(" ", "") in ("باب/بند", "بند/باب"):
                started = True
            continue
        bab, band = r[9] if len(r) > 9 else None, r[10] if len(r) > 10 else None
        if not isinstance(bab, (int, float)) or not isinstance(band, (int, float)):
            continue
        suf = norm_suffix(r[11]) if len(r) > 11 else ""
        qty = r[3] if isinstance(r[3], (int, float)) else 0
        if qty:
            key = (int(bab), int(band), suf)
            lines[key] = lines.get(key, 0) + float(qty)
    return lines


# ── كشف حساب parser (ported from the validated survey) ───────────────


def parse_hesab(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = None
    for nm in wb.sheetnames:
        if "دفع" in nm:
            sheet = wb[nm]
            break
    if sheet is None:
        wb.close()
        return None
    rows = list(sheet.iter_rows(values_only=True))
    hdr_i = id_col = None
    for i, r in enumerate(rows[:20]):
        for j, c in enumerate(r or []):
            if clean(c).replace(" ", "") in ("باب/بند", "بند/باب"):
                hdr_i, id_col = i, j
                break
        if hdr_i is not None:
            break
    if hdr_i is None:
        wb.close()
        return None
    hdr = rows[hdr_i]

    def find_col(*names):
        for j, c in enumerate(hdr or []):
            if any(n in clean(c) for n in names):
                return j
        return None

    qty_col = find_col("كمية امر العمل", "كمية أمر العمل", "الكمية")
    rate_col = find_col("سعر")
    unit_col = find_col("الوحدة")
    total_col = None
    for j, c in enumerate(hdr or []):
        if j > id_col and clean(c) == "مجموع":
            total_col = j
            break
    partial_cols = []
    for j, c in enumerate(hdr or []):
        if "جزئي" in clean(c):
            dt = None
            for back in range(1, 4):
                if hdr_i - back >= 0:
                    rr = rows[hdr_i - back] or []
                    v = rr[j] if j < len(rr) else None
                    if isinstance(v, datetime.datetime):
                        dt = v.date()
                        break
            partial_cols.append((j, dt))

    lines = []
    for r in rows[hdr_i + 1:]:
        if not r:
            continue
        raw = clean(r[id_col] if id_col < len(r) else None)
        m = ID_RE.match(raw)
        if not m:
            continue
        a, sa, b, sb = int(m.group(1)), m.group(2), int(m.group(3)), m.group(4)
        suf = norm_suffix(sa or sb or "")
        if a in KNOWN_BABS:
            bab, band = a, b
        elif b in KNOWN_BABS:
            bab, band = b, a
        else:
            continue
        qty = r[qty_col] if qty_col is not None and qty_col < len(r) and isinstance(r[qty_col], (int, float)) else None
        rate = r[rate_col] if rate_col is not None and rate_col < len(r) and isinstance(r[rate_col], (int, float)) else None
        psum = 0.0
        last_date = None
        for (pc, pd) in partial_cols:
            v = r[pc] if pc < len(r) else None
            if isinstance(v, (int, float)) and v:
                psum += v
                if pd:
                    last_date = pd
        tot = r[total_col] if total_col is not None and total_col < len(r) else None
        # trust مجموع unless it is stale (0 while the جزئي columns carry values — WO 72)
        cum = tot if isinstance(tot, (int, float)) and (tot != 0 or psum == 0) else psum
        lines.append({
            "bab": bab, "band": band, "suffix": suf, "raw": raw,
            "qty": qty, "rate": rate,
            "unit": clean(r[unit_col]) if unit_col is not None and unit_col < len(r) else "",
            "cumExec": cum, "partialSum": psum, "lastDate": last_date,
        })
    wb.close()
    return lines


# ── payment folders scan ─────────────────────────────────────────────

PAY_FOLDER_RE = re.compile(r"^دفعة\s*(\d+)-(\d+)-(\d{4})$")


def payment_folders():
    out = []
    for name in os.listdir(PAY_ROOT):
        p = os.path.join(PAY_ROOT, name)
        if not os.path.isdir(p):
            continue
        m = PAY_FOLDER_RE.match(norm_ar(name).replace("دفعه", "دفعة"))
        if not m:
            continue  # e.g. 'دفعة 5-9-2025 - دفعة المقاولين' — contractor variant, skip
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        out.append((datetime.date(y, mo, d), p))
    return sorted(out)


def wo_key_of_folder(name, aliases):
    """Return ('wo', N) or ('alias', key) or None for a payment/hesab WO folder name."""
    s = norm_ar(name)
    m = re.search(r"امر عمل\s*\(?\s*(\d+)\s*\)?", s)
    if m and int(m.group(1)) > 0:
        return ("wo", int(m.group(1)))
    for key, alias in aliases.items():
        for tok in alias["matchTokens"]:
            if all(norm_ar(t) in s for t in tok.split()):
                return ("alias", key)
    return None


def scan_payments(aliases, progress=True):
    """→ {woKey: {company: {'cum': {line: qty} from latest cum sheet,
                            'cumDate': date, 'plainSum': {line: qty}, 'plainMonths': n}}}"""
    data = {}
    folders = payment_folders()
    for fdate, fpath in folders:
        if progress:
            print(f"  scanning {os.path.basename(fpath)}…")
        for sub in os.listdir(fpath):
            spath = os.path.join(fpath, sub)
            if not os.path.isdir(spath):
                continue
            wk = wo_key_of_folder(sub, aliases)
            if wk is None:
                continue
            books = glob.glob(os.path.join(spath, "جميع الشركات*.xls*"))
            if not books:
                continue
            try:
                wb = openpyxl.load_workbook(books[0], data_only=True, read_only=True)
            except Exception:
                continue
            # resolve unnamed slots from named siblings
            slot_names = {}
            for nm in wb.sheetnames:
                kind, comp, slot = classify_sheet(nm)
                if slot and comp:
                    slot_names[slot] = comp
            rec = data.setdefault(wk, {})
            for nm in wb.sheetnames:
                kind, comp, slot = classify_sheet(nm)
                if kind in ("skip", "carry"):
                    continue
                if comp is None and slot is not None:
                    comp = slot_names.get(slot)
                if comp is None:
                    if kind == "cumulative":
                        rec.setdefault("_unresolvedCum", []).append(f"{os.path.basename(spath)}::{nm}")
                    continue
                try:
                    lines = parse_company_sheet(wb[nm])
                except Exception:
                    continue
                c = rec.setdefault(comp, {"cum": None, "cumDate": None, "plainSum": {}, "plainMonths": 0})
                if kind == "cumulative":
                    if lines or c["cum"] is None:
                        c["cum"] = lines
                        c["cumDate"] = fdate
                elif kind == "plain":
                    if lines:
                        c["plainMonths"] += 1
                        c["lastPlainDate"] = fdate.isoformat()
                        for k, v in lines.items():
                            c["plainSum"][k] = c["plainSum"].get(k, 0) + v
            wb.close()
    return data


# ── register parse ───────────────────────────────────────────────────


def parse_date_str(s):
    s = clean(s)
    m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})", s)
    if m:
        return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return None


def parse_register():
    wb = openpyxl.load_workbook(REGISTER, data_only=True, read_only=True)
    ws = None
    for nm in wb.sheetnames:
        if "أوامر العمل" in nm or "اوامر العمل" in norm_ar(nm):
            ws = wb[nm]
            break
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    for r in rows:
        if not r or not isinstance(r[0], (int, float)):
            continue
        wo = int(r[0])
        out[wo] = {
            "wo": wo,
            "desc": clean(r[1]),
            "area": clean(r[2]),
            "type": clean(r[3]),
            "issued": parse_date_str(r[4]),
            "duration": int(r[5]) if len(r) > 5 and isinstance(r[5], (int, float)) else None,
            "start": parse_date_str(r[6]),
            "currentValue": r[11] if len(r) > 11 and isinstance(r[11], (int, float)) else None,
        }
    wb.close()
    return out


def parse_location(reg):
    """register row → (area, loc_type, block_no, street_name, work_type)"""
    desc = reg["desc"]
    area = reg["area"] if reg["area"] not in ("متفرقات", "اخرى", "أخرى") else ""
    work = {"مدني": "أعمال مدنية", "اسفلت": "أسفلت", "صحي": "صحي", "متفرقات": "متفرقات"}.get(reg["type"], reg["type"])
    # letter-boundary guard: bare 'ق N' must not match the ق inside نطاق etc.
    # قطع[ةه] tolerates the ta-marbuta/ha spelling variants in the register.
    m = re.search(r"قطع[ةه]\s*\(?\s*(\d+\w*)\s*\)?|(?<![ء-ي])ق\s*(\d+)", desc)
    if m:
        return (area or desc.split("قطعة")[0].strip(), "block", m.group(1) or m.group(2), "", work)
    m = re.search(r"(شارع\s+\S.*?)(?:\(|$)", desc)
    if m:
        return (area or "", "street", "", m.group(1).strip(), work)
    return (area or desc[:40], "misc", "", "", work)


# ── config ───────────────────────────────────────────────────────────

DEFAULT_CONFIG = {
    "confirmed": False,
    "_help": [
        "Review everything below, correct vendorId/active/notes, then set confirmed: true and re-run.",
        "subs: observed company name → vendors.id (null = create/resolve first; script refuses SQL for",
        "active companies with null vendorId). active:false = exclude that company's quantities entirely.",
        "aliases: unnumbered kashef folders → kashef numbers (900-range keeps them clearly out of the WO series).",
    ],
    "copriSelfVendorName": "كوبري — تنفيذ ذاتي",
    "subs": {},
    "aliases": {
        "سلوى ق12 مدني": {"matchTokens": ["مدني سلوي ق 12", "مدني سلوى ق12", "مدني سلوي ق12"], "kashefNo": 901,
                          "area": "سلوى", "locType": "block", "blockNo": "12", "workType": "أعمال مدنية"},
        "سلوى ق12 اسفلت": {"matchTokens": ["اسفلت سلوي ق12", "اسفلت سلوى ق12", "اسفلت سلوي ق 12"], "kashefNo": 902,
                           "area": "سلوى", "locType": "block", "blockNo": "12", "workType": "أسفلت"},
        "سلوى ق12 صحي": {"matchTokens": ["صحي سلوي ق 12", "صحي سلوى ق12", "0-صحي"], "kashefNo": 903,
                         "area": "سلوى", "locType": "block", "blockNo": "12", "workType": "صحي"},
        "سلوى ق12 أمطار": {"matchTokens": ["سلوي ق 12 اممطار", "سلوي ق 12 امطار"], "kashefNo": 904,
                           "area": "سلوى", "locType": "block", "blockNo": "12", "workType": "أمطار"},
        "الجاليات": {"matchTokens": ["الجاليات"], "kashefNo": 905,
                     "area": "الجاليات", "locType": "misc", "blockNo": "", "workType": ""},
        "متفرقات جديد": {"matchTokens": ["متفرقات جديد"], "kashefNo": 906,
                         "area": "متفرقات", "locType": "misc", "blockNo": "", "workType": "متفرقات"},
    },
}

VENDOR_PROPOSALS = {
    "الجارحي": {"vendorId": 494, "note": "AL JARHI UNITED CONTRACTING CO."},
    "المثنى": {"vendorId": 490, "note": "AL MOTHANNA PLUS CO.GEN.TRAD.&CONT."},
    "المد الاخضر": {"vendorId": 54, "note": "AL-KHAT AL-AKHDAR GEN.TRADING & CONT.CO. — CONFIRM المد=الخط"},
    "اليمامة": {"vendorId": 538, "note": "AL YAMAMA AL KHALEEJIA GEN.TRAD.&CONT.CO."},
    "بحر الابداع": {"vendorId": None, "note": "7 duplicate rows (423/424/428/429/482/489/498) — pick canonical"},
    "الكندية": {"vendorId": 500, "note": "The Regional Canadian Co. General Contracting For Buildings — CONFIRM"},
    "عبيد": {"vendorId": None, "note": "candidate 371 OBAID RAMADAN ALASWAD — CONFIRM"},
    "فتيح": {"vendorId": None, "note": "no vendor row found — create in Table Editor, put id here"},
    "الوفرة": {"vendorId": 536, "note": "WAFRA NATIONAL BUILDING CONSTRUCTION COMPANY — CONFIRM"},
    "بوبيان": {"vendorId": None, "note": "525 vs 530 United Boubyan duplicates — pick canonical"},
    "وايت بروجكت": {"vendorId": None, "note": "468 vs 488 WHITE PROJECT duplicates — pick canonical"},
    "دالكو": {"vendorId": None, "note": "493 DALCO CO. vs 487 DALCO LIMITED — pick canonical"},
    "الجود": {"vendorId": 495, "note": "AL-JOUD AL-MUTAMAIYZA GEN.TRAD.&CONT.CO."},
    "دانة الرتاج": {"vendorId": None, "note": "not searched — fill if it shows up with quantities"},
    "CCC": {"vendorId": None, "mergeInto": "كوبري",
            "note": "Copri's own asphalt operation (per Fouad 2026-08-12) — folded into كوبري تنفيذ ذاتي"},
    "كوبري": {"vendorId": None, "note": "self-performed — leave null; script uses copriSelfVendorName row"},
}


def load_config():
    if os.path.exists(CONFIG_PATH):
        return json.load(open(CONFIG_PATH, encoding="utf-8"))
    return json.loads(json.dumps(DEFAULT_CONFIG))


def save_config(cfg):
    json.dump(cfg, open(CONFIG_PATH, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


# ── main ─────────────────────────────────────────────────────────────


def main():
    cfg = load_config()
    bop = load_bop_keys()
    print(f"BOP keys: {len(bop)}")
    register = parse_register()
    print(f"register WOs: {len(register)}")

    # hesab corpus
    kashefs = {}  # key ('wo', N) or ('alias', key) → record
    for folder in sorted(os.listdir(HESAB_ROOT)):
        fpath = os.path.join(HESAB_ROOT, folder)
        if not os.path.isdir(fpath):
            continue
        wk = wo_key_of_folder(folder, cfg["aliases"])
        if wk is None:
            # unnumbered hesab folders match aliases by tokens too
            print(f"  !! hesab folder not recognized: {folder}")
            continue
        files = glob.glob(os.path.join(fpath, "كشف حساب*.xls*"))
        if not files:
            continue
        lines = parse_hesab(files[0])
        if lines is None:
            print(f"  !! parse failed: {folder}")
            continue
        kashefs[wk] = {"folder": folder, "lines": lines}
    print(f"hesab kashefs parsed: {len(kashefs)}")

    # payments scan
    print("scanning payment folders…")
    pay = scan_payments(cfg["aliases"])

    # collect observed companies → refresh config sub proposals
    observed = {}
    for wk, comps in pay.items():
        for comp, c in comps.items():
            if comp.startswith("_"):
                continue
            cum_total = sum((c["cum"] or {}).values()) if c["cum"] else 0
            plain_total = sum(c["plainSum"].values())
            o = observed.setdefault(comp, {"wos": 0, "cumQty": 0.0, "plainQty": 0.0})
            o["wos"] += 1
            o["cumQty"] += cum_total
            o["plainQty"] += plain_total
    for comp in sorted(observed):
        if comp not in cfg["subs"]:
            prop = VENDOR_PROPOSALS.get(comp, {"vendorId": None, "note": "NEW — map me"})
            cfg["subs"][comp] = {"vendorId": prop.get("vendorId"),
                                 "active": prop.get("active", True),
                                 "note": prop.get("note", "")}
    save_config(cfg)

    # build final per-kashef dataset
    dataset = {}
    flags = []
    for wk, hes in kashefs.items():
        kind, ident = wk
        if kind == "wo":
            reg = register.get(ident)
            if reg is None:
                flags.append(f"WO {ident}: has كشف حساب but NOT in the register — skipped")
                continue
            area, loc, block, street, work = parse_location(reg)
            head = {"kashefNo": ident, "status": "wo", "woNo": str(ident),
                    "woDate": reg["issued"].isoformat() if reg["issued"] else None,
                    "duration": reg["duration"],
                    "area": area, "locType": loc, "blockNo": block, "streetName": street,
                    "workType": work, "desc": reg["desc"], "registerValue": reg["currentValue"]}
        else:
            # direct-WO model (QA feedback 2026-08-12): unnumbered kashef
            # folders import as WOs too, numbered in the 900 range.
            al = cfg["aliases"][ident]
            head = {"kashefNo": al["kashefNo"], "status": "wo", "woNo": str(al["kashefNo"]),
                    "woDate": None, "duration": None,
                    "area": al["area"], "locType": al["locType"], "blockNo": al["blockNo"],
                    "streetName": al.get("streetName", ""), "workType": al["workType"],
                    "desc": ident, "registerValue": None}

        # kashef lines: aggregate duplicate ids (amendment rows repeat ids)
        lines = {}
        exec_total = {}
        last_date = None
        for l in hes["lines"]:
            key = (l["bab"], l["band"], l["suffix"])
            if key not in bop:
                flags.append(f"{head['desc']}: id {l['raw']} not in BOP — line skipped")
                continue
            lines[key] = lines.get(key, 0) + (l["qty"] or 0)
            exec_total[key] = exec_total.get(key, 0) + (l["cumExec"] or 0)
            if l["lastDate"] and (last_date is None or l["lastDate"] > last_date):
                last_date = l["lastDate"]

        # per-company executed
        comps = pay.get(wk, {})
        by_sub = {}
        merges = {}   # target company → {line: qty} merged from mergeInto companies
        for comp, c in comps.items():
            if comp.startswith("_"):
                continue
            sub_cfg = cfg["subs"].get(comp, {})
            target = sub_cfg.get("mergeInto")
            if target:
                # e.g. CCC = Copri's own asphalt operation → part of كوبري.
                # Under residual mode the target's residual absorbs it anyway;
                # the merge matters only when residual doesn't run.
                src = c["cum"] if c["cum"] else c["plainSum"]
                tgt = merges.setdefault(target, {})
                for key, q in (src or {}).items():
                    if key in bop and q > 0.0005:
                        tgt[key] = tgt.get(key, 0) + q
                continue
            if sub_cfg.get("active") is False:
                if c["cum"] or c["plainSum"]:
                    tot = sum((c["cum"] or c["plainSum"]).values())
                    if tot > 0.001:
                        flags.append(f"{head['desc']}: INACTIVE company {comp} has qty {tot:,.1f} — excluded, review!")
                continue
            src = c["cum"] if c["cum"] else c["plainSum"]
            date = c["cumDate"].isoformat() if c["cumDate"] else c.get("lastPlainDate")
            if not src:
                continue
            entries = {}
            for key, q in src.items():
                if key not in bop:
                    flags.append(f"{head['desc']} / {comp}: id {key} not in BOP — skipped")
                    continue
                if q > 0.0005:
                    entries[key] = round(q, 3)
            if entries:
                by_sub[comp] = {"lines": entries, "date": date,
                                "source": "قديم+جديد" if c["cum"] else f"Σ {c['plainMonths']} monthly sheets"}

        # apply merges (CCC → كوبري): matters when the residual doesn't run
        for target, lines_m in merges.items():
            if not lines_m:
                continue
            t = by_sub.setdefault(target, {"lines": {}, "date": None, "source": "incl. merged"})
            for key, q in lines_m.items():
                t["lines"][key] = round(t["lines"].get(key, 0) + q, 3)

        # كوبري residual mode: the ministry مجموع per line is ground truth
        # (the QA reconciles it with the ministry to the fils). كوبري's own
        # monthly sheets over-count when summed (some months are cumulative
        # restatements), so self-performed = per-line residual after the
        # named subs. Only when كوبري has no قديم+جديد sheet of its own.
        if cfg.get("copriMode", "residual") == "residual":
            copri_has_cum = "كوبري" in comps and comps["كوبري"].get("cum")
            if exec_total and not copri_has_cum:
                sheet_sum = sum(by_sub.get("كوبري", {}).get("lines", {}).values())
                resid = {}
                overrun = 0
                for key, tot in exec_total.items():
                    others = sum(s["lines"].get(key, 0) for c, s in by_sub.items() if c != "كوبري")
                    r_ = tot - others
                    if r_ > 0.0005:
                        if key in bop:
                            resid[key] = round(r_, 3)
                    elif r_ < -0.0005:
                        overrun += 1
                if overrun:
                    flags.append(f"{head['desc']}: named subs exceed the ministry cumulative on {overrun} lines (kept as-is; كوبري 0 there)")
                if resid:
                    by_sub["كوبري"] = {"lines": resid, "date": last_date.isoformat() if last_date else None,
                                       "source": f"residual (sheets Σ was {sheet_sum:,.1f})"}
                elif "كوبري" in by_sub:
                    del by_sub["كوبري"]

        # reconciliation: Σ subs vs ministry cumulative
        recon = {"ministry": round(sum(exec_total.values()), 3),
                 "subsTotal": round(sum(q for s in by_sub.values() for q in s["lines"].values()), 3)}
        dataset[str(wk)] = {"head": head, "lines": {f"{k[0]}/{k[1]}{k[2]}": round(v, 3) for k, v in lines.items()},
                            "linesRaw": [(k[0], k[1], k[2], round(v, 3)) for k, v in lines.items()],
                            "execMinistry": {f"{k[0]}/{k[1]}{k[2]}": round(v, 3) for k, v in exec_total.items() if v},
                            "bySub": {c: {"date": s["date"], "source": s["source"],
                                          "lines": {f"{k[0]}/{k[1]}{k[2]}": v for k, v in s["lines"].items()},
                                          "linesRaw": [(k[0], k[1], k[2], v) for k, v in s["lines"].items()]}
                                      for c, s in by_sub.items()},
                            "lastExecDate": last_date.isoformat() if last_date else None,
                            "recon": recon}

    # payment-only per-sub data helper (no كشف حساب → residual impossible,
    # كوبري keeps its sheet sums; every tadqiq line will be out_of_kashef)
    def pay_subs(wk):
        out = {}
        merges = {}
        for comp, c in pay.get(wk, {}).items():
            if comp.startswith("_"):
                continue
            sub_cfg = cfg["subs"].get(comp, {})
            src = c["cum"] if c["cum"] else c["plainSum"]
            entries = {k: round(q, 3) for k, q in (src or {}).items() if k in bop and q > 0.0005}
            if not entries:
                continue
            target = sub_cfg.get("mergeInto")
            if target:
                tgt = merges.setdefault(target, {})
                for k, q in entries.items():
                    tgt[k] = round(tgt.get(k, 0) + q, 3)
                continue
            if sub_cfg.get("active", True) is False:
                continue
            out[comp] = {"lines": entries,
                         "date": c["cumDate"].isoformat() if c["cumDate"] else c.get("lastPlainDate"),
                         "source": ("قديم+جديد" if c["cum"] else f"Σ {c['plainMonths']} monthly sheets") + " — NO كشف حساب"}
        for target, lines_m in merges.items():
            t = out.setdefault(target, {"lines": {}, "date": None, "source": "incl. merged — NO كشف حساب"})
            for k, q in lines_m.items():
                t["lines"][k] = round(t["lines"].get(k, 0) + q, 3)
        return out

    # register WOs with no hesab → header-only (+ any per-sub payment data)
    for wo, reg in sorted(register.items()):
        if str(("wo", wo)) not in dataset:
            area, loc, block, street, work = parse_location(reg)
            subs = pay_subs(("wo", wo))
            dataset[str(("wo", wo))] = {
                "head": {"kashefNo": wo, "status": "wo", "woNo": str(wo),
                         "woDate": reg["issued"].isoformat() if reg["issued"] else None,
                         "duration": reg["duration"],
                         "area": area, "locType": loc, "blockNo": block, "streetName": street,
                         "workType": work, "desc": reg["desc"], "registerValue": reg["currentValue"]},
                "lines": {}, "linesRaw": [], "execMinistry": {},
                "bySub": {c: {"date": s["date"], "source": s["source"],
                              "lines": {f"{k[0]}/{k[1]}{k[2]}": v for k, v in s["lines"].items()},
                              "linesRaw": [(k[0], k[1], k[2], v) for k, v in s["lines"].items()]}
                          for c, s in subs.items()},
                "lastExecDate": None, "recon": None, "headerOnly": True}

    # alias kashefs seen only in payment folders (no كشف حساب) — e.g. مدني سلوى ق12
    for key, al in cfg["aliases"].items():
        wk = ("alias", key)
        if str(wk) in dataset or wk in kashefs:
            continue
        subs = pay_subs(wk)
        if not subs:
            continue
        flags.append(f"{key}: per-sub payment data but NO كشف حساب — imported header-only, entries flagged out-of-kashef")
        dataset[str(wk)] = {
            "head": {"kashefNo": al["kashefNo"], "status": "wo", "woNo": str(al["kashefNo"]),
                     "woDate": None, "duration": None,
                     "area": al["area"], "locType": al["locType"], "blockNo": al["blockNo"],
                     "streetName": al.get("streetName", ""), "workType": al["workType"],
                     "desc": key, "registerValue": None},
            "lines": {}, "linesRaw": [], "execMinistry": {},
            "bySub": {c: {"date": s["date"], "source": s["source"],
                          "lines": {f"{k[0]}/{k[1]}{k[2]}": v for k, v in s["lines"].items()},
                          "linesRaw": [(k[0], k[1], k[2], v) for k, v in s["lines"].items()]}
                      for c, s in subs.items()},
            "lastExecDate": None, "recon": None, "headerOnly": True}

    json.dump(dataset, open(os.path.join(OUT_DIR, "qm-backfill-data.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1, default=str)

    write_report(dataset, cfg, bop, flags)

    if not cfg.get("confirmed"):
        print("\nconfirmed=false → NO SQL emitted. Review the report + qm-backfill-map.json, "
              "set confirmed:true, re-run.")
        return

    missing = [c for c, s in cfg["subs"].items()
               if s.get("active", True) and s.get("vendorId") is None and c != "كوبري"
               and any(c in d["bySub"] for d in dataset.values())]
    if missing:
        print(f"\nREFUSED: active companies with quantities but no vendorId: {missing}")
        return
    emit_sql(dataset, cfg)


def write_report(dataset, cfg, bop, flags):
    L = []
    L.append("# Quantities backfill — dry-run report\n")
    L.append(f"Generated {datetime.date.today().isoformat()} from D:\\التجميع الشهري new. "
             f"Kashefs: {len(dataset)} ({sum(1 for d in dataset.values() if d.get('headerOnly'))} header-only).\n")
    L.append("\n## Per-kashef coverage + reconciliation\n")
    L.append("| kashef | status | lines | ministry exec (qty Σ) | subs exec (qty Σ) | subs | note |")
    L.append("|---|---|---|---|---|---|---|")
    for key, d in sorted(dataset.items(), key=lambda kv: kv[1]["head"]["kashefNo"]):
        h = d["head"]
        subs = "، ".join(f"{c} ({s['source']})" for c, s in d["bySub"].items()) or "—"
        note = "header-only" if d.get("headerOnly") else ""
        if d["recon"]:
            gap = d["recon"]["ministry"] - d["recon"]["subsTotal"]
            if abs(gap) > 0.5 and d["recon"]["ministry"]:
                note += f" gap(min−subs)={gap:,.1f}"
        L.append(f"| {h['kashefNo']} — {h['desc'][:45]} | {h['status']} | {len(d['lines'])} | "
                 f"{(d['recon'] or {}).get('ministry', 0):,.1f} | {(d['recon'] or {}).get('subsTotal', 0):,.1f} | {subs} | {note} |")
    L.append("\n## Company mapping (edit tools/qm-backfill-map.json)\n")
    L.append("| observed | vendorId | active | note |")
    L.append("|---|---|---|---|")
    for comp, s in sorted(cfg["subs"].items()):
        L.append(f"| {comp} | {s.get('vendorId')} | {s.get('active', True)} | {s.get('note','')} |")
    L.append("\n## Flags\n")
    for f in flags or ["(none)"]:
        L.append(f"- {f}")
    open(os.path.join(OUT_DIR, "qm-backfill-report.md"), "w", encoding="utf-8").write("\n".join(L))
    print(f"report → {os.path.join(OUT_DIR, 'qm-backfill-report.md')}")


def emit_sql(dataset, cfg):
    S = []
    S.append("-- 0036_qm_backfill.sql — GENERATED by tools/qm_backfill.py, do not hand-edit.")
    S.append("-- Historical backfill: work orders + lines + allocations + opening tadqiq entries.")
    S.append("-- Direct-WO model per 0035 (paste 0035 FIRST — this file uses duration_days).")
    S.append("-- Idempotent: each WO block is skipped when (contract, kashef_no) already exists.")
    S.append("do $$")
    S.append("declare")
    S.append("  v_contract bigint;")
    S.append("  v_copri bigint;")
    S.append("  v_k bigint;")
    S.append("  v_line bigint;")
    S.append("  v_t bigint;")
    S.append("  v_item bigint;")
    S.append("begin")
    S.append("  select id into v_contract from qm_contracts where code = 'HAW9';")
    S.append("  if v_contract is null then raise exception 'run 0033 first'; end if;")
    S.append(f"  select id into v_copri from vendors where name = '{esc(cfg['copriSelfVendorName'])}';")
    S.append("  if v_copri is null then")
    S.append(f"    insert into vendors (name, kind, internal, notes) values ('{esc(cfg['copriSelfVendorName'])}', 'internal', true, 'quantities module — COPRI self-performed works (backfill)') returning id into v_copri;")
    S.append("  end if;")
    S.append("  update vendors set qm_subcontractor = true where id = v_copri and not qm_subcontractor;")
    vids = sorted({s["vendorId"] for c, s in cfg["subs"].items()
                   if s.get("active", True) and s.get("vendorId")})
    if vids:
        S.append(f"  update vendors set qm_subcontractor = true where id in ({', '.join(map(str, vids))}) and not qm_subcontractor;")

    def vendor_expr(comp):
        if comp == "كوبري":
            return "v_copri"
        return str(cfg["subs"][comp]["vendorId"])

    for key, d in sorted(dataset.items(), key=lambda kv: kv[1]["head"]["kashefNo"]):
        h = d["head"]
        S.append("")
        S.append(f"  -- ── kashef {h['kashefNo']} — {h['desc'][:60]} ──")
        S.append(f"  if not exists (select 1 from qm_kashefs where contract_id = v_contract and kashef_no = {h['kashefNo']}) then")
        wo_date = f"'{h['woDate']}'" if h["woDate"] else "null"
        kdate = h["woDate"] or datetime.date.today().isoformat()
        duration = h.get("duration") if h.get("duration") else "null"
        S.append(
            "    insert into qm_kashefs (contract_id, kashef_no, area, loc_type, block_no, street_name, work_type, status, wo_no, wo_date, kashef_date, duration_days)"
        )
        S.append(
            f"    values (v_contract, {h['kashefNo']}, '{esc(h['area'])}', '{h['locType']}', '{esc(h['blockNo'])}', "
            f"'{esc(h['streetName'])}', '{esc(h['workType'])}', '{h['status']}', '{esc(h['woNo'])}', {wo_date}, '{kdate}', {duration}) returning id into v_k;"
        )
        S.append("    insert into qm_changelog (entity, entity_id, action, field, line_ref, old_value, new_value, actor_email)")
        S.append(f"    values ('kashef', v_k, 'create', '', '', '', 'استيراد تاريخي — {esc(h['desc'][:80])}', 'backfill');")

        for (bab, band, suf, qty) in sorted(d["linesRaw"]):
            sfx = "''" if not suf else f"'{esc(suf)}'"
            S.append(
                f"    select id into v_item from qm_bop_items where contract_id = v_contract and bab = {bab} and band = {band} and coalesce(suffix,'') = {sfx};"
            )
            S.append(f"    if v_item is null then raise exception 'bop {bab}/{band}{suf} missing'; end if;")
            S.append(f"    insert into qm_kashef_lines (kashef_id, bop_item_id, qty) values (v_k, v_item, {num(qty)});")

        for comp, sub in sorted(d["bySub"].items()):
            v = vendor_expr(comp)
            tdate = sub["date"] or d["lastExecDate"] or kdate
            S.append(f"    -- {comp}: allocations + opening tadqiq ({sub['source']})")
            S.append(
                f"    insert into qm_tadqiq (kashef_id, vendor_id, tadqiq_date, street_no, note, opening)"
            )
            S.append(
                f"    values (v_k, {v}, '{tdate}', '', 'رصيد افتتاحي مستورد ({esc(sub['source'])})', true) returning id into v_t;"
            )
            for (bab, band, suf, qty) in sorted(sub["linesRaw"]):
                sfx = "''" if not suf else f"'{esc(suf)}'"
                S.append(
                    f"    select id into v_item from qm_bop_items where contract_id = v_contract and bab = {bab} and band = {band} and coalesce(suffix,'') = {sfx};"
                )
                S.append(f"    if v_item is null then raise exception 'bop {bab}/{band}{suf} missing'; end if;")
                # kashef line may be absent (out-of-kashef executed work) — allocation needs a line
                S.append(f"    select id into v_line from qm_kashef_lines where kashef_id = v_k and bop_item_id = v_item;")
                S.append(f"    if v_line is not null then")
                S.append(
                    f"      insert into qm_allocations (kashef_line_id, vendor_id, qty) values (v_line, {v}, {num(qty)})"
                )
                S.append("      on conflict (kashef_line_id, vendor_id) do update set qty = qm_allocations.qty + excluded.qty;")
                S.append("    end if;")
                S.append(
                    f"    insert into qm_tadqiq_lines (tadqiq_id, bop_item_id, qty, out_of_kashef, over_allocation)"
                )
                S.append(f"    values (v_t, v_item, {num(qty)}, v_line is null, false);")
            S.append("    insert into qm_changelog (entity, entity_id, action, field, line_ref, old_value, new_value, actor_email)")
            S.append(
                f"    values ('tadqiq', v_k, 'tadqiq_create', '', '{esc(comp)} — {tdate}', '', "
                f"'{len(sub['linesRaw'])} بند (رصيد افتتاحي)', 'backfill');"
            )
        S.append("  end if;")
    S.append("end $$;")
    out = os.path.join(OUT_DIR, "0036_qm_backfill.sql")
    open(out, "w", encoding="utf-8").write("\n".join(S))
    print(f"SQL → {out}")


if __name__ == "__main__":
    main()
