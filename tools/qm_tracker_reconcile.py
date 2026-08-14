# -*- coding: utf-8 -*-
"""Reconcile the QA's WO tracking report against everything the module holds.

Source: D:\\التجميع الشهري new\\اوامر العمل\\متابعة أوامـر العمــل ونسبة انجاز
والمتبقي من عقد 9.-الجديد.xlsx — one sheet per ministry payment; the latest
sheet is the ministry-facing master (issued value, amended value, executed
payment to date, start/end dates, duration, منتهي/جاري status per WO).

Compares, per work order:
  · issued / amended value   vs the WO line totals we hold (after نسبة العقد)
  · executed payment to date vs our executed (ministry cumulative) and vs the
    Σ of the payment certificates built from the جزئي columns
  · duration + dates         vs the values healed from the scanned amendments
  · منتهي / جاري             vs our `closed` flag (nothing set yet)

Writes Desktop/quantities-backfill/tracker-reconciliation.md
"""
import glob
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

TRACK_DIR = r"D:\التجميع الشهري new\اوامر العمل"
OUT_DIR = r"C:\Users\fszog\Desktop\quantities-backfill"
BOP = json.load(open(os.path.join(OUT_DIR, "wo-ocr", "_bop_ref.json"), encoding="utf-8"))
BACKFILL = json.load(open(os.path.join(OUT_DIR, "qm-backfill-data.json"), encoding="utf-8"))
CERTS = json.load(open(os.path.join(OUT_DIR, "paycerts-data.json"), encoding="utf-8"))
DIFFS = json.load(open(os.path.join(OUT_DIR, "wo-ocr", "_diffs.json"), encoding="utf-8"))
PCT = 1.09

# columns in the per-WO table (0-based), verified against the sheet layout
C_PCT, C_DONE_PAY, C_ONGOING_PAY = 0, 1, 2
C_CANCEL, C_ONGOING, C_DONE = 3, 4, 5
C_PENALTY, C_DURATION, C_STATUS = 6, 7, 8
C_END, C_START = 9, 10
C_VALUE, C_VALUE_AMENDED = 11, 12
C_NAME, C_NO = 13, 17


def latest_sheet(wb):
    """The most recent payment sheet: highest 5-MM-YYYY in the tab name."""
    best, best_key = None, None
    for nm in wb.sheetnames:
        m = re.match(r"\s*5-(\d{1,2})-(\d{4})", nm.strip())
        if not m:
            continue
        key = (int(m.group(2)), int(m.group(1)), len(nm))
        if best_key is None or key[:2] > best_key[:2] or (key[:2] == best_key[:2] and "(2)" in nm):
            best, best_key = nm, key
    return best


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def main():
    path = [f for f in glob.glob(os.path.join(TRACK_DIR, "*.xlsx"))
            if "الجديد" in os.path.basename(f)
            and not os.path.basename(f).startswith("~$")
            and "-2.xlsx" not in os.path.basename(f)][0]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = latest_sheet(wb)
    s = wb[sheet_name]
    rows = list(s.iter_rows(values_only=True))

    # header block
    contract_value = num(rows[4][1])
    change_orders = num(rows[5][1])
    total_value = num(rows[6][1])
    fin_pct = num(rows[7][1])
    time_pct = num(rows[8][1])
    elapsed = num(rows[7][11])
    total_days = num(rows[8][11])
    pay_done = num(rows[9][1])
    pay_ongoing = num(rows[10][1])
    pay_total = num(rows[11][1])
    n_issued = num(rows[9][13])
    n_done = num(rows[10][13])
    n_ongoing = num(rows[11][13])
    n_cancelled = num(rows[12][13])

    track = {}
    for r in rows[16:]:
        if not r or len(r) <= C_NO:
            continue
        no = r[C_NO]                       # stored as text in this workbook
        m = re.match(r"^\s*(\d+)\s*$", str(no or ""))
        if not m:
            continue
        wo = int(m.group(1))
        executed = (num(r[C_DONE_PAY]) or 0) + (num(r[C_ONGOING_PAY]) or 0)
        track[wo] = {
            "name": str(r[C_NAME] or "").strip(),
            "value": num(r[C_VALUE]),
            "value_amended": num(r[C_VALUE_AMENDED]),
            "executed": executed,
            "pct": num(r[C_PCT]),
            "duration": num(r[C_DURATION]),
            "penalty": num(r[C_PENALTY]),
            "start": r[C_START], "end": r[C_END],
            "status": str(r[C_STATUS] or "").strip(),
            "cancelled": bool(r[C_CANCEL]),
        }
    wb.close()

    # ── ours ────────────────────────────────────────────────────────────
    ours = {}
    for key, rec in BACKFILL.items():
        m = re.match(r"\('wo', (\d+)\)", key)
        if not m:
            continue
        wo = int(m.group(1))
        head = rec.get("head") or {}
        wo_val = sum((BOP.get(f"{b}/{n}{sfx}", {}).get("rate", 0) or 0) * q
                     for b, n, sfx, q in (rec.get("linesRaw") or []))
        exec_val = sum((BOP.get(k, {}).get("rate", 0) or 0) * q
                       for k, q in (rec.get("execMinistry") or {}).items())
        ours[wo] = {"woValue": wo_val * PCT, "executed": exec_val * PCT,
                    "duration": head.get("duration"), "registerValue": head.get("registerValue"),
                    "desc": head.get("desc", "")}
    # healed values from the official scans (0038)
    healed = {}
    for wo_s, e in DIFFS.items():
        lat = e.get("latest") or {}
        healed[int(wo_s)] = {"duration": lat.get("duration_days"),
                             "value": lat.get("est_cost_after_pct"),
                             "lines": lat.get("lines") or {}}
    # certificates per WO
    cert_by_wo = {}
    for no, c in CERTS.items():
        for l in c["lines"]:
            cert_by_wo[l["wo"]] = cert_by_wo.get(l["wo"], 0) + l["amount"]
    cert_by_wo = {k: v * PCT for k, v in cert_by_wo.items()}

    R = [f"# WO tracking report vs the module — reconciliation\n",
         f"Source sheet: **{sheet_name}** of `{os.path.basename(path)}` "
         f"(one sheet per ministry payment; this is the newest).\n",
         "## Contract-level\n",
         "| figure | tracking report | what the module holds |",
         "|---|---|---|",
         f"| قيمة العقد | {contract_value:,.0f} | 19,000,000 (as told 2026-08-14) |",
         f"| أوامر تغييرية | {change_orders:,.0f} | not modelled |",
         f"| الإجمالي | {total_value:,.0f} | 19,000,000 |",
         f"| نسبة الإنجاز المالي | {fin_pct*100:.2f}% | derived from executed |",
         f"| نسبة الإنجاز الزمني | {time_pct*100:.2f}% ({elapsed:,.0f} / {total_days:,.0f} يوم) | not modelled |",
         f"| قيمة الدفعة — أوامر منتهية | {pay_done:,.3f} | — |",
         f"| قيمة الدفعة — أوامر جارية | {pay_ongoing:,.3f} | — |",
         f"| **إجمالي المنفَّذ** | **{pay_total:,.3f}** | "
         f"**{sum(o['executed'] for o in ours.values()):,.3f}** (ministry cumulative × 1.09) |",
         f"| عدد أوامر العمل الصادرة | {n_issued:,.0f} | {len([w for w in ours if w < 900])} (+ "
         f"{len([w for w in ours if w >= 900])} placeholders in the 900 range) |",
         f"| منتهية / جارية / ملغاة | {n_done:,.0f} / {n_ongoing:,.0f} / {n_cancelled:,.0f} | "
         f"closed flag not set on any WO yet |",
         "",
         "## Per-work-order\n",
         "Values after نسبة العقد. `Δ exec` = tracking executed − ours; "
         "`Δ value` = tracking amended value − our WO total.\n",
         "| WO | tracking value (amended) | our WO total | Δ value | tracking executed | "
         "our executed | Σ our certs | Δ exec | tracking status | dur (track/doc/ours) |",
         "|---|---|---|---|---|---|---|---|---|---|"]

    only_track, only_ours, big_val, big_exec, dur_diff = [], [], [], [], []
    for wo in sorted(set(track) | set(k for k in ours if k < 900)):
        tr = track.get(wo)
        ou = ours.get(wo)
        if tr and not ou:
            only_track.append(wo)
        if ou and not tr:
            only_ours.append(wo)
        tval = (tr.get("value_amended") or tr.get("value")) if tr else None
        oval = ou["woValue"] if ou else None
        texec = tr["executed"] if tr else None
        oexec = ou["executed"] if ou else None
        cexec = cert_by_wo.get(wo)
        dv = (tval - oval) if (tval is not None and oval) else None
        de = (texec - oexec) if (texec is not None and oexec is not None) else None
        hd = healed.get(wo, {}).get("duration")
        od = ou["duration"] if ou else None
        td = tr["duration"] if tr else None
        if dv is not None and abs(dv) > 1000:
            big_val.append((wo, dv, tval, oval))
        if de is not None and abs(de) > 1000:
            big_exec.append((wo, de, texec, oexec))
        if td is not None and hd is not None and int(td) != int(hd):
            dur_diff.append((wo, int(td), int(hd), od))
        f = lambda v: f"{v:,.3f}" if isinstance(v, (int, float)) else "—"
        R.append(f"| {wo} | {f(tval)} | {f(oval)} | {f(dv)} | {f(texec)} | {f(oexec)} | "
                 f"{f(cexec)} | {f(de)} | {tr['status'] if tr else '—'} | "
                 f"{int(td) if td else '—'}/{hd or '—'}/{od or '—'} |")

    R.append("\n## Main divergences\n")
    R.append(f"- **WOs in the tracking report but not in the module:** "
             f"{sorted(only_track) or 'none'}")
    R.append(f"- **WOs in the module but not in the tracking report:** "
             f"{sorted(only_ours) or 'none'}")
    R.append(f"\n### Executed value — {len(big_exec)} WOs differ by more than KD 1,000")
    R.append("| WO | tracking | ours | Δ |")
    R.append("|---|---|---|---|")
    for wo, de, t, o in sorted(big_exec, key=lambda x: -abs(x[1])):
        R.append(f"| {wo} | {t:,.3f} | {o:,.3f} | {de:,.3f} |")
    R.append(f"\n### WO value — {len(big_val)} WOs differ by more than KD 1,000")
    R.append("| WO | tracking (amended) | ours | Δ |")
    R.append("|---|---|---|---|")
    for wo, dv, t, o in sorted(big_val, key=lambda x: -abs(x[1]))[:40]:
        R.append(f"| {wo} | {t:,.3f} | {o:,.3f} | {dv:,.3f} |")
    R.append(f"\n### Duration — {len(dur_diff)} WOs where the tracker and the scanned amendment disagree")
    R.append("| WO | tracker | latest scan (0038) | pre-heal |")
    R.append("|---|---|---|---|")
    for wo, td, hd, od in dur_diff:
        R.append(f"| {wo} | {td} | {hd} | {od or '—'} |")
    done = [w for w, t in track.items() if "منتهي" in t["status"]]
    R.append(f"\n### Completion status (feeds the `closed` flag)\n")
    R.append(f"- منتهي in the tracker: **{len(done)}** WOs — {sorted(done)}")
    R.append(f"- جاري: {sorted([w for w, t in track.items() if 'جاري' in t['status']])}")

    out = os.path.join(OUT_DIR, "tracker-reconciliation.md")
    open(out, "w", encoding="utf-8").write("\n".join(R) + "\n")
    print(f"tracking WOs: {len(track)} | module WOs: {len(ours)}")
    print(f"tracker executed total: {pay_total:,.3f} | ours: {sum(o['executed'] for o in ours.values()):,.3f}")
    print(f"big exec diffs: {len(big_exec)} | big value diffs: {len(big_val)} | duration diffs: {len(dur_diff)}")
    print(f"only in tracker: {sorted(only_track)}")
    print(f"only in module: {sorted(only_ours)}")
    print("written:", out)


if __name__ == "__main__":
    main()
