#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Master-data dedup tool — merge-and-map, never delete (brief v2 appendix).

Turns the polluted SpectroNova vendor master into (a) canonical `vendors`
rows, (b) many-to-one `vendor_spectronova_ids` mapping rows (old contact ID
-> one canonical vendor, survivor unflagged), and (c) a retirement checklist
for finance. THE TOOL PROPOSES; A HUMAN DISPOSES — zero auto-merges.

Stages (separately runnable, config-driven per master via masters.json):

  python tools/dedup/dedup.py propose --master vendors
      classify (test / non-vendor / foreign-demo / vendor) -> normalize
      (Arabic variant unification, legal-suffix tokens dropped, business
      words kept) -> cluster (exact key, fuzzy token-set two bands, hard
      evidence promotes) -> suggest survivor (invoice history from the
      Supplier Invoice exports, completeness, lowest ID). Emits
      output/dedup-review-vendors.xlsx in the SAME sheet layout as the
      de-facto review file COPRI_dedup_review.xlsx, plus a report .md.

  python tools/dedup/dedup.py apply --master vendors
      reads the decided review workbook (defaults to the repo-root
      COPRI_dedup_review.xlsx that the accountant is already filling) plus
      the vendor export (IDs not mentioned in the workbook pass through as
      standalone canonicals) and emits idempotent SQL targeting the 0013
      schema (vendors + vendor_spectronova_ids — no parallel schema),
      output/retirement-checklist.xlsx, and a summary report .md.
      Exclusions map to a single quarantine vendor so historical imports
      referencing a retired ERP contact never crash.

  python tools/dedup/dedup.py apply --master items
      reads the ITEMS sheet (class/UOM review — no merges until the item
      export lands) and targets the 0020 schema (items, unique on
      vendor_norm(name), + item_spectronova_ids). DECISION doubles as the
      corrected UOM; exclude-words quarantine the item; blank = undecided
      passes through with the suggested UOM. items has no notes column, so
      provenance rides in SQL comments.

Traps this handles (all seen in the real inputs):
  * the review workbook covers an OLDER, larger contact universe than the
    current 625-row export — apply accounts for the union of both;
  * literally duplicated workbook rows (C352 survivor row x3) and repeated
    exclusion IDs — deduplicated, conflicts win by EXCLUDE > MERGE > KEEP;
  * two different groups whose survivor names normalize identically —
    the DB unique index on vendor_norm(name) makes them ONE vendor; the
    where-not-exists guards keep that safe and it is reported, not hidden;
  * blank DECISION = undecided -> standalone canonical vendor, listed in
    the undecided report (nothing is ever silently dropped).

Round-trip acceptance is asserted in code: every input contact ID lands in
exactly ONE of {canonical primary, mapped duplicate, excluded->quarantine,
undecided}; any residual aborts the run before outputs are published.

Cluster IDs are deterministic (sha1 of member IDs) so a partially-filled
review survives a re-propose. Console output is ASCII-only (cp1252 console;
Arabic lives in the xlsx/md outputs — reconcile_jv.py convention).

Usage:  python tools/dedup/dedup.py propose|apply [--master vendors|items]
                [--export FILE] [--review FILE] [--out DIR] [--run-ts ISO]
"""
import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl missing — run: pip install --user openpyxl")

ROOT = Path(__file__).resolve().parents[2]
RAW = ROOT / "Accounting raw data" / "spectronova"
OUT = ROOT / "Accounting raw data" / "output"
CFG_PATH = Path(__file__).resolve().parent / "masters.json"
DEFAULT_REVIEW = ROOT / "COPRI_dedup_review.xlsx"

# console: cp1252-safe (never crash on Arabic; Arabic belongs in the files)
try:
    sys.stdout.reconfigure(errors="replace")
    sys.stderr.reconfigure(errors="replace")
except AttributeError:
    pass

BOLD = Font(bold=True)
FILL_A = PatternFill("solid", start_color="FFF2F2F2")
FILL_B = PatternFill("solid", start_color="FFE2EFDA")


# ─────────────────────────────────────────────────────── normalization ──

TASHKEEL = re.compile("[ً-ْٰـ]")   # harakat, dagger alif, tatweel
AR_VARIANTS = [("ة", "ه"),                   # ة -> ه
               ("ى", "ي"),                   # ى -> ي
               ("أ", "ا"), ("إ", "ا"), ("آ", "ا")]  # أإآ -> ا
NON_WORD = re.compile(r"[^0-9a-z؀-ۿ]+")


def norm_text(s):
    """Casefold + Arabic unification + punctuation->space (superset of the
    DB's vendor_norm(); the extra Arabic folding only affects match keys,
    the SQL still resolves by the DB's own vendor_norm)."""
    s = TASHKEEL.sub("", str(s or "").lower())
    for a, b in AR_VARIANTS:
        s = s.replace(a, b)
    return NON_WORD.sub(" ", s).strip()


def match_tokens(name, cfg):
    """Token set for clustering: legal suffixes dropped, abbreviations
    expanded; business words (Trading, Contracting, ...) are KEPT."""
    legal = set(cfg.get("legal_suffix_tokens", []))
    expand = cfg.get("token_expansions", {})
    toks = [expand.get(t, t) for t in norm_text(name).split()]
    kept = [t for t in toks if t not in legal]
    return set(kept) if kept else set(toks)     # never normalize to nothing


def match_key(name, cfg):
    return " ".join(sorted(match_tokens(name, cfg)))


def similarity(key_a, key_b):
    """Token-set similarity in 0..100 (difflib over sorted token strings)."""
    return SequenceMatcher(None, key_a, key_b).ratio() * 100.0


def canon_id(v):
    """Contact IDs arrive as int/float/str — normalize to a plain string."""
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return str(v).strip()


def id_sort_key(cid):
    return (0, int(cid)) if cid.isdigit() else (1, cid)


def sql_quote(s):
    return "'" + str(s).replace("'", "''") + "'"


# ─────────────────────────────────────────────────────── excel helpers ──

def read_table(path, id_candidates):
    """Read the first sheet; locate the header row by any id candidate
    (exports sometimes carry title rows above it). Returns (header, dicts)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    for i, r in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in r]
        if any(c in id_candidates for c in cells):
            return cells, [dict(zip(cells, rr)) for rr in rows[i + 1:]
                           if any(c is not None and str(c).strip() for c in rr)]
    raise ValueError("%s: header row not found (looked for %s)" % (path.name, id_candidates))


def pick_column(header, candidates):
    for c in candidates:
        if c in header:
            return c
    return None


def sheet_rows(wb, wanted):
    """Find a sheet by fuzzy name ('READ ME'/'README' both work)."""
    target = re.sub(r"[^A-Z]", "", wanted.upper())
    for name in wb.sheetnames:
        if re.sub(r"[^A-Z]", "", name.upper()) == target:
            return [list(r) for r in wb[name].iter_rows(values_only=True)]
    return None


def style_sheet(ws, widths, wrap_cols=()):
    for c in ws[1]:
        c.font = BOLD
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in wrap_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(ws.max_column), ws.max_row)


def add_dropdown(ws, col_letter, options, last_row):
    if last_row < 2:
        return
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(options),
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add("%s2:%s%d" % (col_letter, col_letter, last_row))


# ─────────────────────────────────────────────────────── classification ──

def classify(raw_name, cfg):
    """-> (klass, reason). klass in {'test','non-vendor','foreign-demo','vendor'}.
    A suggestion column — human-confirmable like merges, never auto-applied."""
    raw = str(raw_name or "").strip()
    n = norm_text(raw)
    if not n:
        return "test", "empty/garbage name"
    for pat in cfg.get("test_patterns", []):
        if re.search(pat, n):
            return "test", "test pattern '%s'" % pat
    if cfg.get("non_vendor_name_dash_prefix") and re.match(r"^\s*\d{3,4}\s*-", raw):
        return "non-vendor", "cost-center prefix"
    if n in set(cfg.get("non_vendor_exact", [])):
        return "non-vendor", "internal org unit"
    for sub in cfg.get("non_vendor_substrings", []):
        if sub in n:
            return "non-vendor", "gl-concept/internal ('%s')" % sub
    for pat in cfg.get("non_vendor_regex", []):
        if re.search(pat, n):
            return "non-vendor", "job/GL account pattern"
    for sub in cfg.get("foreign_demo_substrings", []):
        if sub in n:
            return "foreign-demo", "demo keyword ('%s')" % sub
    return "vendor", ""


# ─────────────────────────────────────────────────────────── union-find ──

class UnionFind:
    def __init__(self):
        self.parent = {}

    def add(self, x):
        self.parent.setdefault(x, x)

    def find(self, x):
        self.add(x)
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[rb] = ra

    def groups(self):
        g = defaultdict(list)
        for x in self.parent:
            g[self.find(x)].append(x)
        return list(g.values())


def cluster_hash(ids):
    """Deterministic cluster ID: stable across re-runs for the same members."""
    return "C" + hashlib.sha1(",".join(sorted(ids)).encode("utf-8")).hexdigest()[:8]


# ══════════════════════════════════════════════════════════ inputs ══

def find_export(cfg, override):
    if override:
        p = Path(override)
        return p if p.exists() else None
    for glob in cfg["export_globs"]:
        hits = sorted(RAW.glob(glob), key=lambda p: p.stat().st_mtime, reverse=True)
        if hits:
            return hits[0]
    return None


def load_export(path, cfg, notes):
    """-> (records, found_columns). record: {id, name, extras{field:value}}."""
    header, rows = read_table(path, cfg["columns"]["id"])
    cols, missing = {}, []
    for field, candidates in cfg["columns"].items():
        col = pick_column(header, candidates)
        if col:
            cols[field] = col
        else:
            missing.append(field)
    if missing:
        notes.append("export columns not found (handled, reported): %s — header was %s"
                     % (", ".join(missing), header))
    records, seen = [], {}
    for i, r in enumerate(rows, 2):
        cid = canon_id(r.get(cols.get("id"), ""))
        name = str(r.get(cols.get("name"), "") or "").strip()
        if not cid:
            notes.append("export row %d: blank ID skipped (name=%r)" % (i, name))
            continue
        if cid in seen:
            notes.append("export: DUPLICATE contact ID %s (rows %d and %d) — first kept, "
                         "second reported" % (cid, seen[cid], i))
            continue
        seen[cid] = i
        extras = {}
        for field in cfg.get("hard_evidence", []):
            if field in cols:
                v = str(r.get(cols[field], "") or "").strip()
                if v:
                    extras[field] = v
        records.append({"id": cid, "name": name, "row": i, "extras": extras})
    return records, cols


def load_invoice_counts(cfg, notes):
    """Transaction-history evidence: invoices per normalized vendor name
    (the AP exports carry names, not contact IDs — joinable by name only)."""
    counts = Counter()
    files = []
    for glob in cfg.get("evidence_globs", []):
        files += sorted(RAW.glob(glob))
    for path in dict.fromkeys(files):        # dedupe, keep order
        try:
            header, rows = read_table(path, cfg.get("evidence_name_columns", []))
        except ValueError as e:
            notes.append("evidence skipped: %s" % e)
            continue
        col = pick_column(header, cfg.get("evidence_name_columns", []))
        if not col:
            notes.append("evidence %s: no vendor-name column — skipped" % path.name)
            continue
        n = 0
        for r in rows:
            name = norm_text(r.get(col))
            if name:
                counts[name] += 1
                n += 1
        notes.append("evidence %s: %d invoice rows joined by vendor name" % (path.name, n))
    return counts


# ══════════════════════════════════════════════════════════ propose ══

def propose(master, cfg, args):
    notes, ts = [], args.run_ts
    export = find_export(cfg, args.export)
    if export is None:
        print("%s export not found under %s (globs: %s)."
              % (cfg["label"], RAW, ", ".join(cfg["export_globs"])))
        print("The %s config ships in masters.json -- re-run propose when the export lands."
              % master)
        return 0
    print("export      : %s" % export)
    records, cols = load_export(export, cfg, notes)
    inv_counts = load_invoice_counts(cfg, notes) if cfg.get("evidence_globs") else Counter()
    hard_fields = [f for f in cfg.get("hard_evidence", []) if f in cols]
    absent_hard = [f for f in cfg.get("hard_evidence", []) if f not in cols]
    if absent_hard:
        notes.append("hard-evidence columns absent in this export: %s — promotion by "
                     "phone/CR/email activates when a richer export lands" % ", ".join(absent_hard))

    # ── stage 1a: classify ────────────────────────────────────────────
    excl, vendors = [], []
    for rec in records:
        klass, reason = classify(rec["name"], cfg)
        rec["class"], rec["reason"] = klass, reason
        (vendors if klass == "vendor" else excl).append(rec)

    # ── stage 1b: normalize + cluster ─────────────────────────────────
    for rec in vendors:
        rec["key"] = match_key(rec["name"], cfg)
        rec["inv"] = inv_counts.get(norm_text(rec["name"]), 0)
    uf = UnionFind()
    for rec in vendors:
        uf.add(rec["id"])
    by_key = defaultdict(list)
    for rec in vendors:
        by_key[rec["key"]].append(rec)
    for key, members in by_key.items():        # exact normalized key
        for m in members[1:]:
            uf.union(members[0]["id"], m["id"])

    # hard evidence (same phone / CR / email / AR account) promotes + bridges EN<->AR
    hard_edges = 0
    for field in hard_fields:
        by_val = defaultdict(list)
        for rec in vendors:
            v = rec["extras"].get(field)
            if v:
                by_val[norm_text(v) or v] += [rec]
        for v, members in by_val.items():
            if len(members) > 1:
                for m in members[1:]:
                    uf.union(members[0]["id"], m["id"])
                    m.setdefault("evidence", []).append("same %s" % field)
                hard_edges += len(members) - 1

    # fuzzy between distinct keys: candidates share a token (len>=3) or a
    # 4-char latin prefix — then two similarity bands
    keys = sorted(by_key)
    tok_index = defaultdict(set)
    for ki, key in enumerate(keys):
        for t in key.split():
            if len(t) >= 3:
                tok_index[t].add(ki)
            if len(t) >= 4 and t.isascii():
                tok_index["#" + t[:4]].add(ki)
    cand_pairs = set()
    for bucket in tok_index.values():
        b = sorted(bucket)
        for i in range(len(b)):
            for j in range(i + 1, len(b)):
                cand_pairs.add((b[i], b[j]))
    high, low = cfg["high_band"], cfg["low_band"]
    fuzzy_edges, low_pairs = [], []
    for i, j in sorted(cand_pairs):
        sim = similarity(keys[i], keys[j])
        if sim >= high:
            fuzzy_edges.append((keys[i], keys[j], sim))
        elif sim >= low:
            low_pairs.append((keys[i], keys[j], sim))
    for ka, kb, sim in fuzzy_edges:
        uf.union(by_key[ka][0]["id"], by_key[kb][0]["id"])
        for m in by_key[kb]:
            m.setdefault("evidence", []).append("fuzzy %.1f vs '%s'" % (sim, by_key[ka][0]["name"].strip()))

    by_id = {rec["id"]: rec for rec in vendors}
    groups = [sorted(g, key=id_sort_key) for g in uf.groups()]
    clusters = [g for g in groups if len(g) > 1]
    singletons = [g[0] for g in groups if len(g) == 1]
    clusters.sort(key=lambda g: id_sort_key(g[0]))

    # ── stage 1c: survivor suggestion per cluster ─────────────────────
    def survivor_rank(rec):
        completeness = sum(1 for v in rec["extras"].values() if v) + (1 if rec["name"] else 0)
        return (-rec["inv"], -completeness, id_sort_key(rec["id"]))

    cluster_rows = []
    for g in clusters:
        members = sorted((by_id[cid] for cid in g), key=survivor_rank)
        surv = members[0]["id"]
        cid_label = cluster_hash(g)
        for cid in g:
            rec = by_id[cid]
            ev = ["exact key"] if rec["key"] == by_id[surv]["key"] and cid != surv else []
            ev += rec.get("evidence", [])
            cluster_rows.append({
                "cluster": cid_label, "id": cid, "name": rec["name"],
                "account": rec["extras"].get("account", ""),
                "survivor": cid == surv,
                "suggested": "SURVIVOR" if cid == surv else "MERGE",
                "confidence": "high",
                "evidence": "; ".join(ev) or ("suggested survivor: %d invoices" % rec["inv"]
                                              if cid == surv else "exact key"),
                "inv": rec["inv"],
            })

    # low-band pairs between final clusters/singletons (dedup across groups)
    root_of = {rec["id"]: uf.find(rec["id"]) for rec in vendors}
    poss_rows, seen_pair = [], set()
    for ka, kb, sim in sorted(low_pairs, key=lambda t: -t[2]):
        ra, rb = root_of[by_key[ka][0]["id"]], root_of[by_key[kb][0]["id"]]
        if ra == rb or (min(ra, rb), max(ra, rb)) in seen_pair:
            continue
        seen_pair.add((min(ra, rb), max(ra, rb)))
        ids_a = sorted((r["id"] for r in by_key[ka]), key=id_sort_key)
        ids_b = sorted((r["id"] for r in by_key[kb]), key=id_sort_key)
        poss_rows.append({
            "a": "%s  (%s)" % (by_key[ka][0]["name"].strip(), ", ".join(ids_a)),
            "b": "%s  (%s)" % (by_key[kb][0]["name"].strip(), ", ".join(ids_b)),
            "sim": sim,
        })

    # ── round-trip accounting (propose): every export ID in exactly 1 bucket
    b_cluster = {r["id"] for r in cluster_rows}
    b_single = set(singletons)
    b_excl = {r["id"] for r in excl}
    universe = {r["id"] for r in records}
    assert not (b_cluster & b_single) and not (b_cluster & b_excl) and not (b_single & b_excl)
    residual = universe - b_cluster - b_single - b_excl
    if residual:
        raise AssertionError("round-trip residual (propose): %s" % sorted(residual)[:10])

    # ── outputs ───────────────────────────────────────────────────────
    OUT.mkdir(exist_ok=True)
    review_path = OUT / cfg["review_workbook"]
    write_review_workbook(review_path, master, cfg, export, ts, cluster_rows, poss_rows, excl)
    report_path = OUT / ("dedup-propose-%s-report.md" % master)
    src = export.name
    rep = ["# Dedup propose — %s — %s" % (cfg["label"], ts),
           "",
           "Source export: `%s` (mtime %s) — %d records, %d unique IDs."
           % (src, datetime.fromtimestamp(export.stat().st_mtime).isoformat(timespec="seconds"),
              len(records), len(universe)),
           "Columns found: %s." % ", ".join("%s=`%s`" % kv for kv in cols.items()),
           "",
           "## Classification (suggestions — confirm in the EXCLUSIONS sheet)",
           ""]
    kc = Counter(r["class"] for r in records)
    for k in ("vendor", "non-vendor", "test", "foreign-demo"):
        rep.append("- %s: %d" % (k, kc.get(k, 0)))
    rep += ["",
            "## Clustering",
            "",
            "- clusters proposed (>= 2 members): %d covering %d records" % (len(clusters), len(b_cluster)),
            "- standalone vendors (no duplicate suspected): %d — not in the workbook; apply "
            "passes them through from the export" % len(singletons),
            "- possible matches (low band %.0f–%.0f, review carefully): %d" % (low, high, len(poss_rows)),
            "- hard-evidence promotions: %d%s" % (hard_edges, "" if hard_fields else " (no phone/CR/email columns in this export)"),
            "",
            "## Round-trip accounting (asserted)",
            "",
            "| bucket | IDs |", "|---|---|",
            "| clustered vendor | %d |" % len(b_cluster),
            "| standalone vendor | %d |" % len(b_single),
            "| suggested exclusion | %d |" % len(b_excl),
            "| **total = export universe** | **%d** |" % len(universe),
            ""]
    if notes:
        rep += ["## Notes & flags", ""] + ["- %s" % n for n in notes] + [""]
    rep += ["## Next step", "",
            "Review `%s` (or keep filling the existing `COPRI_dedup_review.xlsx`), then run:" % review_path.name,
            "", "    python tools/dedup/dedup.py apply --master %s" % master, ""]
    report_path.write_text("\n".join(rep), encoding="utf-8")

    print("records     : %d  (vendor %d / non-vendor %d / test %d / demo %d)"
          % (len(records), kc.get("vendor", 0), kc.get("non-vendor", 0),
             kc.get("test", 0), kc.get("foreign-demo", 0)))
    print("clusters    : %d (covering %d records), singletons %d, possible pairs %d"
          % (len(clusters), len(b_cluster), len(singletons), len(poss_rows)))
    print("round-trip  : OK (%d = %d + %d + %d)"
          % (len(universe), len(b_cluster), len(b_single), len(b_excl)))
    print("review      -> %s" % review_path)
    print("report      -> %s" % report_path)
    return 0


def write_review_workbook(path, master, cfg, export, ts, cluster_rows, poss_rows, excl):
    """Same sheet layout as the de-facto COPRI_dedup_review.xlsx so there is
    ONE review format; suggestion columns ride after DECISION."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "READ ME"
    for line in (
        "COPRI %s — dedup review (generated %s from %s)" % (cfg["label"], ts, export.name),
        "",
        "HOW TO REVIEW",
        "1. CLUSTERS sheet: each group is one suspected real vendor. DECISION per row:",
        "   MERGE = same vendor as the group survivor. KEEP SEPARATE = different vendor.",
        "   The suggested survivor row has SURVIVOR = YES; to change it, move the YES",
        "   (or set DECISION = SURVIVOR on the right row). EXCLUDE = not a vendor at all.",
        "2. POSSIBLE MATCHES sheet: lower-confidence pairs. DECISION: MERGE or KEEP SEPARATE.",
        "3. EXCLUSIONS sheet: suggested test / internal / GL-concept records.",
        "   DECISION: CONFIRM EXCLUDE or KEEP (= it IS a real vendor).",
        "4. Blank decision = not yet reviewed (passes through unchanged, listed as undecided).",
        "Records NOT in this workbook are standalone vendors — nothing to decide.",
        "When done, run: python tools/dedup/dedup.py apply --master %s" % master,
        "Nothing is deleted anywhere; merged records become mapped aliases of the survivor.",
    ):
        ws.append((line,))
    ws.column_dimensions["A"].width = 100
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    ws = wb.create_sheet("CLUSTERS")
    ws.append(("Cluster", "Source", "Contact ID", "Company name", "Legacy group",
               "Survivor", "DECISION", "Suggested action", "Confidence", "Evidence", "Invoices"))
    fill, last = FILL_A, None
    for r in cluster_rows:
        ws.append((r["cluster"], export.name, r["id"], r["name"], r["account"],
                   "YES" if r["survivor"] else None, None,
                   r["suggested"], r["confidence"], r["evidence"], r["inv"]))
        if r["cluster"] != last:
            fill = FILL_B if fill is FILL_A else FILL_A
            last = r["cluster"]
        for c in ws[ws.max_row]:
            c.fill = fill
        if r["survivor"]:
            for c in ws[ws.max_row]:
                c.font = BOLD
    style_sheet(ws, (11, 20, 12, 55, 18, 9, 16, 16, 11, 45, 9), wrap_cols=(10,))
    add_dropdown(ws, "G", ["MERGE", "SURVIVOR", "KEEP SEPARATE", "EXCLUDE"], ws.max_row)

    ws = wb.create_sheet("POSSIBLE MATCHES")
    ws.append(("Pair", "Name A (IDs)", "Name B (IDs)", "Similarity", "DECISION"))
    for i, r in enumerate(poss_rows, 1):
        ws.append(("P%d" % i, r["a"], r["b"], r["sim"], None))
    style_sheet(ws, (8, 60, 60, 12, 16), wrap_cols=(2, 3))
    add_dropdown(ws, "E", ["MERGE", "KEEP SEPARATE"], ws.max_row)

    ws = wb.create_sheet("EXCLUSIONS")
    ws.append(("Source", "Contact ID", "Company name", "Reason", "DECISION"))
    for r in sorted(excl, key=lambda r: (r["class"], id_sort_key(r["id"]))):
        ws.append((export.name, r["id"], r["name"],
                   "%s — %s" % (r["class"], r["reason"]), None))
    style_sheet(ws, (20, 12, 55, 40, 18))
    add_dropdown(ws, "E", ["CONFIRM EXCLUDE", "KEEP"], ws.max_row)

    ws = wb.create_sheet("ITEMS")
    ws.append(("ItemID", "ItemCode", "Description", "Group", "Suggested class",
               "Current UOM", "Suggested UOM", "DECISION / corrected UOM"))
    style_sheet(ws, (10, 14, 45, 22, 14, 12, 13, 22))

    wb.save(path)


# ══════════════════════════════════════════════════════════ apply ══

DEC_MERGE = {"MERGE", "MERGE INTO SURVIVOR", "SAME", "DUPLICATE"}
DEC_SURVIVOR = {"SURVIVOR", "KEEP AS SURVIVOR"}
DEC_KEEP = {"KEEP", "KEEP SEPARATE", "SEPARATE", "DIFFERENT", "NOT A DUPLICATE"}
DEC_EXCLUDE = {"EXCLUDE", "CONFIRM EXCLUDE", "CONFIRM-EXCLUDE", "EXCLUDE-TEST",
               "EXCLUDE-NON-VENDOR", "EXCLUDE-DEMO", "CONFIRM"}
PAIR_IDS = re.compile(r"\(([0-9,\s]+)\)\s*$")


def dec_norm(v):
    return re.sub(r"\s+", " ", str(v or "").strip().upper())


def parse_pair_side(text):
    """'Name A  (id1, id2)' -> (name, [ids])."""
    s = str(text or "").strip()
    m = PAIR_IDS.search(s)
    if not m:
        return s, []
    ids = [canon_id(x) for x in m.group(1).split(",") if x.strip()]
    return s[: m.start()].strip(), ids


def apply_stage(master, cfg, args):
    ts, notes, conflicts = args.run_ts, [], []
    review = Path(args.review) if args.review else (
        DEFAULT_REVIEW if DEFAULT_REVIEW.exists() else OUT / cfg["review_workbook"])
    if not review.exists():
        sys.exit("review workbook not found: %s — run propose first" % review)
    review_mtime = datetime.fromtimestamp(review.stat().st_mtime).isoformat(timespec="seconds")
    print("review      : %s (mtime %s)" % (review, review_mtime))

    if master == "items":
        return apply_items(cfg, args, review, review_mtime, notes)
    if master != "vendors":
        sys.exit("apply --master %s: no canonical schema wired for this master yet" % master)

    export = find_export(cfg, args.export)
    export_recs = []
    if export is not None:
        export_recs, _cols = load_export(export, cfg, notes)
        print("export      : %s (%d records)" % (export, len(export_recs)))
    else:
        notes.append("vendor export not found — apply covered only the workbook IDs")

    wb = openpyxl.load_workbook(review, read_only=True, data_only=True)
    cl_rows = sheet_rows(wb, "CLUSTERS") or []
    pm_rows = sheet_rows(wb, "POSSIBLE MATCHES") or []
    ex_rows = sheet_rows(wb, "EXCLUSIONS") or []
    wb.close()

    def cols_of(rows, wanted):
        if not rows:
            return {}
        hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
        out = {}
        for k, cands in wanted.items():
            for c in cands:
                hit = [i for i, h in enumerate(hdr) if h.upper().startswith(c.upper())]
                if hit:
                    out[k] = hit[0]
                    break
        return out

    def cell(row, idx):
        """Ragged-row-safe cell access (read-only sheets can cut rows short)."""
        return row[idx] if idx is not None and idx < len(row) else None

    # ── read CLUSTERS ─────────────────────────────────────────────────
    cc = cols_of(cl_rows, {"cluster": ["Cluster"], "source": ["Source"],
                           "id": ["Contact ID"], "name": ["Company name"],
                           "surv": ["Survivor"], "dec": ["DECISION"]})
    names, sources, provenance = {}, {}, {}
    clusters = defaultdict(list)          # label -> [(id, decision, survivor?, row)]
    seen_cluster_row = set()
    for i, r in enumerate(cl_rows[1:], 2):
        v = cell(r, cc.get("id"))
        cid = canon_id(v) if v is not None else ""
        if not cid:
            continue
        label = str(cell(r, cc.get("cluster")) or "").strip()
        dec = dec_norm(cell(r, cc.get("dec")))
        surv = dec_norm(cell(r, cc.get("surv"))) == "YES"
        key = (label, cid, dec, surv)
        if key in seen_cluster_row:       # literally duplicated rows (seen: C352 x3)
            notes.append("CLUSTERS row %d: exact duplicate of an earlier row (%s / %s) — ignored"
                         % (i, label, cid))
            continue
        seen_cluster_row.add(key)
        names.setdefault(cid, str(cell(r, cc.get("name")) or "").strip())
        sources.setdefault(cid, str(cell(r, cc.get("source")) or "").strip())
        provenance.setdefault(cid, "CLUSTERS!row %d" % i)
        clusters[label].append((cid, dec, surv, i))

    # ── read EXCLUSIONS ───────────────────────────────────────────────
    ec = cols_of(ex_rows, {"source": ["Source"], "id": ["Contact ID"],
                           "name": ["Company name"], "reason": ["Reason"], "dec": ["DECISION"]})
    excluded, excl_keep, excl_undecided = {}, set(), {}
    for i, r in enumerate(ex_rows[1:], 2):
        v = cell(r, ec.get("id"))
        cid = canon_id(v) if v is not None else ""
        if not cid:
            continue
        dec = dec_norm(cell(r, ec.get("dec")))
        reason = str(cell(r, ec.get("reason")) or "").strip()
        names.setdefault(cid, str(cell(r, ec.get("name")) or "").strip())
        sources.setdefault(cid, str(cell(r, ec.get("source")) or "").strip())
        if cid in excluded or cid in excl_keep or cid in excl_undecided:
            notes.append("EXCLUSIONS row %d: contact %s listed again — first row kept" % (i, cid))
            continue
        provenance.setdefault(cid, "EXCLUSIONS!row %d" % i)
        if dec in DEC_EXCLUDE:
            excluded[cid] = (reason, i)
        elif dec in DEC_KEEP:
            excl_keep.add(cid)
        elif dec == "":
            excl_undecided[cid] = i
        else:
            notes.append("EXCLUSIONS row %d: unknown DECISION %r — treated as undecided" % (i, dec))
            excl_undecided[cid] = i

    # CLUSTERS rows the accountant marked EXCLUDE outright join the excluded set
    for label, members in sorted(clusters.items()):
        for cid, dec, _s, row in members:
            if dec in DEC_EXCLUDE and cid not in excluded:
                excluded[cid] = ("excluded in CLUSTERS %s" % label, row)
                excl_undecided.pop(cid, None)
                excl_keep.discard(cid)

    # ── build groups from decided clusters ────────────────────────────
    uf = UnionFind()
    survivor_hint, undecided = {}, {}      # id -> priority / id -> why
    for label, members in sorted(clusters.items()):
        surv_ids = [cid for cid, dec, surv, _row in members
                    if (surv or dec in DEC_SURVIVOR) and cid not in excluded]
        merges = [cid for cid, dec, _s, _row in members if dec in DEC_MERGE and cid not in excluded]
        for cid, dec, _s, row in members:
            if cid in excluded:
                if dec in DEC_MERGE or dec in DEC_SURVIVOR:
                    conflicts.append("%s: contact %s is both clustered (%s) and confirmed "
                                     "excluded — exclusion wins" % (label, cid, dec))
                continue
            uf.add(cid)
            if dec in DEC_KEEP:
                pass                                        # standalone by decision
            elif dec == "" :
                undecided[cid] = "%s: blank DECISION" % label
            elif dec in DEC_MERGE or dec in DEC_SURVIVOR:
                pass                                        # unioned below
            else:
                notes.append("CLUSTERS row %d: unknown DECISION %r — treated as undecided" % (row, dec))
                undecided[cid] = "%s: unknown decision %r" % (label, dec)
        if merges and not surv_ids:
            conflicts.append("%s: MERGE rows but no survivor — whole cluster left undecided" % label)
            for cid in merges:
                undecided[cid] = "%s: merge without survivor" % label
            continue
        if len(set(surv_ids)) > 1:
            conflicts.append("%s: %d survivor rows — first (%s) kept"
                             % (label, len(set(surv_ids)), sorted(set(surv_ids), key=id_sort_key)[0]))
        if surv_ids:
            surv = sorted(set(surv_ids), key=id_sort_key)[0]
            survivor_hint[surv] = 1
            for cid in merges:
                if cid not in undecided:
                    uf.union(surv, cid)

    # ── decided POSSIBLE MATCHES pairs bridge groups ──────────────────
    pc = cols_of(pm_rows, {"pair": ["Pair"], "a": ["Name A"], "b": ["Name B"], "dec": ["DECISION"]})
    pair_merges = pair_pending = 0
    for i, r in enumerate(pm_rows[1:], 2):
        dec = dec_norm(cell(r, pc.get("dec")))
        if dec not in DEC_MERGE:
            if dec == "" and any(c is not None for c in r):
                pair_pending += 1      # advisory: both sides stay as decided elsewhere
            elif dec and dec not in DEC_KEEP:
                notes.append("POSSIBLE MATCHES row %d: unknown DECISION %r — ignored" % (i, dec))
            continue
        name_a, ids_a = parse_pair_side(cell(r, pc.get("a")))
        name_b, ids_b = parse_pair_side(cell(r, pc.get("b")))
        ids_a = [c for c in ids_a if c not in excluded]
        ids_b = [c for c in ids_b if c not in excluded]
        if not ids_a or not ids_b:
            notes.append("POSSIBLE MATCHES row %d: no usable IDs on one side — skipped" % i)
            continue
        for cid, nm in [(c, name_a) for c in ids_a] + [(c, name_b) for c in ids_b]:
            uf.add(cid)
            names.setdefault(cid, nm)
            provenance.setdefault(cid, "POSSIBLE MATCHES!row %d" % i)
            undecided.pop(cid, None)       # an explicit pair decision beats blank
        survivor_hint[ids_a[0]] = survivor_hint.get(ids_a[0], 2)
        for cid in ids_a[1:] + ids_b:
            uf.union(ids_a[0], cid)
        pair_merges += 1

    # ── export pass-through: IDs the workbook never mentions ──────────
    export_only = 0
    for rec in export_recs:
        cid = rec["id"]
        names.setdefault(cid, rec["name"])
        sources.setdefault(cid, export.name)
        provenance.setdefault(cid, "%s!row %d" % (export.name, rec["row"]))
        if cid not in excluded and cid not in undecided and cid not in excl_undecided:
            if cid not in uf.parent:
                export_only += 1
            uf.add(cid)

    for cid in list(excl_undecided):       # undecided exclusions pass through standalone
        if cid not in excluded:
            uf.add(cid)
            undecided[cid] = "EXCLUSIONS: blank DECISION"
    for cid in excl_keep:
        uf.add(cid)

    # undecided ids must stay singletons (nothing merges an undecided row)
    groups = []
    for g in uf.groups():
        dec_members = [c for c in g if c not in undecided]
        groups += [[c] for c in g if c in undecided]
        if dec_members:
            groups.append(sorted(dec_members, key=id_sort_key))
    groups.sort(key=lambda g: id_sort_key(g[0]))

    # ── survivor per final group + name collisions ────────────────────
    final = []                             # {survivor, members, name, undecided}
    for g in groups:
        hints = sorted((c for c in g if c in survivor_hint),
                       key=lambda c: (survivor_hint[c], id_sort_key(c)))
        surv = hints[0] if hints else g[0]
        if len(hints) > 1:
            conflicts.append("group %s: several survivors after pair-merge (%s) — %s kept"
                             % (cluster_hash(g), ", ".join(hints), surv))
        nm = (names.get(surv) or "").strip()
        final.append({"surv": surv, "members": g, "name": nm,
                      "undecided": surv in undecided})
    by_norm = defaultdict(list)
    for grp in final:
        key = norm_text(grp["name"])
        if key:
            by_norm[key].append(grp)
    norm_collisions = {k: v for k, v in by_norm.items() if len(v) > 1}
    for k, grps in sorted(norm_collisions.items()):
        notes.append("normalized-name collision '%s': groups %s become ONE vendor row in the "
                     "DB (unique index on vendor_norm) — their mappings share it"
                     % (k, ", ".join(g["surv"] for g in grps)))

    # ── round-trip accounting: exactly one bucket per ID ──────────────
    b_primary, b_mapped = set(), set()
    b_und = set(undecided) | (set(excl_undecided) - set(excluded))
    for grp in final:
        if grp["undecided"]:
            continue
        b_primary.add(grp["surv"])
        b_mapped |= set(grp["members"]) - {grp["surv"]}
    b_excl = set(excluded)
    universe = set(names) | {r["id"] for r in export_recs}
    buckets = [("canonical primary", b_primary), ("mapped duplicate", b_mapped),
               ("excluded -> quarantine", b_excl), ("undecided (standalone canonical)", b_und)]
    for i in range(len(buckets)):
        for j in range(i + 1, len(buckets)):
            overlap = buckets[i][1] & buckets[j][1]
            if overlap:
                raise AssertionError("bucket overlap %s/%s: %s"
                                     % (buckets[i][0], buckets[j][0], sorted(overlap)[:10]))
    residual = universe - b_primary - b_mapped - b_excl - b_und
    if residual:
        raise AssertionError("round-trip residual (apply): %s" % sorted(residual)[:10])

    # ── SQL (idempotent: where-not-exists guards; safe to run twice) ──
    OUT.mkdir(exist_ok=True)
    quarantine = cfg["quarantine_name"]
    prov = "dedup apply %s (mtime %s) — run %s" % (review.name, review_mtime, ts)
    sql = ["-- %s — generated by tools/dedup/dedup.py" % prov,
           "-- Idempotent: every insert is guarded (where not exists); re-runs are no-ops.",
           "-- Targets 0013: vendors (unique on vendor_norm(name)) + vendor_spectronova_ids.",
           "-- Merge-and-map, never delete: no update/delete statements in this file.",
           "begin;", "",
           "-- quarantine bucket: excluded ERP contacts resolve here so historical",
           "-- imports referencing them never crash (reported, not posted)",
           "insert into vendors (name, kind, active, notes)",
           "  select %s, 'other', false, %s" % (sql_quote(quarantine), sql_quote(prov)),
           "  where not exists (select 1 from vendors where vendor_norm(name) = vendor_norm(%s));"
           % sql_quote(quarantine), ""]
    skipped_unnamed = []

    def mapping_sql(vendor_name, contact_id, flagged):
        return ("insert into vendor_spectronova_ids (vendor_id, contact_id, flagged)\n"
                "  select (select id from vendors where vendor_norm(name) = vendor_norm(%s) limit 1),\n"
                "         %s, %s\n"
                "  where not exists (select 1 from vendor_spectronova_ids where contact_id = %s);"
                % (sql_quote(vendor_name), sql_quote(contact_id),
                   "true" if flagged else "false", sql_quote(contact_id)))

    n_vendor_rows = n_map_rows = 0
    for grp in sorted(final, key=lambda g: id_sort_key(g["surv"])):
        nm = grp["name"]
        if not norm_text(nm):
            skipped_unnamed.append(grp["surv"])
            continue
        tag = "undecided — passes through, review later" if grp["undecided"] else \
              ("survivor of %d" % len(grp["members"]) if len(grp["members"]) > 1 else "standalone")
        sql.append("-- %s [%s] %s — %s" % (grp["surv"], tag, nm, provenance.get(grp["surv"], "")))
        sql.append("insert into vendors (name, kind, notes)")
        sql.append("  select %s, 'supplier', %s" % (sql_quote(nm), sql_quote(prov)))
        sql.append("  where not exists (select 1 from vendors where vendor_norm(name) = vendor_norm(%s));"
                   % sql_quote(nm))
        n_vendor_rows += 1
        for cid in grp["members"]:
            sql.append(mapping_sql(nm, cid, flagged=(cid != grp["surv"])))
            n_map_rows += 1
        sql.append("")
    if excluded:
        sql.append("-- excluded contacts -> quarantine (flagged; never a canonical vendor)")
        for cid in sorted(excluded, key=id_sort_key):
            sql.append("-- %s %s — %s (%s)" % (cid, names.get(cid, ""), excluded[cid][0],
                                               provenance.get(cid, "")))
            sql.append(mapping_sql(quarantine, cid, flagged=True))
            n_map_rows += 1
        sql.append("")
    for cid in skipped_unnamed:
        notes.append("group %s has an empty/garbage survivor name — NO SQL emitted for it "
                     "(listed as undecided in the checklist)" % cid)
    sql += ["commit;", ""]
    sql_path = OUT / cfg["apply_sql"]
    sql_path.write_text("\n".join(sql), encoding="utf-8")

    # ── retirement checklist ──────────────────────────────────────────
    check_path = OUT / "retirement-checklist.xlsx"
    write_checklist(check_path, ts, review.name, final, names, sources, provenance,
                    excluded, undecided, excl_undecided, quarantine)

    # ── summary report ────────────────────────────────────────────────
    n_merged_groups = sum(1 for g in final if len(g["members"]) > 1)
    before, after = len(universe), len([g for g in final if norm_text(g["name"])]) \
        - sum(len(v) - 1 for v in norm_collisions.values()) + (1 if excluded else 0)
    rep = ["# Dedup apply — %s — %s" % (cfg["label"], ts), "",
           "Review workbook: `%s` (mtime %s)." % (review.name, review_mtime),
           "Vendor export: `%s`." % (export.name if export else "(not found)"),
           "",
           "## Counts per action", "",
           "| action | count |", "|---|---|",
           "| clusters applied (merge groups) | %d |" % n_merged_groups,
           "| canonical vendors emitted (SQL insert-if-missing) | %d |" % n_vendor_rows,
           "| mapping rows emitted | %d |" % n_map_rows,
           "| merged (retired) contact IDs | %d |" % len(b_mapped),
           "| excluded -> quarantine | %d |" % len(b_excl),
           "| undecided (standalone canonical, review later) | %d |" % len(b_und),
           "| possible-match pairs applied | %d |" % pair_merges,
           "| possible-match pairs still blank (advisory — both sides kept as decided elsewhere) | %d |" % pair_pending,
           "| export IDs not in the workbook (pass-through) | %d |" % export_only,
           "",
           "## Before / after", "",
           "- ERP contact IDs accounted for: **%d**" % before,
           "- canonical vendors after apply (approx., after DB-level norm collisions): **%d**" % after,
           "",
           "## Round-trip accounting (asserted in code)", "",
           "| bucket | IDs |", "|---|---|"]
    rep += ["| %s | %d |" % (label, len(ids)) for label, ids in buckets]
    rep += ["| **total = input universe** | **%d** |" % len(universe), ""]
    if conflicts:
        rep += ["## Conflicts (resolved by EXCLUDE > MERGE > KEEP, all reported)", ""] \
            + ["- %s" % c for c in conflicts] + [""]
    if notes:
        rep += ["## Notes & flags", ""] + ["- %s" % n for n in notes] + [""]
    rep += ["## Outputs", "",
            "- `%s` — paste in the Supabase SQL editor (safe to run twice)" % sql_path.name,
            "- `retirement-checklist.xlsx` — for finance: losers to retire in SpectroNova",
            "- undecided rows pass through as standalone canonical vendors; finish the review "
            "and re-run apply — the guards make the second run additive-only.", ""]
    report_path = OUT / ("dedup-apply-%s-report.md" % master)
    report_path.write_text("\n".join(rep), encoding="utf-8")

    print("universe    : %d IDs (workbook + export union)" % len(universe))
    print("buckets     : primary %d / mapped %d / excluded %d / undecided %d"
          % (len(b_primary), len(b_mapped), len(b_excl), len(b_und)))
    print("round-trip  : OK (no residual, no overlap)")
    print("SQL         -> %s  (%d vendor inserts, %d mapping inserts)"
          % (sql_path, n_vendor_rows, n_map_rows))
    print("checklist   -> %s" % check_path)
    print("report      -> %s" % report_path)
    if conflicts:
        print("conflicts   : %d (see report)" % len(conflicts))
    return 0


# ── items master (0020: items + item_spectronova_ids) ───────────────────

ITEM_EXCLUDE = DEC_EXCLUDE | {"NOT MATERIAL", "NON MATERIAL", "NON-MATERIAL",
                              "NOT A MATERIAL", "NOT AN ITEM", "REMOVE", "DROP", "DELETE"}
ITEM_CONFIRM = {"OK", "YES", "CONFIRM", "CONFIRMED", "KEEP", "APPROVE", "APPROVED"}


def canon_uom(value, cfg):
    """Map a UOM spelling to its canonical key via uom_equivalence."""
    s = str(value or "").strip()
    if not s:
        return ""
    low = norm_text(s)
    for canon, alts in cfg.get("uom_equivalence", {}).items():
        if low == norm_text(canon) or low in {norm_text(a) for a in alts}:
            return canon
    return s


def apply_items(cfg, args, review, review_mtime, notes):
    """Apply the ITEMS sheet of the review workbook to the 0020 schema.

    The ITEMS sheet is a class/UOM review (ItemID, ItemCode, Description,
    Group, Suggested class, Current UOM, Suggested UOM, DECISION/corrected
    UOM) — not a duplicate-cluster review, so there are no merges here;
    item merge clusters arrive with `propose --master items` once the item
    export lands. DECISION semantics: an exclude word -> quarantine (no
    canonical row); a confirm word -> canonical with the suggested UOM; any
    other text -> treated as the CORRECTED UOM; blank -> undecided, passes
    through as canonical with the suggested UOM and is listed for review."""
    ts = args.run_ts
    wb = openpyxl.load_workbook(review, read_only=True, data_only=True)
    it_rows = sheet_rows(wb, "ITEMS") or []
    wb.close()
    if len(it_rows) < 2:
        print("ITEMS sheet in %s is empty — nothing to apply." % review.name)
        return 0
    hdr = [str(c).strip() if c is not None else "" for c in it_rows[0]]

    def col(*cands):
        for c in cands:
            hit = [i for i, h in enumerate(hdr) if h.upper().startswith(c.upper())]
            if hit:
                return hit[0]
        return None

    ic = {"id": col("ItemID", "Item ID"), "code": col("ItemCode", "Item Code"),
          "name": col("Description", "Item Name"), "group": col("Group", "Category"),
          "klass": col("Suggested class"), "cur": col("Current UOM"),
          "sugg": col("Suggested UOM"), "dec": col("DECISION")}
    missing = [k for k, v in ic.items() if v is None]
    if missing:
        notes.append("ITEMS sheet columns not found (handled): %s — header was %s"
                     % (", ".join(missing), hdr))

    def cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    rows, seen = [], {}
    for i, r in enumerate(it_rows[1:], 2):
        v = cell(r, ic["id"])
        iid = canon_id(v) if v is not None else ""
        code = canon_id(cell(r, ic["code"]) or "")
        name = str(cell(r, ic["name"]) or "").strip()
        if not iid and not code:
            continue
        key = iid or code
        if key in seen:
            notes.append("ITEMS row %d: ItemID %s listed again (row %d kept)" % (i, key, seen[key]))
            continue
        seen[key] = i
        dec_raw = str(cell(r, ic["dec"]) or "").strip()
        dec = dec_norm(dec_raw)
        sugg = canon_uom(cell(r, ic["sugg"]), cfg) or canon_uom(cell(r, ic["cur"]), cfg)
        if dec in ITEM_EXCLUDE:
            status, unit = "excluded", ""
        elif dec == "":
            status, unit = "undecided", sugg
        elif dec in ITEM_CONFIRM:
            status, unit = "decided", sugg
        else:                       # the column doubles as "corrected UOM"
            status, unit = "decided", canon_uom(dec_raw, cfg)
        rows.append({"id": iid, "code": code, "name": name, "row": i,
                     "group": str(cell(r, ic["group"]) or "").strip(),
                     "unit": unit, "status": status, "dec": dec_raw})

    # round-trip accounting: every item lands in exactly one bucket
    b_primary = {r["id"] or r["code"] for r in rows if r["status"] == "decided"}
    b_und = {r["id"] or r["code"] for r in rows if r["status"] == "undecided"}
    b_excl = {r["id"] or r["code"] for r in rows if r["status"] == "excluded"}
    universe = {r["id"] or r["code"] for r in rows}
    assert not (b_primary & b_und) and not (b_primary & b_excl) and not (b_und & b_excl)
    residual = universe - b_primary - b_und - b_excl
    if residual:
        raise AssertionError("round-trip residual (items apply): %s" % sorted(residual)[:10])

    # normalized-name collisions become ONE items row (unique index on vendor_norm)
    by_norm = defaultdict(list)
    for r in rows:
        if r["status"] != "excluded" and norm_text(r["name"]):
            by_norm[norm_text(r["name"])].append(r)
    for k, grp in sorted(by_norm.items()):
        if len(grp) > 1:
            notes.append("normalized-name collision '%s': ItemIDs %s share ONE items row "
                         "in the DB — their codes map to it"
                         % (k, ", ".join(g["id"] or g["code"] for g in grp)))

    # ── SQL (0020: items has NO notes column — provenance rides in comments)
    OUT.mkdir(exist_ok=True)
    quarantine = cfg["quarantine_name"]
    prov = "dedup apply %s (mtime %s) — run %s" % (review.name, review_mtime, ts)
    sql = ["-- %s — generated by tools/dedup/dedup.py --master items" % prov,
           "-- Idempotent: every insert is guarded (where not exists); re-runs are no-ops.",
           "-- Targets 0020: items (unique on vendor_norm(name)) + item_spectronova_ids.",
           "-- Merge-and-map, never delete: no update/delete statements in this file.",
           "begin;", "",
           "-- quarantine bucket: excluded ERP items resolve here so historical",
           "-- imports referencing them never crash (reported, not posted)",
           "insert into items (name, category, active)",
           "  select %s, 'quarantine', false" % sql_quote(quarantine),
           "  where not exists (select 1 from items where vendor_norm(name) = vendor_norm(%s));"
           % sql_quote(quarantine), ""]

    def map_sql(item_name, code, flagged):
        return ("insert into item_spectronova_ids (item_id, item_code, flagged)\n"
                "  select (select id from items where vendor_norm(name) = vendor_norm(%s) limit 1),\n"
                "         %s, %s\n"
                "  where not exists (select 1 from item_spectronova_ids where item_code = %s);"
                % (sql_quote(item_name), sql_quote(code),
                   "true" if flagged else "false", sql_quote(code)))

    n_item_rows = n_map_rows = 0
    skipped_unnamed = []
    for r in sorted(rows, key=lambda r: id_sort_key(r["id"] or r["code"])):
        if r["status"] == "excluded":
            continue
        if not norm_text(r["name"]):
            skipped_unnamed.append(r["id"] or r["code"])
            notes.append("ITEMS row %d: empty description — NO SQL emitted (listed undecided)" % r["row"])
            continue
        tag = "undecided — passes through, review later" if r["status"] == "undecided" else "decided"
        sql.append("-- %s [%s] %s — ITEMS!row %d" % (r["id"] or r["code"], tag, r["name"], r["row"]))
        sql.append("insert into items (name, unit, category)")
        sql.append("  select %s, %s, %s" % (sql_quote(r["name"]), sql_quote(r["unit"]),
                                            sql_quote(r["group"])))
        sql.append("  where not exists (select 1 from items where vendor_norm(name) = vendor_norm(%s));"
                   % sql_quote(r["name"]))
        n_item_rows += 1
        codes = [c for c in dict.fromkeys([r["code"], r["id"]]) if c]  # code first, id if distinct
        for c in codes:
            sql.append(map_sql(r["name"], c, flagged=False))
            n_map_rows += 1
        sql.append("")
    excl_rows = [r for r in rows if r["status"] == "excluded"]
    if excl_rows:
        sql.append("-- excluded items -> quarantine (flagged; never a canonical row)")
        for r in excl_rows:
            sql.append("-- %s %s — DECISION %r (ITEMS!row %d)" % (r["id"] or r["code"],
                                                                  r["name"], r["dec"], r["row"]))
            for c in [c for c in dict.fromkeys([r["code"], r["id"]]) if c]:
                sql.append(map_sql(quarantine, c, flagged=True))
                n_map_rows += 1
        sql.append("")
    sql += ["commit;", ""]
    sql_path = OUT / cfg["apply_sql"]
    sql_path.write_text("\n".join(sql), encoding="utf-8")

    # ── checklist (no MERGED sheet: the ITEMS sheet carries no merge decisions)
    check_path = OUT / "retirement-checklist-items.xlsx"
    cwb = openpyxl.Workbook()
    ws = cwb.active
    ws.title = "SUMMARY"
    for line in (("COPRI item retirement checklist",),
                 ("generated %s from %s" % (ts, review.name),),
                 (),
                 ("EXCLUDED sheet", "non-material / excluded item codes -> quarantine bucket", len(b_excl)),
                 ("UNDECIDED sheet", "blank decisions — passed through with the suggested UOM", len(b_und)),
                 (),
                 ("Merge clusters for items arrive with `propose --master items` once the",),
                 ("SpectroNova item export lands — this sheet reviewed class/UOM only.",)):
        ws.append(line)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws = cwb.create_sheet("EXCLUDED")
    ws.append(("ItemID", "ItemCode", "Description", "Group", "DECISION", "Resolves to"))
    for r in excl_rows:
        ws.append((r["id"], r["code"], r["name"], r["group"], r["dec"], quarantine))
    style_sheet(ws, (10, 14, 45, 22, 18, 40))
    ws = cwb.create_sheet("UNDECIDED")
    ws.append(("ItemID", "ItemCode", "Description", "Group", "Unit applied", "Where"))
    for r in rows:
        if r["status"] == "undecided":
            ws.append((r["id"], r["code"], r["name"], r["group"], r["unit"],
                       "ITEMS!row %d" % r["row"]))
    style_sheet(ws, (10, 14, 45, 22, 12, 14))
    cwb.save(check_path)

    # ── summary report ────────────────────────────────────────────────
    rep = ["# Dedup apply — %s — %s" % (cfg["label"], ts), "",
           "Review workbook: `%s` (mtime %s), ITEMS sheet." % (review.name, review_mtime),
           "",
           "## Counts per action", "",
           "| action | count |", "|---|---|",
           "| canonical items emitted (SQL insert-if-missing) | %d |" % n_item_rows,
           "| mapping rows emitted (ItemCode + distinct ItemID) | %d |" % n_map_rows,
           "| decided (class/UOM confirmed or corrected) | %d |" % len(b_primary),
           "| excluded -> quarantine | %d |" % len(b_excl),
           "| undecided (canonical with suggested UOM, review later) | %d |" % len(b_und),
           "",
           "## Round-trip accounting (asserted in code)", "",
           "| bucket | items |", "|---|---|",
           "| canonical (decided) | %d |" % len(b_primary),
           "| excluded -> quarantine | %d |" % len(b_excl),
           "| undecided (standalone canonical) | %d |" % len(b_und),
           "| **total = ITEMS sheet universe** | **%d** |" % len(universe), ""]
    if notes:
        rep += ["## Notes & flags", ""] + ["- %s" % n for n in notes] + [""]
    rep += ["## Outputs", "",
            "- `%s` — paste in the Supabase SQL editor AFTER migration 0020 (safe to run twice)" % sql_path.name,
            "- `retirement-checklist-items.xlsx` — excluded + undecided items",
            "- item MERGE clusters come from `propose --master items` when the export lands.", ""]
    report_path = OUT / "dedup-apply-items-report.md"
    report_path.write_text("\n".join(rep), encoding="utf-8")

    print("universe    : %d items (ITEMS sheet)" % len(universe))
    print("buckets     : decided %d / excluded %d / undecided %d"
          % (len(b_primary), len(b_excl), len(b_und)))
    print("round-trip  : OK (no residual, no overlap)")
    print("SQL         -> %s  (%d item inserts, %d mapping inserts)"
          % (sql_path, n_item_rows, n_map_rows))
    print("checklist   -> %s" % check_path)
    print("report      -> %s" % report_path)
    return 0


def write_checklist(path, ts, review_name, final, names, sources, provenance,
                    excluded, undecided, excl_undecided, quarantine):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUMMARY"
    n_mapped = sum(len(g["members"]) - 1 for g in final)
    for line in (("COPRI vendor retirement checklist",),
                 ("generated %s from %s" % (ts, review_name),),
                 (),
                 ("MERGED sheet", "losers to mark inactive in SpectroNova (or rename ZZ — DO NOT USE — see <survivor>)", n_mapped),
                 ("EXCLUDED sheet", "test / non-vendor / demo records -> quarantine bucket", len(excluded)),
                 ("UNDECIDED sheet", "blank decisions — passed through, finish the review", len(undecided) + len(set(excl_undecided) - set(excluded))),
                 (),
                 ("Nothing was deleted. Every old contact ID stays resolvable via vendor_spectronova_ids.",)):
        ws.append(line)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    ws = wb.create_sheet("MERGED")
    ws.append(("Contact ID", "Company name", "Source", "Survivor contact ID",
               "Survivor name", "ERP action"))
    for g in sorted(final, key=lambda g: id_sort_key(g["surv"])):
        for cid in g["members"]:
            if cid == g["surv"]:
                continue
            ws.append((cid, names.get(cid, ""), sources.get(cid, ""), g["surv"], g["name"],
                       "Mark inactive / rename: ZZ — DO NOT USE — see %s" % g["name"]))
    style_sheet(ws, (14, 50, 20, 16, 50, 60))

    ws = wb.create_sheet("EXCLUDED")
    ws.append(("Contact ID", "Company name", "Source", "Reason", "Resolves to", "ERP action"))
    for cid in sorted(excluded, key=id_sort_key):
        ws.append((cid, names.get(cid, ""), sources.get(cid, ""), excluded[cid][0],
                   quarantine, "Mark inactive / rename: ZZ — DO NOT USE — excluded"))
    style_sheet(ws, (14, 50, 20, 30, 40, 50))

    ws = wb.create_sheet("UNDECIDED")
    ws.append(("Contact ID", "Company name", "Where", "Why undecided"))
    und_all = dict(undecided)
    for cid, row in excl_undecided.items():
        und_all.setdefault(cid, "EXCLUSIONS row %d: blank DECISION" % row)
    for cid in sorted(und_all, key=id_sort_key):
        ws.append((cid, names.get(cid, ""), provenance.get(cid, ""), und_all[cid]))
    style_sheet(ws, (14, 50, 26, 40))
    wb.save(path)


# ══════════════════════════════════════════════════════════ main ══

def main():
    ap = argparse.ArgumentParser(
        description="Master-data dedup: propose clusters for human review, "
                    "apply the decided workbook as idempotent SQL")
    ap.add_argument("stage", choices=["propose", "apply"])
    ap.add_argument("--master", default="vendors", help="vendors | items (per masters.json)")
    ap.add_argument("--export", help="master export xlsx (default: newest match in "
                                     "Accounting raw data/spectronova)")
    ap.add_argument("--review", help="review workbook for apply (default: repo-root "
                                     "COPRI_dedup_review.xlsx, else output/dedup-review-*.xlsx)")
    ap.add_argument("--run-ts", default=datetime.now().isoformat(timespec="seconds"),
                    help="provenance timestamp override (ISO; default: now)")
    args = ap.parse_args()

    masters = json.loads(CFG_PATH.read_text(encoding="utf-8"))
    if args.master not in masters or args.master.startswith("_"):
        sys.exit("unknown master %r — masters.json defines: %s"
                 % (args.master, ", ".join(k for k in masters if not k.startswith("_"))))
    cfg = masters[args.master]

    if args.stage == "propose":
        return propose(args.master, cfg, args)
    return apply_stage(args.master, cfg, args)


if __name__ == "__main__":
    sys.exit(main())
